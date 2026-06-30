import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def generate_library_excel(items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    headers = [
        "Code", "Titel", "Autor", "Typ", "ISBN/Code", 
        "Filter 1", "Filter 2", "Filter 3", 
        "Status", "Ausgeliehen von", "Rückgabe", "Kosten"
    ]
    
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for item in items:
        status = "Verfuegbar" if str(item.get("Verfuegbar", "True")).lower() == "true" else "Ausgeliehen"
        row = [
            item.get("Code_4", ""),
            item.get("Name", ""),
            item.get("Author", ""),
            item.get("ItemType", ""),
            item.get("ISBN", ""),
            item.get("Filter", ""),
            item.get("Filter2", ""),
            item.get("Filter3", ""),
            status,
            item.get("User", ""),
            item.get("ReturnDate", ""),
            item.get("Anschaffungskosten", "")
        ]
        
        # FIX: Clean the row data INSIDE the loop for every single item
        clean_row = []
        for value in row:
            if isinstance(value, list):
                # Converts ['A', 'B'] to "A, B" and an empty list [] to an empty string ""
                clean_row.append(", ".join(map(str, value)) if value else "")
            elif isinstance(value, dict):
                # Just in case there is an embedded MongoDB sub-document
                clean_row.append(str(value))
            else:
                clean_row.append(value)
        
        # Append the safe, flattened row to the worksheet
        ws.append(clean_row)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer