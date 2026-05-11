'''
   Copyright 2025-2026 AIIrondev

   Licensed under the Inventarsystem EULA (Endbenutzer-Lizenzvertrag).
   See Legal/LICENSE for the full license text.
   Unauthorized commercial use, SaaS hosting, or removal of branding is prohibited.
   For commercial licensing inquiries: https://github.com/AIIrondev
'''
"""
Inventarsystem - Flask Web Application

This application provides an inventory management system with user authentication,
item tracking, QR code generation, and borrowing/returning functionality.

The system uses MongoDB for data storage and provides separate interfaces for
regular users and administrators.

Features:
- User authentication (login/logout)
- Item management (add, delete, view)
- Borrowing and returning items
- QR code generation for items
- Administrative functions
- History logging of item usage
- Booking and reservation of items
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, get_flashed_messages, jsonify, Response, make_response, send_file
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import user as us
import items as it
import ausleihung as au
import audit_log as al
import push_notifications as pn
import pdf_export 
import datetime
from apscheduler.schedulers.background import BackgroundScheduler
from bson.objectid import ObjectId
from urllib.parse import urlparse, urlunparse
import requests
import csv
import ipaddress
import os
import json
import datetime
import time
import traceback
import re
import socket
import io
import html
import logging
from logging.handlers import RotatingFileHandler
import secrets
import importlib
try:
    redis = importlib.import_module('redis')
except Exception:
    redis = None
# QR Code functionality deactivated
# import qrcode
# from qrcode.constants import ERROR_CORRECT_L
import threading
import sys
import shutil
import uuid
from PIL import Image, ImageOps
import mimetypes
import subprocess
from data_protection import (
    decrypt_document_fields,
    encrypt_document_fields,
    encrypt_soft_deleted_media_pack,
)

# Set base directory and centralized settings
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
import settings as cfg
from settings import MongoClient
from tenant import get_tenant_context


app = Flask(__name__, static_folder='static')  # Correctly set static folder
app.logger.setLevel(logging.DEBUG)
if not os.path.exists(cfg.LOGS_FOLDER):
    os.makedirs(cfg.LOGS_FOLDER, exist_ok=True)
log_file_path = os.path.join(cfg.LOGS_FOLDER, 'application.log')
file_handler = RotatingFileHandler(log_file_path, maxBytes=10 * 1024 * 1024, backupCount=5, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(name)s %(message)s'))
app.logger.handlers = []
app.logger.addHandler(file_handler)
app.secret_key = cfg.SECRET_KEY
app.debug = cfg.DEBUG
app.config['UPLOAD_FOLDER'] = cfg.UPLOAD_FOLDER
app.config['THUMBNAIL_FOLDER'] = cfg.THUMBNAIL_FOLDER
app.config['PREVIEW_FOLDER'] = cfg.PREVIEW_FOLDER
app.config['ALLOWED_EXTENSIONS'] = set(cfg.ALLOWED_EXTENSIONS)
app.config['MAX_CONTENT_LENGTH'] = max(cfg.MAX_UPLOAD_MB, cfg.IMAGE_MAX_UPLOAD_MB, cfg.VIDEO_MAX_UPLOAD_MB) * 1024 * 1024
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = cfg.SSL_ENABLED if os.getenv('INVENTAR_SESSION_COOKIE_SECURE') is None else os.getenv('INVENTAR_SESSION_COOKIE_SECURE', '').strip().lower() in ('1', 'true', 'yes', 'on')
app.config['PREFERRED_URL_SCHEME'] = 'https' if app.config['SESSION_COOKIE_SECURE'] else 'http'
# app.config['QR_CODE_FOLDER'] = cfg.QR_CODE_FOLDER  # QR Code storage deactivated
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

"""--------------------------------------------------------------Path Init-------------------------------------------------------"""

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])
# QR Code directory creation deactivated
# if not os.path.exists(app.config['QR_CODE_FOLDER']):
#     os.makedirs(app.config['QR_CODE_FOLDER'])

BACKUP_FOLDER = cfg.BACKUP_FOLDER
if not os.path.exists(BACKUP_FOLDER):
    try:
        os.makedirs(BACKUP_FOLDER, exist_ok=True)
    except PermissionError:
        # Fallback: use a backup directory inside the application directory (writable)
        fallback_backup = os.path.join(BASE_DIR, 'backups')
        try:
            os.makedirs(fallback_backup, exist_ok=True)
            BACKUP_FOLDER = fallback_backup
            print(f"Warnung: Konnte BACKUP_FOLDER nicht erstellen. Fallback genutzt: {BACKUP_FOLDER}")
        except Exception as e:
            print(f"Fehler: Backup-Verzeichnis konnte nicht erstellt werden: {e}")


def print(*args, **kwargs):
    if not args:
        return None

    message = " ".join(str(arg) for arg in args)
    stripped = message.lstrip()
    if stripped.startswith(('Error', 'Fehler', 'Warning', 'Warnung', '[WARN]', '[KONFLIKT]', 'Failed', 'Fehl')):
        app.logger.warning(message)
    elif stripped.startswith(('Exception', 'Traceback')):
        app.logger.error(message)
    else:
        app.logger.info(message)


STUDENT_CARD_ENCRYPTED_FIELDS = ('SchülerName', 'Klasse', 'Notizen')


def _decrypt_student_card_doc(card_doc):
    if not card_doc:
        return card_doc
    return decrypt_document_fields(card_doc, STUDENT_CARD_ENCRYPTED_FIELDS)

# Thumbnail sizes
THUMBNAIL_SIZE = cfg.THUMBNAIL_SIZE
PREVIEW_SIZE = cfg.PREVIEW_SIZE

__version__ = cfg.APP_VERSION
Host = cfg.HOST
Port = cfg.PORT

MONGODB_HOST = cfg.MONGODB_HOST
MONGODB_PORT = cfg.MONGODB_PORT
MONGODB_DB = cfg.MONGODB_DB
SCHEDULER_INTERVAL = cfg.SCHEDULER_INTERVAL_MIN
SSL_CERT = cfg.SSL_CERT
SSL_KEY = cfg.SSL_KEY

LIBRARY_ITEM_TYPES = ['book', 'cd', 'dvd', 'media']
INVOICE_CURRENCY = 'EUR'

NOTIFICATION_STATUS_CACHE_TTL = max(3, int(os.getenv('INVENTAR_NOTIFICATION_STATUS_CACHE_TTL', '8')))
_NOTIFICATION_CACHE_PREFIX = 'inventar:notif'
_NOTIFICATION_REDIS_CLIENT = None
_NOTIFICATION_REDIS_FAILED = False
_NOTIFICATION_LOCAL_CACHE = {}
_NOTIFICATION_LOCAL_VERSIONS = {'admin': 0}
_NOTIFICATION_CACHE_LOCK = threading.Lock()


SCHOOL_PERIODS = cfg.SCHOOL_PERIODS

PERMISSION_ACTION_OPTIONS = [
    ('can_borrow', 'Ausleihe erlauben'),
    ('can_insert', 'Einfügen/Hochladen erlauben'),
    ('can_edit', 'Bearbeiten erlauben'),
    ('can_delete', 'Löschen erlauben'),
    ('can_manage_users', 'Benutzerverwaltung erlauben'),
    ('can_manage_settings', 'Systemverwaltung erlauben'),
    ('can_view_logs', 'Logs/Audit einsehen erlauben'),
]

PERMISSION_PAGE_OPTIONS = [
    ('home', 'Artikel (Inventar)'),
    ('tutorial_page', 'Tutorial'),
    ('my_borrowed_items', 'Meine Ausleihen'),
    ('notifications_view', 'Benachrichtigungen'),
    ('impressum', 'Impressum'),
    ('license', 'Lizenz'),
    ('library_view', 'Bibliothek (Medien)'),
    ('terminplan', 'Terminplan'),
    ('home_admin', 'Admin Startseite'),
    ('upload_admin', 'Admin Upload Inventar'),
    ('library_admin', 'Admin Upload Bibliothek'),
    ('admin_borrowings', 'Admin Ausleihen'),
    ('library_loans_admin', 'Admin Bibliotheks-Ausleihen'),
    ('admin_damaged_items', 'Admin Defekte Items'),
    ('admin_audit_dashboard', 'Admin Audit Dashboard'),
    ('logs', 'System-Logs'),
    ('user_del', 'Benutzerverwaltung'),
    ('register', 'Benutzer anlegen'),
    ('manage_filters', 'Filter verwalten'),
    ('manage_locations', 'Orte verwalten'),
]

PERMISSION_EXEMPT_ENDPOINTS = {
    'static',
    'login',
    'logout',
    'impressum',
    'license',
    'uploaded_file',
    'thumbnail_file',
    'preview_file',
}

PERMISSION_ACTION_ENDPOINTS = {
    'upload_item': 'can_insert',
    'upload_inventory_excel': 'can_insert',
    'upload_library_excel': 'can_insert',
    'upload_student_cards_excel': 'can_insert',
    'edit_item': 'can_edit',
    'api_library_item_update': 'can_edit',
    'admin_update_user_name': 'can_edit',
    'delete_item': 'can_delete',
    'delete_user': 'can_delete',
    'ausleihen': 'can_borrow',
    'zurueckgeben': 'can_borrow',
    'api_library_scan_action': 'can_borrow',
    'user_del': 'can_manage_users',
    'register': 'can_manage_users',
    'admin_reset_user_password': 'can_manage_users',
    'admin_update_user_permissions': 'can_manage_users',
    'admin_anonymize_names': 'can_manage_users',
    'home_admin': 'can_manage_settings',
    'upload_admin': 'can_insert',
    'library_admin': 'can_insert',
    'admin_borrowings': 'can_manage_settings',
    'library_loans_admin': 'can_manage_settings',
    'admin_damaged_items': 'can_manage_settings',
    'manage_filters': 'can_manage_settings',
    'manage_locations': 'can_manage_settings',
    'admin_audit_dashboard': 'can_view_logs',
    'logs': 'can_view_logs',
}

# Apply the configuration for general use throughout the app
APP_VERSION = __version__
RELEASE_STATE_FILE = os.path.join(os.path.dirname(BASE_DIR), '.release-version')

""" -----------------------------------------------------------------------Before Request Handlers---------------------------------------------------------------------------- """

def _get_csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token

def _is_csrf_exempt_request():
    return request.method in {'GET', 'HEAD', 'OPTIONS', 'TRACE'}


@app.before_request
def _enforce_csrf_protection():
    if _is_csrf_exempt_request():
        _get_csrf_token()
        return None

    expected_token = _get_csrf_token()
    provided_token = (
        request.headers.get('X-CSRFToken')
        or request.headers.get('X-CSRF-Token')
        or request.form.get('csrf_token')
        or request.args.get('csrf_token')
    )

    if not provided_token or not secrets.compare_digest(provided_token, expected_token):
        return _csrf_error_response()

    return None

@app.before_request
def _enforce_user_permissions():
    endpoint = request.endpoint
    if not endpoint:
        return None

    if endpoint == 'static' or endpoint.startswith('static'):
        return None

    if endpoint in PERMISSION_EXEMPT_ENDPOINTS:
        return None

    if 'username' not in session:
        return None

    permissions = _get_current_user_permissions()
    if not permissions:
        return None

    if not _page_access_allowed(permissions, endpoint):
        message = 'Diese Seite ist für Ihren Benutzer aktuell gesperrt.'
        if request.path.startswith('/api/') or request.is_json:
            return jsonify({'ok': False, 'message': message}), 403

        flash(message, 'error')
        fallback_endpoint = _permission_denied_fallback_endpoint(permissions, current_endpoint=endpoint)
        return redirect(url_for(fallback_endpoint))

    action_key = PERMISSION_ACTION_ENDPOINTS.get(endpoint)
    if action_key and not _action_access_allowed(permissions, action_key):
        message = 'Für diese Aktion fehlen Ihnen die erforderlichen Berechtigungen.'
        if request.path.startswith('/api/') or request.is_json:
            return jsonify({'ok': False, 'message': message}), 403

        flash(message, 'error')
        fallback_endpoint = _permission_denied_fallback_endpoint(permissions, current_endpoint=endpoint)
        return redirect(url_for(fallback_endpoint))

    return None

@app.before_request
def _enforce_active_session_user():
    endpoint = request.endpoint or ''
    if endpoint == 'static' or endpoint.startswith('static'):
        return None

    username = session.get('username')
    if not username:
        return None

    user = us.get_user(username)
    if user:
        return None

    session.clear()
    if request.path.startswith('/api/') or request.is_json:
        return jsonify({'ok': False, 'message': 'Sitzung ungültig. Bitte erneut anmelden.'}), 401

    flash('Ihre Sitzung ist nicht mehr gültig. Bitte erneut anmelden.', 'error')
    return redirect(url_for('login'))

@app.before_request
def _enforce_module_access():
    endpoint = request.endpoint or ''
    if endpoint == 'static' or endpoint.startswith('static') or request.path.startswith('/api/'):
        return None
    
    for name, matcher in cfg.MODULES._path_matchers.items():
        if matcher(request.path) and not cfg.MODULES.is_enabled(name):
            msg = {
                'library': "Bibliotheks-Modul ist deaktiviert.",
                'inventory': "Inventar-Modul ist deaktiviert.",
                'student_cards': "Schülerausweis-Modul ist deaktiviert."
            }.get(name, f"{name.capitalize()}-Modul ist deaktiviert.")
            return msg, 403

""" -----------------------------------------------------------After Request Handlers----------------------------------------------------------------------------- """

@app.after_request
def _set_security_headers(response):
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    if cfg.SSL_ENABLED:
        response.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    
    # Optimize caching for static resources (images, etc.)
    path = request.path
    
    # Aggressive caching for optimized images (480p) - they're immutable
    if '/image/optimized/' in path:
        response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'  # 30 days
    
    # Moderate caching for thumbnails
    elif '/thumbnails/' in path:
        response.headers['Cache-Control'] = 'public, max-age=604800'  # 7 days
    
    # Moderate caching for previews
    elif '/previews/' in path:
        response.headers['Cache-Control'] = 'public, max-age=604800'  # 7 days
    
    # Short cache for regular uploads (in case they're updated/deleted)
    elif '/uploads/' in path:
        response.headers['Cache-Control'] = 'public, max-age=3600'  # 1 hour
    
    # Ensure WebP images are served with correct content-type
    if path.endswith('.webp') or '.webp' in path:
        response.headers['Content-Type'] = 'image/webp'
    
    return response


def _csrf_error_response(message='CSRF token fehlt oder ist ungültig.'):
    if request.is_json or request.path.startswith('/api/') or request.path in {'/download_book_cover', '/proxy_image', '/log_mobile_issue'}:
        return jsonify({'error': message}), 400
    flash(message, 'error')
    return redirect(request.referrer or url_for('home'))

def _get_current_module(path):
    """Resolve the active UI module for navbar separation."""
    mod = cfg.MODULES.get_module_for_path(path)
    if mod:
        session['last_module'] = mod
        return mod
        
    last_module = session.get('last_module')
    if last_module and cfg.MODULES.is_enabled(last_module):
        return last_module
        
    if cfg.MODULES.is_enabled('inventory'):
        return 'inventory'
    return 'library' if cfg.MODULES.is_enabled('library') else 'inventory'

"""---------------------------------------------User Access Permissions----------------------------------------------------------------------------- """

def _get_current_user_permissions():
    username = session.get('username')
    if not username:
        return None
    try:
        return us.get_effective_permissions(username)
    except Exception:
        return us.build_default_permission_payload('standard_user')


def _page_access_allowed(permissions, endpoint):
    if not permissions or not endpoint:
        return True
    page_permissions = permissions.get('pages', {})
    return bool(page_permissions.get(endpoint, True))


def _action_access_allowed(permissions, action_key):
    if not permissions or not action_key:
        return True
    action_permissions = permissions.get('actions', {})
    return bool(action_permissions.get(action_key, True))


def _permission_denied_fallback_endpoint(permissions, current_endpoint=None):
    username = session.get('username')
    is_admin_user = bool(username and us.check_admin(username))
    admin_home_allowed = _page_access_allowed(permissions, 'home_admin') and _action_access_allowed(permissions, 'can_manage_settings')

    for candidate in ('my_borrowed_items', 'tutorial_page', 'notifications_view', 'impressum', 'home'):
        if current_endpoint and candidate == current_endpoint:
            continue
        if candidate == 'home' and is_admin_user and not admin_home_allowed:
            continue
        if _page_access_allowed(permissions, candidate):
            return candidate
    return 'logout'

"""-------------------------------------------------------------Audit Logging----------------------------------------------------------------------------- """

def _append_audit_event_standalone(event_type, payload):
    """Write audit event by opening a short-lived DB connection."""
    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        try:
            al.append_audit_event(
                db=db,
                event_type=event_type,
                actor=session.get('username', 'system'),
                payload=payload,
                request_ip=request.remote_addr,
                source='web',
            )
        except Exception as exc:
            app.logger.warning(f"Audit write failed for {event_type}: {exc}")
    except Exception as exc:
        app.logger.warning(f"Standalone audit write failed for {event_type}: {exc}")
    finally:
        if client:
            client.close()

AUDIT_INDEXES_READY = False

def _ensure_audit_indexes_once():
    """Ensure audit indexes exist once per process."""
    if AUDIT_INDEXES_READY:
        return

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        al.ensure_audit_indexes(db)
        AUDIT_INDEXES_READY = True
    except Exception as exc:
        app.logger.warning(f"Could not ensure audit indexes: {exc}")
    finally:
        if client:
            client.close()

"""-------------------------------------------------------------School Info & Logo Handling----------------------------------------------------------------------------- """

def _get_school_info_for_export():
    """
    Get school information for PDF exports from configuration or database.
    Returns default info if not configured.
    """
    try:
        if hasattr(cfg, 'get_school_info'):
            return cfg.get_school_info()

        school_info = {
            'name': 'Schulname',
            'address': 'Schuladresse',
            'postal_code': 'PLZ',
            'city': 'Stadt',
            'school_number': '000000',
            'it_admin': 'IT-Beauftragter/in',
            'logo_path': '',
        }
        return school_info
    except Exception:
        # Return defaults if anything fails
        return {
            'name': 'Schulname',
            'address': 'Schuladresse', 
            'postal_code': 'PLZ',
            'city': 'Stadt',
            'school_number': '000000',
            'it_admin': 'IT-Beauftragter/in',
            'logo_path': '',
        }


def _save_school_logo_upload(upload_file, tenant_id=None, tenant_db=None):
    """Save an uploaded school logo to the shared upload folder with a tenant-specific filename."""
    if not upload_file or not getattr(upload_file, 'filename', ''):
        return None

    is_allowed, error_message = allowed_file(upload_file.filename, upload_file, max_size_mb=cfg.IMAGE_MAX_UPLOAD_MB)
    if not is_allowed:
        raise ValueError(error_message)

    safe_tenant = re.sub(r'[^a-zA-Z0-9_\-]+', '_', str(tenant_id or tenant_db or 'default').strip())
    safe_tenant = safe_tenant.strip('_') or 'default'
    _, original_ext = os.path.splitext(secure_filename(upload_file.filename))
    extension = (original_ext or '').lower()
    if not extension:
        raise ValueError('Das Logo benötigt eine Dateiendung.')

    logo_filename = f'school-logo-{safe_tenant}{extension}'
    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    upload_file.save(logo_path)

    # Generate a thumbnail for consistent navbar sizing
    try:
        thumb_filename = f'school-logo-{safe_tenant}-thumb.png'
        thumb_path = os.path.join(app.config['UPLOAD_FOLDER'], thumb_filename)

        with Image.open(logo_path) as img:
            # Ensure RGBA for transparency and convert if needed
            if img.mode not in ('RGBA', 'RGB'):
                img = img.convert('RGBA')

            # Target max size: width up to 520px, height up to 114px (matches CSS constraints)
            max_thumb_size = (520, 114)
            img.thumbnail(max_thumb_size, Image.LANCZOS)

            # Save as PNG for predictable rendering in web UI
            img.save(thumb_path, format='PNG', optimize=True)
    except Exception as e:
        app.logger.warning(f"Thumbnail generation failed for {logo_path}: {e}")
        thumb_filename = None

    return (logo_filename, thumb_filename)

"""---------------------------------------------Invoice Generation----------------------------------------------------------------------------- """

def _parse_money_value(value):
    """Parse a user-facing money value into a float when possible."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    cleaned = re.sub(r'[^0-9,\.\-]', '', text)
    if ',' in cleaned and '.' in cleaned:
        cleaned = cleaned.replace('.', '').replace(',', '.')
    else:
        cleaned = cleaned.replace(',', '.')
    try:
        return float(cleaned)
    except ValueError:
        return None


def _format_money_value(value):
    """Format a money value in German notation for display."""
    parsed_value = _parse_money_value(value)
    if parsed_value is None:
        return '-'
    formatted = f"{parsed_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"{formatted} {INVOICE_CURRENCY}"


def _build_invoice_number(borrow_id, created_at):
    """Build a stable, human-readable invoice number."""
    short_id = str(borrow_id)[-6:].upper()
    return f"INV-{created_at.strftime('%Y%m%d-%H%M%S')}-{short_id}"


def _prepare_invoice_pdf_payload(invoice_data, borrow_doc=None, item_doc=None):
    """Normalize stored invoice data for robust PDF rendering."""
    borrow_doc = borrow_doc or {}
    item_doc = item_doc or {}

    created_at_raw = invoice_data.get('created_at')
    created_at_display = ''
    if isinstance(created_at_raw, datetime.datetime):
        created_at_display = created_at_raw.strftime('%d.%m.%Y %H:%M')
    elif created_at_raw:
        created_at_display = str(created_at_raw)

    amount_value = _parse_money_value(invoice_data.get('amount'))
    if amount_value is None:
        amount_value = _parse_money_value(invoice_data.get('amount_text'))
    amount_text = invoice_data.get('amount_text')
    if not amount_text:
        amount_text = _format_money_value(amount_value)

    item_id = invoice_data.get('item_id')
    if not item_id and item_doc.get('_id'):
        item_id = str(item_doc.get('_id'))

    return {
        'invoice_number': invoice_data.get('invoice_number') or _build_invoice_number(borrow_doc.get('_id', ''), datetime.datetime.now()),
        'created_at': created_at_raw,
        'created_at_display': created_at_display,
        'borrower': invoice_data.get('borrower') or borrow_doc.get('User', '-'),
        'item_name': invoice_data.get('item_name') or item_doc.get('Name', '-'),
        'item_code': invoice_data.get('item_code') or item_doc.get('Code_4', '-'),
        'item_id': str(item_id or '-'),
        'damage_reason': invoice_data.get('damage_reason') or 'Keine Schadensbeschreibung hinterlegt.',
        'amount': amount_value,
        'amount_text': amount_text or '-',
    }

"""---------------------------------------------Notification System----------------------------------------------------------------------------- """

def _create_notification(db, *, audience, notif_type, title, message, target_user=None, reference=None, unique_key=None, severity='info'):
    """Create a notification entry with optional deduplication via unique_key."""
    notifications_col = db['notifications']

    if unique_key:
        existing = notifications_col.find_one({'UniqueKey': unique_key}, {'_id': 1})
        if existing:
            return False

    now = datetime.datetime.now()
    payload = {
        'Audience': audience,
        'Type': notif_type,
        'Title': title,
        'Message': message,
        'TargetUser': target_user,
        'Reference': reference or {},
        'UniqueKey': unique_key,
        'Severity': severity,
        'ReadBy': [],
        'CreatedAt': now,
        'UpdatedAt': now,
    }
    notifications_col.insert_one(payload)

    if audience == 'user' and target_user:
        _bump_notification_version(f'user:{target_user}')
        # Send push notification to user
        try:
            pn.send_push_notification(
                target_user,
                title,
                message,
                url=reference.get('url', '/') if reference else '/',
                reference=reference,
                tag=f'notification-{notif_type}'
            )
        except Exception as e:
            app.logger.warning(f'Failed to send push notification to {target_user}: {e}')
    elif audience == 'admin':
        _bump_notification_version('admin')
        # Send push notification to all admins
        try:
            pn.send_push_to_all_admins(
                title,
                message,
                url=reference.get('url', '/') if reference else '/',
                reference=reference
            )
        except Exception as e:
            app.logger.warning(f'Failed to send push to admins: {e}')

    return True


def _get_notification_cache_client():
    """Return shared Redis client for distributed notification cache if available."""
    global _NOTIFICATION_REDIS_CLIENT, _NOTIFICATION_REDIS_FAILED
    if _NOTIFICATION_REDIS_CLIENT is not None:
        return _NOTIFICATION_REDIS_CLIENT
    if _NOTIFICATION_REDIS_FAILED or redis is None:
        return None

    try:
        redis_host = os.getenv('INVENTAR_REDIS_HOST', 'redis')
        redis_port = int(os.getenv('INVENTAR_REDIS_PORT', 6379))
        redis_db = int(os.getenv('INVENTAR_REDIS_CACHE_DB', 1))
        client = redis.Redis(
            host=redis_host,
            port=redis_port,
            db=redis_db,
            decode_responses=True,
            socket_keepalive=True,
            socket_timeout=1,
            socket_connect_timeout=1,
        )
        client.ping()
        _NOTIFICATION_REDIS_CLIENT = client
        app.logger.info(f'Notification cache backend enabled: {redis_host}:{redis_port}/db{redis_db}')
        return _NOTIFICATION_REDIS_CLIENT
    except Exception as exc:
        _NOTIFICATION_REDIS_FAILED = True
        app.logger.warning(f'Notification cache backend unavailable, using local cache fallback: {exc}')
        return None


def _notification_scope_key(scope):
    return f'{_NOTIFICATION_CACHE_PREFIX}:ver:{scope}'


def _notification_status_key(username, is_admin, version_tag):
    role = 'admin' if is_admin else 'user'
    return f'{_NOTIFICATION_CACHE_PREFIX}:status:{role}:{username}:{version_tag}'


def _get_notification_version_tag(username, is_admin=False):
    user_scope = f'user:{username}'
    cache_client = _get_notification_cache_client()

    if cache_client:
        user_version = cache_client.get(_notification_scope_key(user_scope)) or '0'
        if is_admin:
            admin_version = cache_client.get(_notification_scope_key('admin')) or '0'
            return f'u{user_version}-a{admin_version}'
        return f'u{user_version}'

    with _NOTIFICATION_CACHE_LOCK:
        user_version = _NOTIFICATION_LOCAL_VERSIONS.get(user_scope, 0)
        if is_admin:
            admin_version = _NOTIFICATION_LOCAL_VERSIONS.get('admin', 0)
            return f'u{user_version}-a{admin_version}'
        return f'u{user_version}'


def _bump_notification_version(scope):
    cache_client = _get_notification_cache_client()
    if cache_client:
        try:
            cache_client.incr(_notification_scope_key(scope))
            return
        except Exception as exc:
            app.logger.warning(f'Could not bump notification cache version for {scope}: {exc}')

    with _NOTIFICATION_CACHE_LOCK:
        current = int(_NOTIFICATION_LOCAL_VERSIONS.get(scope, 0))
        _NOTIFICATION_LOCAL_VERSIONS[scope] = current + 1
        # Prevent stale local payload reuse after version bump.
        _NOTIFICATION_LOCAL_CACHE.clear()


def _get_cached_unread_status(username, is_admin=False):
    version_tag = _get_notification_version_tag(username, is_admin=is_admin)
    key = _notification_status_key(username, is_admin, version_tag)
    cache_client = _get_notification_cache_client()

    if cache_client:
        try:
            cached = cache_client.get(key)
            if cached:
                return json.loads(cached), version_tag
        except Exception as exc:
            app.logger.warning(f'Could not read notification status cache for {username}: {exc}')
        return None, version_tag

    now = time.time()
    with _NOTIFICATION_CACHE_LOCK:
        cached = _NOTIFICATION_LOCAL_CACHE.get(key)
        if not cached:
            return None, version_tag
        if cached.get('expires_at', 0) <= now:
            _NOTIFICATION_LOCAL_CACHE.pop(key, None)
            return None, version_tag
        return cached.get('payload'), version_tag


def _set_cached_unread_status(username, is_admin, version_tag, payload):
    key = _notification_status_key(username, is_admin, version_tag)
    cache_client = _get_notification_cache_client()

    if cache_client:
        try:
            cache_client.setex(key, NOTIFICATION_STATUS_CACHE_TTL, json.dumps(payload, default=str))
            return
        except Exception as exc:
            app.logger.warning(f'Could not write notification status cache for {username}: {exc}')

    with _NOTIFICATION_CACHE_LOCK:
        _NOTIFICATION_LOCAL_CACHE[key] = {
            'expires_at': time.time() + NOTIFICATION_STATUS_CACHE_TTL,
            'payload': payload,
        }


def _build_unread_status_etag(version_tag, payload):
    unread_count = int(payload.get('unread_count', 0))
    latest = payload.get('latest_unread') or {}
    latest_created = latest.get('created_at') or ''
    latest_type = latest.get('type') or ''
    return f'W/"notif-{version_tag}-{unread_count}-{latest_type}-{latest_created}"'


def _build_cached_json_response(payload, etag_value):
    incoming_etag = (request.headers.get('If-None-Match') or '').strip()
    if incoming_etag and incoming_etag == etag_value:
        response = make_response('', 304)
    else:
        response = jsonify(payload)

    response.headers['Cache-Control'] = f'private, max-age={NOTIFICATION_STATUS_CACHE_TTL}, must-revalidate'
    response.headers['ETag'] = etag_value
    response.headers['Vary'] = 'Cookie'
    return response


def _get_notifications_for_user(db, username, is_admin=False, limit=150):
    """Fetch notifications visible to the current user, newest first."""
    query = {
        '$or': [
            {'Audience': 'user', 'TargetUser': username},
        ]
    }
    if is_admin:
        query['$or'].append({'Audience': 'admin'})

    cursor = db['notifications'].find(query).sort('CreatedAt', -1).limit(limit)
    return list(cursor)


def _get_unread_notification_count(db, username, is_admin=False):
    """Count unread notifications for navbar badge rendering."""
    query = {
        '$and': [
            {
                '$or': [
                    {'Audience': 'user', 'TargetUser': username},
                ]
            },
            {'ReadBy': {'$ne': username}},
        ]
    }
    if is_admin:
        query['$and'][0]['$or'].append({'Audience': 'admin'})

    return db['notifications'].count_documents(query)


def _build_reminder_message(item_name, start_dt=None, end_dt=None):
    """Build a concise reminder text for borrowed items."""
    start_text = start_dt.strftime('%d.%m.%Y %H:%M') if isinstance(start_dt, datetime.datetime) else '-'
    end_text = end_dt.strftime('%d.%m.%Y %H:%M') if isinstance(end_dt, datetime.datetime) else '-'
    return (
        f"Bitte denke an die Rueckgabe von '{item_name}'. "
        f"Ausleihe seit: {start_text}. "
        f"Geplantes Ende: {end_text}."
    )


def create_return_reminders():
    """Create one-time reminders for day-1 and planned-end events."""
    now = datetime.datetime.now()
    one_day_ago = now - datetime.timedelta(days=1)

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen_col = db['ausleihungen']
        items_col = db['items']

        candidates = list(ausleihungen_col.find(
            {
                'Status': {'$in': ['active', 'planned']},
                'User': {'$exists': True, '$ne': ''},
                '$or': [
                    {'Start': {'$lte': one_day_ago}},
                    {'End': {'$lte': now}},
                ]
            },
            {'User': 1, 'Item': 1, 'Status': 1, 'Start': 1, 'End': 1}
        ))

        item_ids = []
        for entry in candidates:
            item_id = entry.get('Item')
            if item_id:
                item_ids.append(item_id)

        item_docs = {}
        for raw_id in item_ids:
            try:
                doc = items_col.find_one({'_id': ObjectId(raw_id)}, {'Name': 1, 'Code_4': 1})
                if doc:
                    item_docs[raw_id] = doc
            except Exception:
                continue

        for borrow_doc in candidates:
            borrow_id = str(borrow_doc.get('_id'))
            username = str(borrow_doc.get('User', '')).strip()
            if not borrow_id or not username:
                continue

            item_id = str(borrow_doc.get('Item', '')).strip()
            item_doc = item_docs.get(item_id, {})
            item_name = item_doc.get('Name') or f'Item {item_id}'

            start_dt = borrow_doc.get('Start')
            end_dt = borrow_doc.get('End')

            if isinstance(start_dt, datetime.datetime) and start_dt <= one_day_ago:
                _create_notification(
                    db,
                    audience='user',
                    notif_type='return_day_1',
                    title='Erinnerung: Rueckgabe nach 1 Tag',
                    message=_build_reminder_message(item_name, start_dt=start_dt, end_dt=end_dt),
                    target_user=username,
                    reference={'borrow_id': borrow_id, 'item_id': item_id, 'event': 'day_1'},
                    unique_key=f'reminder:day1:{borrow_id}',
                    severity='warning',
                )

            if isinstance(end_dt, datetime.datetime) and end_dt <= now:
                _create_notification(
                    db,
                    audience='user',
                    notif_type='return_after_end',
                    title='Erinnerung: Geplante Ausleihe ist beendet',
                    message=_build_reminder_message(item_name, start_dt=start_dt, end_dt=end_dt),
                    target_user=username,
                    reference={'borrow_id': borrow_id, 'item_id': item_id, 'event': 'after_end'},
                    unique_key=f'reminder:end:{borrow_id}',
                    severity='warning',
                )
    except Exception as exc:
        app.logger.warning(f"Reminder creation failed: {exc}")
    finally:
        if client:
            client.close()

@app.context_processor
def inject_version():
    """Inject global template variables."""
    is_admin = False
    asset_version = APP_VERSION
    csrf_token = _get_csrf_token()
    unread_notification_count = 0
    current_permissions = us.build_default_permission_payload('standard_user')
    if 'username' in session:
        try:
            is_admin = us.check_admin(session['username'])
        except Exception:
            is_admin = False

        try:
            current_permissions = us.get_effective_permissions(session['username'])
        except Exception:
            current_permissions = us.build_default_permission_payload('standard_user')

        client = None
        try:
            client = MongoClient(MONGODB_HOST, MONGODB_PORT)
            db = client[MONGODB_DB]
            unread_notification_count = _get_unread_notification_count(db, session['username'], is_admin=is_admin)
        except Exception:
            unread_notification_count = 0
        finally:
            if client:
                client.close()

    current_module = _get_current_module(request.path)
    current_tenant_db = MONGODB_DB
    current_tenant_id = None
    try:
        ctx = get_tenant_context()
        if ctx:
            current_tenant_db = ctx.db_name or current_tenant_db
            current_tenant_id = ctx.tenant_id
    except Exception:
        current_tenant_db = MONGODB_DB

    return {
        'APP_VERSION': APP_VERSION,
        'ASSET_VERSION': asset_version,
        'csrf_token': csrf_token,
        'CURRENT_MODULE': current_module,
        'school_periods': SCHOOL_PERIODS,
        'inventory_module_enabled': cfg.MODULES.is_enabled('inventory'),
        'library_module_enabled': cfg.MODULES.is_enabled('library'),
        'student_cards_module_enabled': cfg.MODULES.is_enabled('student_cards'),
        'is_admin': is_admin,
        'unread_notification_count': unread_notification_count,
        'current_permissions': current_permissions,
        'current_tenant_db': current_tenant_db,
        'current_tenant_id': current_tenant_id,
        'school_info': _get_school_info_for_export(),
        'permission_action_options': PERMISSION_ACTION_OPTIONS,
        'permission_page_options': PERMISSION_PAGE_OPTIONS,
        'permission_presets': us.get_permission_preset_definitions(),
    }


"""-------------------------------------------------------------Scheduled Tasks----------------------------------------------------------------------------- """

def create_daily_backup():
    """
    Erstellt täglich ein Backup der Ausleihungsdatenbank
    """
    try:
        result = au.create_backup_database()
        if not result:
            app.logger.warning("Daily backup creation returned false")
    except Exception as e:
        app.logger.error(f"Daily backup creation failed: {e}")

def update_appointment_statuses():
    """
    Aktualisiert automatisch die Status aller Terminplaner-Einträge.
    Diese Funktion wird jede Minute ausgeführt und überprüft:
    - Geplante Termine, die aktiviert werden sollten
    - Aktive Termine, die beendet werden sollten
    """
    current_time = datetime.datetime.now()

    try:
        # Hole alle Termine mit Status 'planned' oder 'active'
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']

        # Finde alle Termine, die status updates benötigen
        appointments_to_check = list(ausleihungen.find({
            'Status': {'$in': ['planned', 'active']}
        }))

        updated_count = 0
        activated_count = 0
        completed_count = 0

        for appointment in appointments_to_check:
            old_status = appointment.get('Status')
            activation_user = str(appointment.get('User') or '').strip()
            activation_item_name = str(appointment.get('Item') or 'Termin')

            # Aktuellen Status bestimmen
            new_status = au.get_current_status(appointment, log_changes=True, user='scheduler')

            # Wenn sich der Status geändert hat, aktualisiere in der Datenbank
            if new_status != old_status:
                extra_fields = {}

                # --- Conflict resolver: planned → active transition ---
                # Check if the physical item is already borrowed by someone else
                if old_status == 'planned' and new_status == 'active':
                    items_col = db['items']
                    item_id_str = appointment.get('Item')
                    conflict_detected = False
                    conflict_note = ''
                    item_name = item_id_str or 'Termin'
                    if item_id_str:
                        try:
                            item_doc = items_col.find_one(
                                {'_id': ObjectId(item_id_str)},
                                {'Verfuegbar': 1, 'User': 1, 'Name': 1, 'Exemplare': 1}
                            )
                            if item_doc:
                                item_name = item_doc.get('Name', item_id_str)
                                activation_item_name = item_name
                                total_exemplare = int(item_doc.get('Exemplare', 1))
                                # Count how many active (non-planned) borrows currently hold this item
                                active_borrows = ausleihungen.count_documents({
                                    'Item': item_id_str,
                                    'Status': 'active',
                                    '_id': {'$ne': appointment['_id']}
                                })
                                if active_borrows >= total_exemplare or item_doc.get('Verfuegbar') is False:
                                    conflict_detected = True
                                    borrower = item_doc.get('User', 'unbekannter Benutzer')
                                    item_name = item_doc.get('Name', item_id_str)
                                    conflict_note = (
                                        f"Gegenstand '{item_name}' war beim Aktivieren von "
                                        f"'{appointment.get('User', '?')}' bereits ausgeliehen "
                                        f"von '{borrower}' (aktive Borrows: {active_borrows}/{total_exemplare})."
                                    )
                                    extra_fields['ConflictDetected'] = True
                                    extra_fields['ConflictNote'] = conflict_note
                                    extra_fields['ConflictAt'] = current_time
                                    conflict_log = (
                                        f"  [KONFLIKT] Termin {appointment['_id']}: "
                                        f"planned → active, aber {conflict_note}"
                                    )
                                    app.logger.warning(conflict_log)
                                else:
                                    # No conflict — clear any previously stored conflict flag
                                    extra_fields['ConflictDetected'] = False
                                    extra_fields['ConflictNote'] = ''
                        except Exception as conflict_err:
                            app.logger.warning(
                                f"Conflict check failed for appointment {appointment['_id']}: {conflict_err}"
                            )

                result = ausleihungen.update_one(
                    {'_id': appointment['_id']},
                    {'$set': {
                        'Status': new_status,
                        'LastUpdated': current_time,
                        **extra_fields
                    }}
                )

                if result.modified_count > 0:
                    updated_count += 1
                    if new_status == 'active':
                        activated_count += 1
                        # Make item unshareable if no conflict is detected
                        if old_status == 'planned' and appointment.get('Item') and not extra_fields.get('ConflictDetected', False):
                            try:
                                it.update_item_status(str(appointment.get('Item')), False, activation_user)
                            except Exception as e:
                                app.logger.warning(f"Could not update item status to False for {appointment['_id']}: {e}")

                    elif new_status == 'completed':
                        completed_count += 1
                        # Make item available again
                        if appointment.get('Item'):
                            try:
                                it.update_item_status(str(appointment.get('Item')), True)
                            except Exception as e:
                                app.logger.warning(f"Could not update item status to True for {appointment['_id']}: {e}")

                # Create activation notification even if another worker already updated the status.
                if old_status == 'planned' and new_status == 'active' and activation_user:
                    try:
                        _create_notification(
                            db,
                            audience='user',
                            notif_type='appointment_activated',
                            title='Reservierung ist jetzt aktiv',
                            message=(
                                f"Deine geplante Ausleihe für {activation_item_name} startet jetzt."
                            ),
                            target_user=activation_user,
                            reference={
                                'appointment_id': str(appointment.get('_id')),
                                'item_id': str(appointment.get('Item') or ''),
                                'event': 'activated',
                            },
                            unique_key=f"appointment:activated:{appointment.get('_id')}",
                            severity='info',
                        )
                    except Exception as notif_err:
                        app.logger.warning(
                            f"Failed to create activation notification for {appointment.get('_id')}: {notif_err}"
                        )

        client.close()

        if updated_count > 0:
            app.logger.warning(
                f"Appointment status update finished: {updated_count} changed ({activated_count} active, {completed_count} completed)"
            )

    except Exception as e:
        app.logger.error(f"Automatic appointment status update failed: {e}")

# Schedule jobs - only start scheduler if this is the main process or a single-worker deployment
# This prevents race conditions in multi-worker Gunicorn environments
scheduler = BackgroundScheduler()
_scheduler_initialized = False

def _initialize_scheduler():
    """Initialize the background scheduler in a safe way for multi-worker deployments."""
    global _scheduler_initialized
    if _scheduler_initialized or not cfg.SCHEDULER_ENABLED:
        return
    
    try:
        # For multi-worker Gunicorn, use a lock file to ensure only one instance starts the scheduler
        # Clean up any stale lock file from previous runs (older than 5 minutes)
        scheduler_lock_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.scheduler_lock')
        try:
            if os.path.exists(scheduler_lock_path):
                lock_age = time.time() - os.path.getmtime(scheduler_lock_path)
                if lock_age > 300:  # 5 minutes - indicates a stale lock from a previous container run
                    os.remove(scheduler_lock_path)
                    app.logger.info(f"Removed stale scheduler lock file (age: {lock_age:.0f}s)")
        except Exception as e:
            app.logger.warning(f"Could not clean up scheduler lock file: {e}")

        # Always try to remove lock file on startup (extra safety)
        try:
            if os.path.exists(scheduler_lock_path):
                os.remove(scheduler_lock_path)
                app.logger.info("Scheduler lock file removed on startup.")
        except Exception as e:
            app.logger.warning(f"Could not remove scheduler lock file on startup: {e}")

        try:
            # Try to create the lock file - only succeeds if it doesn't exist
            lock_fd = os.open(scheduler_lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o644)
            os.close(lock_fd)
            should_start = True
        except FileExistsError:
            should_start = False
            app.logger.warning("Scheduler lock exists - another process is already running the scheduler")

        if should_start:
            scheduler.add_job(func=create_daily_backup, trigger="interval", hours=cfg.BACKUP_INTERVAL_HOURS)
            scheduler.add_job(func=update_appointment_statuses, trigger="interval", minutes=cfg.SCHEDULER_INTERVAL_MIN)
            scheduler.add_job(func=create_return_reminders, trigger="interval", minutes=cfg.SCHEDULER_INTERVAL_MIN)
            scheduler.start()
            _scheduler_initialized = True
            app.logger.info(f"Scheduler started successfully (interval={cfg.SCHEDULER_INTERVAL_MIN} min)")
        else:
            app.logger.info("Scheduler skipped - another worker instance is running it")
    except Exception as e:
        app.logger.error(f"Failed to initialize scheduler: {e}")
        _scheduler_initialized = False

# Initialize scheduler on app startup
_initialize_scheduler()

# Register shutdown handler to stop scheduler when app is terminated
import atexit
def _shutdown_scheduler():
    if cfg.SCHEDULER_ENABLED and _scheduler_initialized:
        try:
            scheduler.shutdown()
            lock_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.scheduler_lock')
            try:
                os.remove(lock_path)
            except Exception:
                pass
        except Exception as e:
            app.logger.error(f"Error during scheduler shutdown: {e}")

atexit.register(_shutdown_scheduler)

"""-------------------------------------------------------------File Upload Validation----------------------------------------------------------------------------- """

def allowed_file(filename, file_content=None, max_size_mb=cfg.MAX_UPLOAD_MB):
    """
    Check if a file has an allowed extension and valid content.
    
    Args:
        filename (str): Name of the file to check
        file_content (FileStorage, optional): The actual file object to validate content
        max_size_mb (int, optional): Maximum allowed file size in MB
        
    Returns:
        tuple: (bool, str) - True if the file is valid, False otherwise
               along with an error message if not valid
    """
    # Check file extension
    if '.' not in filename:
        return False, f"Datei '{filename}' hat keine Dateiendung. Erlaubte Formate: {', '.join(app.config['ALLOWED_EXTENSIONS'])}"
    
    extension = filename.rsplit('.', 1)[1].lower()
    allowed_extensions_lower = {ext.lower() for ext in app.config['ALLOWED_EXTENSIONS']}
    if extension not in allowed_extensions_lower:
        app.logger.warning(f"File extension not allowed: {extension} for file {filename}. Allowed: {allowed_extensions_lower}")
        return False, f"Datei '{filename}' hat ein nicht unterstütztes Format ({extension}). Erlaubte Formate: {', '.join(app.config['ALLOWED_EXTENSIONS'])}"
    
    # Check file size if content is provided
    if file_content is not None:
        # Check file size
        file_content.seek(0, os.SEEK_END)
        file_size = file_content.tell() / (1024 * 1024)  # Size in MB
        file_content.seek(0)  # Reset file pointer to beginning
        
        if file_size > max_size_mb:
            return False, f"Datei '{filename}' ist zu groß ({file_size:.1f} MB). Maximale Größe: {max_size_mb} MB."
        
        # Verify file content matches extension
        try:
            if extension in ['jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp']:
                try:
                    # Special debug for PNG files
                    if extension == 'png':
                        app.logger.info(f"PNG DEBUG: Validating PNG file: {filename}, size: {file_size:.2f}MB")
                        # Save first few bytes for magic number checking
                        header_bytes = file_content.read(32)  # Reading more bytes for deeper analysis
                        file_content.seek(0)  # Reset pointer
                        
                        # Check PNG magic number (first 8 bytes)
                        png_signature = b'\x89PNG\r\n\x1a\n'
                        is_valid_signature = header_bytes.startswith(png_signature)
                        
                        # Create a hex dump of the header for debugging
                        hex_dump = ' '.join([f"{b:02x}" for b in header_bytes[:16]])
                        app.logger.info(f"PNG DEBUG: File {filename} - Header hex dump: {hex_dump}")
                        app.logger.info(f"PNG DEBUG: Valid PNG signature: {is_valid_signature}, first bytes: {header_bytes[:8]!r}")
                        
                        # More detailed analysis of PNG chunks
                        if is_valid_signature:
                            try:
                                # IHDR chunk should start at byte 8
                                if header_bytes[8:12] == b'IHDR':
                                    app.logger.info(f"PNG DEBUG: Found IHDR chunk at correct position")
                                else:
                                    app.logger.warning(f"PNG DEBUG: IHDR chunk not found at expected position. Found: {header_bytes[8:12]!r}")
                            except Exception as chunk_err:
                                app.logger.error(f"PNG DEBUG: Error analyzing PNG chunks: {str(chunk_err)}")
                        else:
                            app.logger.error(f"PNG DEBUG: Invalid PNG signature for {filename}. Expected: {png_signature!r}")
                    
                    with Image.open(file_content) as img:
                        # Verify it's a valid image by accessing its format and size
                        img_format = img.format
                        img_mode = img.mode
                        img_size = img.size
                        
                        if extension == 'png':
                            app.logger.info(f"PNG DEBUG: Successfully opened PNG - Format: {img_format}, Mode: {img_mode}, Size: {img_size[0]}x{img_size[1]}")
                            # Add more PNG-specific checks
                            app.logger.info(f"PNG DEBUG: Image info - Bands: {len(img.getbands())}, Bands: {img.getbands()}")
                            # Check if there's transparency
                            has_alpha = 'A' in img.getbands() or img.mode == 'P' and img.info.get('transparency') is not None
                            app.logger.info(f"PNG DEBUG: Has transparency: {has_alpha}")
                        
                        if not img_format:
                            if extension == 'png':
                                app.logger.error(f"PNG DEBUG: Invalid format - got None for {filename}")
                            return False, f"Datei '{filename}' scheint keine gültige Bilddatei zu sein."
                        
                        file_content.seek(0)  # Reset file pointer after reading
                except Exception as e:
                    error_msg = f"Error validating image content for {filename}: {str(e)}"
                    app.logger.error(error_msg)
                    
                    if extension == 'png':
                        app.logger.error(f"PNG DEBUG: Failed to process PNG file: {filename}")
                        app.logger.error(f"PNG DEBUG: Error details: {str(e)}")
                        app.logger.error(f"PNG DEBUG: Error type: {type(e).__name__}")
                        
                        # Get the full traceback as string and log it
                        import io
                        tb_output = io.StringIO()
                        traceback.print_exc(file=tb_output)
                        app.logger.error(f"PNG DEBUG: Full traceback for {filename}:\n{tb_output.getvalue()}")
                        
                        # Try to manually read the file data and check it
                        try:
                            file_content.seek(0)
                            file_bytes = file_content.read(1024)  # Read first KB
                            file_content.seek(0)  # Reset again
                            
                            hex_signature = ' '.join([f"{b:02x}" for b in file_bytes[:16]])
                            app.logger.error(f"PNG DEBUG: Raw file bytes (first 16): {hex_signature}")
                            
                            # Check for common corruption patterns
                            if not file_bytes.startswith(png_signature):
                                app.logger.error(f"PNG DEBUG: File doesn't start with PNG signature")
                                if file_bytes.startswith(b'<'):
                                    app.logger.error(f"PNG DEBUG: File appears to be XML/HTML, not PNG")
                                elif file_bytes.startswith(b'\xff\xd8'):
                                    app.logger.error(f"PNG DEBUG: File appears to be JPEG, not PNG")
                            
                            # Check file size again to make sure it's not empty
                            file_content.seek(0, os.SEEK_END)
                            actual_size = file_content.tell()
                            file_content.seek(0)
                            if actual_size < 100:  # Very small for a PNG
                                app.logger.error(f"PNG DEBUG: File is suspiciously small: {actual_size} bytes")
                        except Exception as raw_err:
                            app.logger.error(f"PNG DEBUG: Error during raw file analysis: {str(raw_err)}")
                        
                        traceback.print_exc()
                    
                    return False, f"Datei '{filename}' konnte nicht als Bild erkannt werden. Fehler: {str(e)}"
                    
            # Add more content type validations as needed for other file types
                
        except Exception as e:
            app.logger.error(f"Error during content validation for {filename}: {str(e)}")
            file_content.seek(0)  # Reset file pointer in case of error
            # Don't reject the file based on content validation failure alone
    
    return True, ""

"""-------------------------------------------------------------Form Value Sanitization----------------------------------------------------------------------------- """

def strip_whitespace(value):
    """
    Strip leading and trailing whitespace from a string or from each item in a list.
    
    Args:
        value: String or list of strings to strip
        
    Returns:
        String or list of strings with whitespace stripped
    """
    if isinstance(value, str):
        return value.strip()
    elif isinstance(value, list):
        return [item.strip() if isinstance(item, str) else item for item in value]
    return value


def sanitize_form_value(value):
    """
    Strip whitespace and escape HTML for a string or each string item in a list.

    Args:
        value: String, list of strings, or None

    Returns:
        Sanitized string, list of sanitized strings, or the original value if unsupported
    """
    value = strip_whitespace(value)
    if isinstance(value, str):
        return html.escape(value)
    if isinstance(value, list):
        return [html.escape(item) if isinstance(item, str) else item for item in value]
    return value

def normalize_isbn(isbn_raw):
    """Return only digits/X from ISBN input in canonical uppercase form."""
    if isbn_raw is None:
        return ''
    normalized = re.sub(r'[^0-9Xx]', '', str(isbn_raw)).upper()
    return normalized


def is_valid_isbn10(isbn10):
    if not re.fullmatch(r'[0-9]{9}[0-9X]', isbn10):
        return False
    checksum = 0
    for idx, char in enumerate(isbn10):
        value = 10 if char == 'X' else int(char)
        checksum += value * (10 - idx)
    return checksum % 11 == 0


def is_valid_isbn13(isbn13):
    if not re.fullmatch(r'[0-9]{13}', isbn13):
        return False
    checksum = 0
    for idx, char in enumerate(isbn13[:12]):
        checksum += int(char) * (1 if idx % 2 == 0 else 3)
    check_digit = (10 - (checksum % 10)) % 10
    return check_digit == int(isbn13[12])


def normalize_and_validate_isbn(isbn_raw):
    """Normalize ISBN and return a valid canonical ISBN-13/10 or empty string."""
    isbn = normalize_isbn(isbn_raw)
    if len(isbn) == 13 and is_valid_isbn13(isbn):
        return isbn
    if len(isbn) == 10 and is_valid_isbn10(isbn):
        return isbn
    return ''


def _normalize_excel_header(value):
    """Normalize Excel header labels for robust auto-mapping."""
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = text.replace('ä', 'ae').replace('ö', 'oe').replace('ü', 'ue').replace('ß', 'ss')
    text = re.sub(r'[^a-z0-9]+', '_', text)
    text = re.sub(r'_+', '_', text).strip('_')
    return text


def _excel_bool(value, default=True):
    """Parse a flexible Excel bool-like cell value."""
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'1', 'true', 'yes', 'y', 'ja', 'j', 'x'}:
        return True
    if text in {'0', 'false', 'no', 'n', 'nein'}:
        return False
    return default


def _excel_int(value):
    """Parse int-like values from Excel cells."""
    if value is None or str(value).strip() == '':
        return None
    try:
        return int(float(str(value).replace(',', '.')))
    except Exception:
        return None


def _excel_float(value):
    """Parse float-like values from Excel cells."""
    if value is None or str(value).strip() == '':
        return None
    try:
        return float(str(value).replace(',', '.'))
    except Exception:
        return None


def _excel_list(value):
    """Parse list-like filter values from comma/semicolon/newline separated strings."""
    if value is None:
        return []
    if isinstance(value, list):
        return [sanitize_form_value(v) for v in value if str(v).strip()]
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r'[;,\n\|]+', text)
    cleaned = [sanitize_form_value(p.strip()) for p in parts if p and p.strip()]
    # Deduplicate while preserving order.
    seen = set()
    unique = []
    for entry in cleaned:
        if entry not in seen:
            unique.append(entry)
            seen.add(entry)
    return unique


def _load_tabular_upload(uploaded_file):
    """Load a CSV or XLSX upload and return header row plus data rows."""
    filename = (getattr(uploaded_file, 'filename', '') or '').lower()
    file_bytes = uploaded_file.read()
    uploaded_file.stream.seek(0)

    if filename.endswith('.csv'):
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = file_bytes.decode(encoding)
                break
            except Exception:
                text = None
        if text is None:
            raise ValueError('CSV-Datei konnte nicht dekodiert werden.')

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,\t,|')
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ';' if sample.count(';') >= sample.count(',') else ','

        reader = csv.reader(io.StringIO(text), dialect)
        rows = [row for row in reader if any(str(cell).strip() for cell in row)]
        if not rows:
            raise ValueError('CSV-Datei enthält keine Daten.')
        return rows[0], rows[1:]

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError(f'Excel-Import benötigt openpyxl: {exc}') from exc

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        workbook.close()
        raise ValueError('Die Datei enthält keine Kopfzeile.')

    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    return header_row, data_rows

def _deny_if_unauthenticated_file_access():
    """Block file-serving routes unless a user is logged in."""
    if 'username' not in session:
        return Response('Forbidden', status=403)
    return None


def _student_card_id_slug(value):
    """Build a compact identifier fragment from a name or class value."""
    normalized = _normalize_excel_header(value)
    if not normalized:
        return ''
    return re.sub(r'[^a-z0-9]+', '', normalized).upper()

"""-------------------------------------------------------------Filter----------------------------------------------------------------------------- """
FILTER_SELECT_ALL_TOKEN = '__ALL__'


def expand_filter_selection(selected_values, filter_num):
    """Expand special filter token into all predefined values for a filter."""
    if not isinstance(selected_values, list):
        return []

    has_select_all = False
    cleaned_values = []
    seen = set()

    for raw_value in selected_values:
        value = str(raw_value).strip() if raw_value is not None else ''
        if not value:
            continue
        if value == FILTER_SELECT_ALL_TOKEN:
            has_select_all = True
            continue
        if value not in seen:
            cleaned_values.append(value)
            seen.add(value)

    if not has_select_all:
        return cleaned_values

    predefined_values = it.get_predefined_filter_values(filter_num)
    for predefined_value in predefined_values:
        value = str(predefined_value).strip() if predefined_value is not None else ''
        if value and value not in seen:
            cleaned_values.append(value)
            seen.add(value)

    return cleaned_values

def _is_public_host(hostname):
    """Return True only for hosts that resolve to public IPs."""
    if not hostname:
        return False

    hostname = hostname.strip().lower()
    if hostname in {'localhost', '127.0.0.1', '::1'}:
        return False

    try:
        resolved_infos = socket.getaddrinfo(hostname, None)
    except Exception:
        return False

    public_seen = False
    for info in resolved_infos:
        address = info[4][0]
        try:
            ip_obj = ipaddress.ip_address(address)
        except ValueError:
            continue
        if ip_obj.is_private or ip_obj.is_loopback or ip_obj.is_link_local or ip_obj.is_multicast or ip_obj.is_reserved or ip_obj.is_unspecified:
            return False
        public_seen = True
 
    return public_seen

"""-------------------------------------------------------------Student Cards Excel Import----------------------------------------------------------------------------- """

def _build_student_card_excel_id(student_name, class_name, row_number, used_ids):
    """Create a stable student-card ID without embedding personal names."""
    class_slug = _student_card_id_slug(class_name)

    base_parts = [part for part in (class_slug,) if part]
    if base_parts:
        base_id = f"SC-{'-'.join(base_parts[:1])}-ROW-{row_number}"
    else:
        base_id = f"SC-ROW-{row_number}"

    candidate = base_id
    suffix = 2
    while candidate in used_ids:
        candidate = f"{base_id}-{suffix}"
        suffix += 1

    used_ids.add(candidate)
    return candidate


def _upload_student_cards_excel():
    """Bulk import student cards from Excel with automatic name/class mapping."""
    if 'username' not in session:
        flash('Nicht angemeldet.', 'error')
        return redirect(url_for('login'))

    if not us.check_admin(session['username']):
        flash('Administratorrechte erforderlich.', 'error')
        return redirect(url_for('home_admin'))

    if not cfg.MODULES.is_enabled('student_cards'):
        flash('Schülerausweis-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    excel_file = request.files.get('student_cards_excel')
    if not excel_file or not excel_file.filename:
        flash('Bitte eine Excel-Datei auswählen.', 'error')
        return redirect(url_for('student_cards_admin'))

    filename_lower = excel_file.filename.lower()
    if not filename_lower.endswith(('.xlsx', '.csv')):
        flash('Nur .xlsx oder .csv Dateien werden unterstützt.', 'error')
        return redirect(url_for('student_cards_admin'))

    try:
        header_row, data_rows = _load_tabular_upload(excel_file)
    except Exception as exc:
        flash(f'Datei konnte nicht gelesen werden: {exc}', 'error')
        return redirect(url_for('student_cards_admin'))

    header_map = {}
    for idx, raw_header in enumerate(header_row):
        normalized = _normalize_excel_header(raw_header)
        if normalized:
            header_map[normalized] = idx

    synonyms = {
        'ausweis_id': ['ausweis_id', 'ausweisid', 'ausweis-id', 'karte', 'kartennummer', 'card_id', 'id'],
        'student_name': ['student_name', 'schuelername', 'schülername', 'schueler', 'schüler', 'name', 'vollname', 'vorname_nachname', 'nachname_vorname'],
        'first_name': ['vorname', 'first_name', 'firstname'],
        'last_name': ['nachname', 'last_name', 'lastname'],
        'class_name': ['klasse', 'class', 'class_name', 'jahrgang', 'jahrgangsstufe', 'stufe', 'gruppe', 'asv_klasse'],
        'notes': ['notizen', 'notes', 'bemerkungen', 'bemerkung', 'hinweis', 'hinweise'],
        'default_borrow_days': ['standard_ausleihdauer', 'ausleihdauer', 'borrow_days', 'tage', 'leihtage', 'max_borrow_days'],
    }

    def col_index(key):
        for candidate in synonyms.get(key, []):
            normalized = _normalize_excel_header(candidate)
            if normalized in header_map:
                return header_map[normalized]
        return None

    mapped_indices = {
        'ausweis_id': col_index('ausweis_id'),
        'student_name': col_index('student_name'),
        'first_name': col_index('first_name'),
        'last_name': col_index('last_name'),
        'class_name': col_index('class_name'),
        'notes': col_index('notes'),
        'default_borrow_days': col_index('default_borrow_days'),
    }

    validation_only = (request.form.get('excel_action') or '').strip().lower() == 'validate'
    max_rows = 15000

    planned_rows = []
    validation_errors = []
    validation_warnings = []
    existing_ids = set()

    client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
    try:
        db = client[cfg.MONGODB_DB]
        student_cards = db['student_cards']
        existing_ids.update(
            str(card.get('AusweisId', '')).strip().upper()
            for card in student_cards.find({}, {'AusweisId': 1})
            if card.get('AusweisId')
        )

        processed_rows = 0
        for row_number, row_values in enumerate(data_rows, start=2):
            processed_rows += 1
            if processed_rows > max_rows:
                validation_errors.append((row_number, f'Maximal {max_rows} Zeilen pro Datei erlaubt'))
                break

            def val(key):
                idx = mapped_indices.get(key)
                if idx is None or idx >= len(row_values):
                    return None
                return row_values[idx]

            ausweis_id = sanitize_form_value(val('ausweis_id'))
            student_name = sanitize_form_value(val('student_name'))
            first_name = sanitize_form_value(val('first_name'))
            last_name = sanitize_form_value(val('last_name'))
            class_name = sanitize_form_value(val('class_name'))
            notes = sanitize_form_value(val('notes'))
            default_borrow_days = _excel_int(val('default_borrow_days')) or cfg.STUDENT_DEFAULT_BORROW_DAYS

            if not student_name and first_name and last_name:
                student_name = f'{first_name} {last_name}'.strip()
                validation_warnings.append((row_number, 'Schülername wurde aus Vorname und Nachname zusammengesetzt'))

            if not ausweis_id and not student_name and not class_name:
                continue

            row_errors = []
            if not student_name:
                row_errors.append('Schülername fehlt')

            if not ausweis_id and student_name:
                ausweis_id = _build_student_card_excel_id(student_name, class_name, row_number, existing_ids)
                validation_warnings.append((row_number, f'Ausweis-ID wurde automatisch erzeugt: {ausweis_id}'))
            elif ausweis_id:
                ausweis_id = str(ausweis_id).strip().upper()
                if ausweis_id in existing_ids:
                    row_errors.append(f'Ausweis-ID {ausweis_id} existiert bereits')
                else:
                    existing_ids.add(ausweis_id)

            if row_errors:
                validation_errors.append((row_number, '; '.join(row_errors)))
                continue

            planned_rows.append({
                'row_number': row_number,
                'ausweis_id': ausweis_id,
                'student_name': student_name,
                'class_name': class_name,
                'notes': notes,
                'default_borrow_days': default_borrow_days,
            })
    finally:
        client.close()

    if validation_errors:
        details = '; '.join([f'Zeile {n}: {msg}' for n, msg in validation_errors[:15]])
        flash(f'Validierung fehlgeschlagen ({len(validation_errors)} Zeilen). {details}', 'error')
        return redirect(url_for('student_cards_admin'))

    if validation_only:
        warning_text = ''
        if validation_warnings:
            warning_details = '; '.join([f'Zeile {n}: {msg}' for n, msg in validation_warnings[:10]])
            warning_text = f' Hinweise: {warning_details}'
        flash(f'Validierung erfolgreich: {len(planned_rows)} Ausweise würden importiert.{warning_text}', 'success')
        return redirect(url_for('student_cards_admin'))

    client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
    try:
        db = client[cfg.MONGODB_DB]
        student_cards = db['student_cards']

        created_total = 0
        for row in planned_rows:
            encrypted_payload = encrypt_document_fields(
                {
                    'SchülerName': row['student_name'],
                    'Klasse': row['class_name'],
                    'Notizen': row['notes'],
                },
                STUDENT_CARD_ENCRYPTED_FIELDS
            )
            student_cards.insert_one({
                'AusweisId': row['ausweis_id'],
                'StandardAusleihdauer': int(row['default_borrow_days']),
                'Erstellt': datetime.datetime.now(),
                **encrypted_payload,
            })
            created_total += 1
    except Exception as exc:
        app.logger.error(f'Error importing student cards from Excel: {exc}')
        flash(f'Fehler beim Import der Bibliotheksausweise: {exc}', 'error')
        return redirect(url_for('student_cards_admin'))
    finally:
        client.close()

    if validation_warnings:
        warning_details = '; '.join([f'Zeile {n}: {msg}' for n, msg in validation_warnings[:10]])
        flash(f'Excel-Import erfolgreich: {created_total} Ausweise importiert. Hinweise: {warning_details}', 'warning')
    else:
        flash(f'Excel-Import erfolgreich: {created_total} Ausweise importiert.', 'success')

    return redirect(url_for('student_cards_admin'))


def _upload_excel_items(scope='inventory'):
    """Bulk import inventory/library items from Excel with validation-first workflow."""
    if 'username' not in session:
        flash('Nicht angemeldet.', 'error')
        return redirect(url_for('login'))

    permissions = _get_current_user_permissions() or us.build_default_permission_payload('standard_user')
    if not _action_access_allowed(permissions, 'can_insert'):
        flash('Einfüge-Rechte erforderlich.', 'error')
        return redirect(url_for('home'))

    is_library_scope = scope == 'library'
    file_field = 'library_excel' if is_library_scope else 'inventory_excel'
    fallback_route = 'library_admin' if is_library_scope else 'upload_admin'

    if is_library_scope:
        if not cfg.MODULES.is_enabled('library'):
            flash('Bibliotheks-Modul ist deaktiviert.', 'error')
            return redirect(url_for('home'))

    excel_file = request.files.get(file_field)
    if not excel_file or not excel_file.filename:
        flash('Bitte eine Excel-Datei auswählen.', 'error')
        return redirect(url_for(fallback_route))

    filename_lower = excel_file.filename.lower()
    if not filename_lower.endswith(('.xlsx', '.csv')):
        flash('Nur .xlsx oder .csv Dateien werden unterstützt.', 'error')
        return redirect(url_for(fallback_route))

    try:
        header_row, data_rows = _load_tabular_upload(excel_file)
    except Exception as exc:
        flash(f'Datei konnte nicht gelesen werden: {exc}', 'error')
        return redirect(url_for(fallback_route))

    header_map = {}
    for idx, raw_header in enumerate(header_row):
        normalized = _normalize_excel_header(raw_header)
        if normalized:
            header_map[normalized] = idx

    synonyms = {
        'name': ['name', 'titel', 'artikel', 'item', 'bezeichnung'],
        'ort': ['ort', 'location', 'standort', 'platz'],
        'beschreibung': ['beschreibung', 'description', 'desc', 'details'],
        'filter1': ['filter1', 'filter', 'fach', 'unterrichtsfach', 'kategorie1', 'category1'],
        'filter2': ['filter2', 'jahrgang', 'jahrgangsstufe', 'klasse', 'kategorie2', 'category2'],
        'filter3': ['filter3', 'bereich', 'gruppe', 'typ', 'kategorie3', 'category3'],
        'anschaffungsjahr': ['anschaffungsjahr', 'jahr', 'year'],
        'anschaffungskosten': ['anschaffungskosten', 'kosten', 'cost', 'preis', 'price'],
        'code_4': ['code_4', 'code4', 'code', 'inventarnummer', 'inventar_nr', 'id_code'],
        'reservierbar': ['reservierbar', 'reservable', 'bookable'],
        'anzahl': ['anzahl', 'menge', 'quantity', 'count'],
        'isbn': ['isbn'],
        'item_type': ['item_type', 'typ', 'type'],
    }

    def col_index(key):
        for candidate in synonyms.get(key, []):
            normalized = _normalize_excel_header(candidate)
            if normalized in header_map:
                return header_map[normalized]
        return None

    required_columns = {'name': col_index('name'), 'ort': col_index('ort'), 'beschreibung': col_index('beschreibung')}
    missing_required = [k for k, v in required_columns.items() if v is None]
    if missing_required:
        missing_label = ', '.join(missing_required)
        flash(f'Pflichtspalten fehlen in Excel: {missing_label}', 'error')
        return redirect(url_for(fallback_route))

    mapped_indices = {
        'name': col_index('name'),
        'ort': col_index('ort'),
        'beschreibung': col_index('beschreibung'),
        'filter1': col_index('filter1'),
        'filter2': col_index('filter2'),
        'filter3': col_index('filter3'),
        'anschaffungsjahr': col_index('anschaffungsjahr'),
        'anschaffungskosten': col_index('anschaffungskosten'),
        'code_4': col_index('code_4'),
        'reservierbar': col_index('reservierbar'),
        'anzahl': col_index('anzahl'),
        'isbn': col_index('isbn'),
        'item_type': col_index('item_type'),
    }

    validation_only = (request.form.get('excel_action') or '').strip().lower() == 'validate'

    max_rows = 15000
    max_generated_items = 50000
    max_items_per_row = 1000

    # Validation phase: parse + map + validate every row before any write.
    planned_rows = []
    validation_errors = []
    validation_warnings = []
    reserved_codes = set()
    known_locations = set(it.get_predefined_locations())
    known_filters = {
        1: set(it.get_predefined_filter_values(1)),
        2: set(it.get_predefined_filter_values(2)),
        3: set(it.get_predefined_filter_values(3))
    }

    def is_code_available(candidate_code):
        return candidate_code not in reserved_codes and it.is_code_unique(candidate_code)

    def resolve_unique_code(base_candidate):
        if is_code_available(base_candidate):
            reserved_codes.add(base_candidate)
            return base_candidate

        suffix = 1
        while suffix <= 1000:
            alt = f"{base_candidate}-{suffix}"
            if is_code_available(alt):
                reserved_codes.add(alt)
                return alt
            suffix += 1
        return None

    processed_rows = 0
    planned_item_total = 0
    row_limit_exceeded = False

    for row_number, row_values in enumerate(data_rows, start=2):
        processed_rows += 1
        if processed_rows > max_rows:
            row_limit_exceeded = True
            break

        def val(key):
            idx = mapped_indices.get(key)
            if idx is None or idx >= len(row_values):
                return None
            return row_values[idx]

        name = sanitize_form_value(val('name'))
        ort = sanitize_form_value(val('ort'))
        beschreibung = sanitize_form_value(val('beschreibung'))

        if not name and not ort and not beschreibung:
            # Fully empty row.
            continue

        row_errors = []

        if not name or not ort or not beschreibung:
            row_errors.append('Pflichtfeld Name/Ort/Beschreibung fehlt')

        if ort and ort not in known_locations:
            validation_warnings.append((row_number, f'Ort "{ort}" wird neu angelegt'))

        filter1 = expand_filter_selection(_excel_list(val('filter1')), 1)
        for f in filter1:
            if f and f not in known_filters[1]:
                validation_warnings.append((row_number, f'Kategorie 1 "{f}" wird neu angelegt'))

        filter2 = expand_filter_selection(_excel_list(val('filter2')), 2)
        for f in filter2:
            if f and f not in known_filters[2]:
                validation_warnings.append((row_number, f'Kategorie 2 "{f}" wird neu angelegt'))

        filter3 = _excel_list(val('filter3'))
        for f in filter3:
            if f and f not in known_filters[3]:
                validation_warnings.append((row_number, f'Kategorie 3 "{f}" wird neu angelegt'))

        anschaffungsjahr = _excel_int(val('anschaffungsjahr'))
        raw_year = val('anschaffungsjahr')
        if raw_year not in (None, '') and anschaffungsjahr is None:
            row_errors.append('Anschaffungsjahr ist kein gültiger Integer')

        anschaffungskosten = _excel_float(val('anschaffungskosten'))
        raw_cost = val('anschaffungskosten')
        if raw_cost not in (None, '') and anschaffungskosten is None:
            row_errors.append('Anschaffungskosten sind keine gültige Zahl')

        reservierbar = _excel_bool(val('reservierbar'), default=True)

        count = _excel_int(val('anzahl')) or 1
        raw_count = val('anzahl')
        if raw_count not in (None, '') and _excel_int(raw_count) is None:
            row_errors.append('Anzahl ist kein gültiger Integer')
        count = max(1, min(count, max_items_per_row))

        raw_code = val('code_4')
        base_code = sanitize_form_value(str(raw_code).strip()) if raw_code is not None and str(raw_code).strip() else None

        item_isbn = ''
        raw_isbn = val('isbn')
        if raw_isbn:
            item_isbn = normalize_and_validate_isbn(str(raw_isbn))
            if not item_isbn:
                row_errors.append('ISBN ist ungültig (erwartet ISBN-10/13)')
        if is_library_scope and not item_isbn:
            row_errors.append('Für den Bibliotheksimport ist eine gültige ISBN erforderlich')

        raw_item_type = sanitize_form_value(str(val('item_type')).strip().lower()) if val('item_type') else ''
        item_type = 'book' if is_library_scope else (raw_item_type or ('book' if item_isbn else 'general'))

        planned_codes = []
        if base_code:
            for position in range(1, count + 1):
                candidate = base_code if position == 1 else f"{base_code}-{position}"
                code = resolve_unique_code(candidate)
                if not code:
                    row_errors.append(f'Code konnte nicht eindeutig erzeugt werden ({candidate})')
                    break
                planned_codes.append(code)
        else:
            planned_codes = [None] * count

        if row_errors:
            validation_errors.append((row_number, '; '.join(row_errors)))
            continue

        planned_rows.append({
            'row_number': row_number,
            'name': name,
            'ort': ort,
            'beschreibung': beschreibung,
            'filter1': filter1,
            'filter2': filter2,
            'filter3': filter3,
            'anschaffungsjahr': anschaffungsjahr,
            'anschaffungskosten': anschaffungskosten,
            'reservierbar': reservierbar,
            'count': count,
            'codes': planned_codes,
            'isbn': item_isbn,
            'item_type': item_type,
        })
        planned_item_total += count

        if planned_item_total > max_generated_items:
            row_limit_exceeded = True
            break

    if row_limit_exceeded:
        flash(
            f'Importgrenze überschritten. Maximal erlaubt: {max_rows} Zeilen und {max_generated_items} erzeugte Artikel pro Datei.',
            'error'
        )
        return redirect(url_for(fallback_route))

    if validation_errors:
        details = '; '.join([f'Zeile {n}: {msg}' for n, msg in validation_errors[:15]])
        flash(f'Validierung fehlgeschlagen ({len(validation_errors)} Zeilen). {details}', 'error')
        return redirect(url_for(fallback_route))

    if validation_only:
        warning_text = ''
        if validation_warnings:
            warning_details = '; '.join([f'Zeile {n}: {msg}' for n, msg in validation_warnings[:10]])
            warning_text = f' Hinweise: {warning_details}'
        planned_count = sum(row['count'] for row in planned_rows)
        flash(f'Validierung erfolgreich: {len(planned_rows)} Zeilen geprüft, {planned_count} Artikel würden importiert.{warning_text}', 'success')
        return redirect(url_for(fallback_route))

    # Import phase: only runs after successful validation.
    created_total = 0
    import_errors = []
    for row in planned_rows:
        if row['ort'] and row['ort'] not in known_locations:
            it.add_predefined_location(row['ort'])
            known_locations.add(row['ort'])

        for f in row.get('filter1', []):
            if f and f not in known_filters[1]:
                it.add_predefined_filter_value(1, f)
                known_filters[1].add(f)

        for f in row.get('filter2', []):
            if f and f not in known_filters[2]:
                it.add_predefined_filter_value(2, f)
                known_filters[2].add(f)

        for f in row.get('filter3', []):
            if f and f not in known_filters[3]:
                it.add_predefined_filter_value(3, f)
                known_filters[3].add(f)

        series_group_id = str(uuid.uuid4()) if row['count'] > 1 else None
        row_created_ids = []

        for position in range(1, row['count'] + 1):
            parent_item_id = str(row_created_ids[0]) if row_created_ids else None
            code = row['codes'][position - 1] if position - 1 < len(row['codes']) else None
            item_id = it.add_item(
                name=row['name'],
                ort=row['ort'],
                beschreibung=row['beschreibung'],
                images=[],
                filter=row['filter1'],
                filter2=row['filter2'],
                filter3=row['filter3'],
                ansch_jahr=row['anschaffungsjahr'],
                ansch_kost=row['anschaffungskosten'],
                code_4=code,
                reservierbar=row['reservierbar'],
                series_group_id=series_group_id,
                series_count=row['count'],
                series_position=position,
                is_grouped_sub_item=(position > 1),
                parent_item_id=parent_item_id,
                isbn=row['isbn'],
                item_type=row['item_type'],
            )
            if not item_id:
                import_errors.append((row['row_number'], 'Fehler beim Anlegen des Artikels'))
                break
            row_created_ids.append(item_id)

        created_total += len(row_created_ids)

    if created_total == 0:
        details = '; '.join([f'Zeile {n}: {msg}' for n, msg in import_errors[:10]])
        flash(f'Keine Artikel importiert. {details}', 'error')
        return redirect(url_for(fallback_route))

    if import_errors:
        flash(f'Excel-Import abgeschlossen: {created_total} Artikel importiert, {len(import_errors)} Zeilen fehlgeschlagen.', 'warning')
    else:
        flash(f'Excel-Import erfolgreich: {created_total} Artikel importiert.', 'success')
    return redirect(url_for('home_admin'))


@app.route('/upload_inventory_excel', methods=['POST'])
def upload_inventory_excel():
    """Bulk import for non-library inventory."""
    return _upload_excel_items(scope='inventory')


@app.route('/upload_library_excel', methods=['POST'])
def upload_library_excel():
    """Bulk import dedicated to library items (ISBN required)."""
    return _upload_excel_items(scope='library')


@app.route('/upload_student_cards_excel', methods=['POST'])
def upload_student_cards_excel():
    """Bulk import student cards from Excel."""
    return _upload_student_cards_excel()

"""-------------------------------------------------------------File Serving-----------------------------------------------------------------------------"""

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """
    Serve uploaded files from the uploads directory.
    
    Args:
        filename (str): Name of the file to serve
        
    Returns:
        flask.Response: The requested file or placeholder image if not found
    """
    try:
        denied = _deny_if_unauthenticated_file_access()
        if denied:
            return denied

        # Check production path first (deployed environment)
        prod_path = "/opt/Inventarsystem/Web/uploads"
        dev_path = app.config['UPLOAD_FOLDER']
        if os.path.exists(os.path.join(prod_path, filename)):
            return send_from_directory(prod_path, filename)
        # Then check development path
        if os.path.exists(os.path.join(dev_path, filename)):
            return send_from_directory(dev_path, filename)
            
        # Use a placeholder image if file not found - first try SVG, then PNG
        svg_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
        png_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
        
        if os.path.exists(svg_placeholder_path):
            return send_from_directory(app.static_folder, 'img/no-image.svg')
        elif os.path.exists(png_placeholder_path):
            return send_from_directory(app.static_folder, 'img/no-image.png')
        
        # Default placeholder from static folder
        return send_from_directory(app.static_folder, 'favicon.ico')
    except Exception as e:
        app.logger.error(f"Error serving file {filename}: {str(e)}")
        return Response("Image not found", status=404)


@app.route('/thumbnails/<filename>')
def thumbnail_file(filename):
    """
    Serve thumbnail files from the thumbnails directory.
    
    Args:
        filename (str): Name of the thumbnail file to serve
        
    Returns:
        flask.Response: The requested thumbnail file or placeholder image if not found
    """
    try:
        denied = _deny_if_unauthenticated_file_access()
        if denied:
            return denied

        # Check production path first
        prod_path = "/var/Inventarsystem/Web/thumbnails"
        dev_path = app.config['THUMBNAIL_FOLDER']
        if os.path.exists(os.path.join(prod_path, filename)):
            return send_from_directory(prod_path, filename)
        if os.path.exists(os.path.join(dev_path, filename)):
            return send_from_directory(dev_path, filename)
            
        # Use a placeholder image if file not found - first try SVG, then PNG
        svg_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
        png_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
        
        if os.path.exists(svg_placeholder_path):
            return send_from_directory(app.static_folder, 'img/no-image.svg')
        elif os.path.exists(png_placeholder_path):
            return send_from_directory(app.static_folder, 'img/no-image.png')
        else:
            return send_from_directory(app.static_folder, 'favicon.ico')
    except Exception as e:
        app.logger.error(f"Error serving thumbnail {filename}: {str(e)}")
        return Response("Thumbnail not found", status=404)


@app.route('/previews/<filename>')
def preview_file(filename):
    """
    Serve preview files from the previews directory.
    
    Args:
        filename (str): Name of the preview file to serve
        
    Returns:
        flask.Response: The requested preview file or placeholder image if not found
    """
    try:
        denied = _deny_if_unauthenticated_file_access()
        if denied:
            return denied

        # Check production path first
        prod_path = "/var/Inventarsystem/Web/previews"
        dev_path = app.config['PREVIEW_FOLDER']
        if os.path.exists(os.path.join(prod_path, filename)):
            return send_from_directory(prod_path, filename)
        if os.path.exists(os.path.join(dev_path, filename)):
            return send_from_directory(dev_path, filename)
            
        # Use a placeholder image if file not found - first try SVG, then PNG
        svg_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
        png_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
        
        if os.path.exists(svg_placeholder_path):
            return send_from_directory(app.static_folder, 'img/no-image.svg')
        elif os.path.exists(png_placeholder_path):
            return send_from_directory(app.static_folder, 'img/no-image.png')
        else:
            return send_from_directory(app.static_folder, 'favicon.ico')
    except Exception as e:
        app.logger.error(f"Error serving preview {filename}: {str(e)}")
        return Response("Preview not found", status=404)


@app.route('/image/optimized/<filename>')
def optimized_image(filename):
    """
    Serve optimized images at 480p maximum resolution (854px width).
    Images are cached and converted to WebP for maximum compression.
    This endpoint minimizes server RAM usage and bandwidth.
    
    Args:
        filename (str): Original image filename
        
    Returns:
        flask.Response: Optimized image (WebP preferred, JPEG fallback) or placeholder
    """
    try:
        denied = _deny_if_unauthenticated_file_access()
        if denied:
            return denied

        # Sanitize filename to prevent directory traversal
        filename = os.path.basename(filename)
        name_part, ext_part = os.path.splitext(filename)
        
        # Determine cache directory (use unique subdirectory for 480p optimized images)
        cache_dir = app.config['THUMBNAIL_FOLDER']  # Reuse existing directory structure
        cache_subdir = os.path.join(cache_dir, 'optimized_480p')
        os.makedirs(cache_subdir, exist_ok=True)
        
        # Try to find the cached optimized image first (WebP preferred)
        cached_webp = os.path.join(cache_subdir, f"{name_part}_480p.webp")
        if os.path.exists(cached_webp):
            response = send_from_directory(cache_subdir, f"{name_part}_480p.webp")
            response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'  # 30 days
            response.headers['Content-Type'] = 'image/webp'
            return response
        
        # Try cached JPEG fallback
        cached_jpeg = os.path.join(cache_subdir, f"{name_part}_480p.jpg")
        if os.path.exists(cached_jpeg):
            response = send_from_directory(cache_subdir, f"{name_part}_480p.jpg")
            response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'  # 30 days
            response.headers['Content-Type'] = 'image/jpeg'
            return response
        
        # Find the original image
        original_paths = [
            os.path.join(app.config['UPLOAD_FOLDER'], filename),
            os.path.join("/var/Inventarsystem/Web/uploads", filename),
            os.path.join(app.config['UPLOAD_FOLDER'], f"{name_part}.webp"),
            os.path.join("/var/Inventarsystem/Web/uploads", f"{name_part}.webp"),
        ]
        
        original_image_path = None
        for path in original_paths:
            if os.path.exists(path):
                original_image_path = path
                break
        
        # If original image not found, serve placeholder
        if not original_image_path:
            svg_placeholder = os.path.join(app.static_folder, 'img', 'no-image.svg')
            if os.path.exists(svg_placeholder):
                return send_from_directory(app.static_folder, 'img/no-image.svg')
            return send_from_directory(app.static_folder, 'img/no-image.png')
        
        # Skip if it's not a supported image format
        if not is_image_file(original_image_path):
            return send_from_directory(app.static_folder, 'img/no-image.png')
        
        # Create optimized version (480p = ~854px width max)
        MAX_WIDTH = 854
        MAX_HEIGHT = 480
        
        try:
            with Image.open(original_image_path) as img:
                # Normalize orientation (fix EXIF rotation)
                img = normalize_image_orientation(img)
                
                # Resize maintaining aspect ratio
                img.thumbnail((MAX_WIDTH, MAX_HEIGHT), Image.Resampling.LANCZOS)
                
                # Try to save as WebP first (best compression)
                try:
                    img = img.convert('RGB') if img.mode in ('RGBA', 'P') else img
                    img.save(cached_webp, 'WEBP', quality=80, method=6)  # Quality 80, slowest method for best compression
                    
                    response = send_from_directory(cache_subdir, f"{name_part}_480p.webp")
                    response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'  # 30 days
                    response.headers['Content-Type'] = 'image/webp'
                    return response
                except Exception as webp_err:
                    app.logger.warning(f"WebP encoding failed for {filename}, falling back to JPEG: {str(webp_err)}")
                
                # Fallback to JPEG if WebP fails
                img = img.convert('RGB') if img.mode in ('RGBA', 'P', 'L') else img
                img.save(cached_jpeg, 'JPEG', quality=75, optimize=True)  # Quality 75, optimized
                
                response = send_from_directory(cache_subdir, f"{name_part}_480p.jpg")
                response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'  # 30 days
                response.headers['Content-Type'] = 'image/jpeg'
                return response
                
        except Exception as img_err:
            app.logger.error(f"Error processing image {filename}: {str(img_err)}")
            return send_from_directory(app.static_folder, 'img/no-image.png')
    
    except Exception as e:
        app.logger.error(f"Error serving optimized image {filename}: {str(e)}")
        return Response("Optimized image not found", status=404)


# @app.route('/QRCodes/<filename>')
# def qrcode_file(filename):
#     """
#     Serve QR code files from the QRCodes directory.
#     
#     Args:
#         filename (str): Name of the QR code file to serve
#         
#     Returns:
#         flask.Response: The requested QR code file or placeholder image if not found
#     """
#     try:
#         # Check production path first
#         prod_path = "/var/Inventarsystem/Web/QRCodes"
#         dev_path = app.config['QR_CODE_FOLDER']
#         if os.path.exists(os.path.join(prod_path, filename)):
#             return send_from_directory(prod_path, filename)
#         if os.path.exists(os.path.join(dev_path, filename)):
#             return send_from_directory(dev_path, filename)
#             
#             # Use a placeholder image if file not found - first try SVG, then PNG
#             svg_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
#             png_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
#             
#             if os.path.exists(svg_placeholder_path):
#                 return send_from_directory(app.static_folder, 'img/no-image.svg')
#             elif os.path.exists(png_placeholder_path):
#                 return send_from_directory(app.static_folder, 'img/no-image.png')
#             else:
#                 return send_from_directory(app.static_folder, 'favicon.ico')
#     except Exception as e:
#         print(f"Error serving QR code {filename}: {str(e)}")
#         return Response("QR code not found", status=404)


@app.route('/<path:filename>')
def catch_all_files(filename):
    """
    Fallback route to serve files from various directories.
    Tries to find the requested file in known directories.
    
    Args:
        filename (str): Name of the file to serve
        
    Returns:
        flask.Response: The requested file or placeholder image if not found
    """
    try:
        denied = _deny_if_unauthenticated_file_access()
        if denied:
            return denied

        # Check if the file exists in any of our directories
        possible_dirs = [
            app.config['UPLOAD_FOLDER'],
            app.config['THUMBNAIL_FOLDER'],
            app.config['PREVIEW_FOLDER'],
            # app.config['QR_CODE_FOLDER'],  # QR Code serving deactivated
            os.path.join(BASE_DIR, 'static')
        ]
        
        for directory in possible_dirs:
            file_path = os.path.join(directory, filename)
            if os.path.isfile(file_path):
                return send_from_directory(directory, os.path.basename(filename))
        
        # Check if this looks like an image request
        if any(filename.lower().endswith(ext) for ext in ['png', 'jpg', 'jpeg', 'gif', 'svg']):
            # Use a placeholder image if file not found - first try SVG, then PNG
            svg_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
            png_placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
            
            if os.path.exists(svg_placeholder_path):
                return send_from_directory(app.static_folder, 'img/no-image.svg')
            elif os.path.exists(png_placeholder_path):
                return send_from_directory(app.static_folder, 'img/no-image.png')
            else:
                return send_from_directory(app.static_folder, 'favicon.ico')
        
        # If we get here, the file wasn't found
        return Response(f"File {filename} not found", status=404)
    except Exception as e:
        print(f"Error in catch-all route for {filename}: {str(e)}")
        return Response(f"Error serving file: {str(e)}", status=500)

"""-------------------------------------------------------------Main Views-----------------------------------------------------------------------------"""

@app.route('/test_connection', methods=['GET'])
def test_connection():
    """
    Test API endpoint to verify the server is running.
    
    Returns:
        dict: Status information including version and status code
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return {'status': 'forbidden'}, 403
    return {'status': 'success', 'message': 'Connection successful', 'status_code': 200}


@app.route('/user_status')
def user_status():
    """
    API endpoint to get the current user's status (username, admin status).
    Used by JavaScript in templates to personalize the UI.
    
    Returns:
        JSON: User status information or error if not authenticated
    """
    if 'username' in session:
        is_admin = us.check_admin(session['username'])
        return jsonify({
            'authenticated': True,
            'username': session['username'],
            'is_admin': is_admin
        })
    else:
        return jsonify({
            'authenticated': False,
            'error': 'Not logged in'
        }), 401


@app.route('/')
def home():
    """
    Main route for the application homepage.
    Redirects to the appropriate view based on user role.
    
    Returns:
        flask.Response: Rendered template or redirect
    """
    if 'username' not in session:
        flash('Bitte mit registriertem Konto anmelden!', 'error')
        return redirect(url_for('login'))
        
    if not cfg.MODULES.is_enabled('inventory'):
        if cfg.MODULES.is_enabled('library'):
            return redirect(url_for('library_view'))
        else:
            return "Weder Inventar- noch Bibliotheks-Modul sind aktiviert.", 403

    elif not us.check_admin(session['username']):
        return render_template(
            'main.html',
            username=session['username'],
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
            student_default_borrow_days=cfg.STUDENT_DEFAULT_BORROW_DAYS,
            student_max_borrow_days=cfg.STUDENT_MAX_BORROW_DAYS
        )
    else:
        permissions = _get_current_user_permissions() or us.build_default_permission_payload('standard_user')
        if _page_access_allowed(permissions, 'home_admin') and _action_access_allowed(permissions, 'can_manage_settings'):
            return redirect(url_for('home_admin'))

        fallback_endpoint = _permission_denied_fallback_endpoint(permissions, current_endpoint='home')
        if fallback_endpoint == 'logout':
            flash('Für diesen Benutzer sind aktuell keine Seiten freigegeben.', 'error')
        return redirect(url_for(fallback_endpoint))


@app.route('/home_admin')
def home_admin():
    """
    Admin homepage route.
    Only accessible by users with admin privileges.
    
    Returns:
        flask.Response: Rendered template or redirect
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
        
    if not cfg.MODULES.is_enabled('inventory'):
        if cfg.MODULES.is_enabled('library'):
            return redirect(url_for('library_admin'))
        else:
            return "Weder Inventar- noch Bibliotheks-Modul sind aktiviert.", 403

    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    return render_template(
        'main_admin.html',
        username=session['username'],
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        student_default_borrow_days=cfg.STUDENT_DEFAULT_BORROW_DAYS,
        student_max_borrow_days=cfg.STUDENT_MAX_BORROW_DAYS,
        school_info=_get_school_info_for_export(),
    )


@app.route('/tutorial')
def tutorial_page():
    """Guided onboarding page (2-5 minutes) for first-time users."""
    if 'username' not in session:
        flash('Bitte mit registriertem Konto anmelden!', 'error')
        return redirect(url_for('login'))

    return render_template(
        'tutorial.html',
        username=session['username'],
        is_admin=us.check_admin(session['username']),
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        student_default_borrow_days=cfg.STUDENT_DEFAULT_BORROW_DAYS,
        student_max_borrow_days=cfg.STUDENT_MAX_BORROW_DAYS
    )

@app.route('/library')
def library_view():
    """
    Dedicated page for viewing library items (books, CDs, etc.).
    Only available when library module is enabled.
    Table-only view with customizable filter.
    
    Returns:
        flask.Response: Rendered template or redirect
    """
    if 'username' not in session:
        flash('Bitte mit registriertem Konto anmelden!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('library'):
        flash('Bibliotheks-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home'))
    
    return render_template(
        'library_table.html',
        username=session['username'],
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        is_admin=us.check_admin(session['username']),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        student_default_borrow_days=cfg.STUDENT_DEFAULT_BORROW_DAYS,
        student_max_borrow_days=cfg.STUDENT_MAX_BORROW_DAYS
    )


@app.route('/library_loans_admin')
def library_loans_admin():
    """Admin overview for library borrowings and damaged library items."""
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('library'):
        flash('Bibliotheks-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    _ensure_audit_indexes_once()

    def fmt_dt(dt):
        try:
            return dt.strftime('%d.%m.%Y %H:%M') if dt else ''
        except Exception:
            return str(dt) if dt else ''

    def fmt_money(value):
        return _format_money_value(value)

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        ausleihungen_col = db['ausleihungen']

        library_items = list(items_col.find(
            {'ItemType': {'$in': LIBRARY_ITEM_TYPES}, 'Deleted': {'$ne': True}},
            {'Name': 1, 'Code_4': 1, 'Anschaffungskosten': 1, 'Condition': 1, 'HasDamage': 1, 'DamageReports': 1, 'Verfuegbar': 1, 'User': 1, 'ItemType': 1, 'Author': 1, 'ISBN': 1}
        ))
        item_map = {str(item['_id']): item for item in library_items if item.get('_id')}
        item_ids = list(item_map.keys())

        active_records = []
        if item_ids:
            active_records = list(ausleihungen_col.find(
                {'Item': {'$in': item_ids}, 'Status': {'$in': ['active', 'planned', 'completed']}},
                {'User': 1, 'Item': 1, 'Status': 1, 'Start': 1, 'End': 1, 'Period': 1, 'Notes': 1, 'InvoiceData': 1}
            ).sort('Start', -1))

        active_item_ids = set()
        loan_entries = []
        for record in active_records:
            item_id = str(record.get('Item') or '')
            item_doc = item_map.get(item_id)
            if item_id and record.get('Status') == 'active':
                active_item_ids.add(item_id)

            if not item_doc:
                continue

            invoice_data = record.get('InvoiceData') or {}
            condition_value = str(item_doc.get('Condition', '')).strip().lower()
            item_has_damage = bool(item_doc.get('HasDamage')) or condition_value == 'destroyed' or bool(item_doc.get('DamageReports'))
            damage_reports = item_doc.get('DamageReports', []) or []

            loan_entries.append({
                'id': str(record.get('_id')),
                'item_id': item_id,
                'item_name': item_doc.get('Name', ''),
                'item_code': item_doc.get('Code_4', ''),
                'item_author': item_doc.get('Author', ''),
                'item_isbn': item_doc.get('ISBN', ''),
                'user': record.get('User', ''),
                'status': record.get('Status', ''),
                'start': fmt_dt(record.get('Start')),
                'end': fmt_dt(record.get('End')),
                'period': record.get('Period', ''),
                'notes': record.get('Notes', ''),
                'invoice_number': invoice_data.get('invoice_number', ''),
                'invoice_amount': fmt_money(invoice_data.get('amount')) if invoice_data.get('amount') is not None else fmt_money(item_doc.get('Anschaffungskosten')),
                'invoice_paid': bool(invoice_data.get('paid', False)),
                'invoice_paid_at': fmt_dt(invoice_data.get('paid_at')) if isinstance(invoice_data.get('paid_at'), datetime.datetime) else '',
                'invoice_corrections_count': len(record.get('InvoiceCorrections', []) or []),
                'has_damage': item_has_damage,
                'damage_count': len(damage_reports),
                'damage_text': (damage_reports[0].get('description', '') if damage_reports else ''),
            })

        damaged_items = []
        for item_doc in library_items:
            item_id = str(item_doc.get('_id') or '')
            condition_value = str(item_doc.get('Condition', '')).strip().lower()
            damage_reports = item_doc.get('DamageReports', []) or []
            item_has_damage = bool(item_doc.get('HasDamage')) or condition_value == 'destroyed' or bool(damage_reports)
            if not item_has_damage or item_id in active_item_ids:
                continue

            damaged_items.append({
                'id': item_id,
                'name': item_doc.get('Name', ''),
                'code': item_doc.get('Code_4', ''),
                'author': item_doc.get('Author', ''),
                'isbn': item_doc.get('ISBN', ''),
                'condition': item_doc.get('Condition', ''),
                'damage_count': len(damage_reports),
                'damage_text': (damage_reports[0].get('description', '') if damage_reports else ''),
                'available': bool(item_doc.get('Verfuegbar', False)),
                'last_updated': fmt_dt(item_doc.get('LastUpdated')),
            })

        return render_template(
            'library_borrowings_admin.html',
            loan_entries=loan_entries,
            damaged_items=damaged_items,
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        )
    except Exception as e:
        app.logger.error(f"Error loading library loans admin view: {e}")
        flash('Fehler beim Laden der Bibliotheksverwaltung.', 'error')
        return redirect(url_for('home_admin'))
    finally:
        if client:
            client.close()


@app.route('/api/library_items')
def api_library_items():
    """
    API endpoint to fetch library items (books, CDs, DVDs, media).
    Supports pagination via query params: offset, limit.
    """
    if 'username' not in session:
        return jsonify({'items': []}), 401
    
    offset_raw = request.args.get('offset', '0')
    limit_raw = request.args.get('limit', '120')

    try:
        offset = max(0, int(offset_raw))
    except (TypeError, ValueError):
        offset = 0

    try:
        limit = int(limit_raw)
    except (TypeError, ValueError):
        limit = 120
    limit = min(max(limit, 1), 500)

    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_db = db['items']
        ausleihungen_db = db['ausleihungen']

        query = {
            'ItemType': {'$in': ['book', 'cd', 'dvd', 'media']},
            'Deleted': {'$ne': True}
        }

        projection = {
            'Name': 1,
            'Autor': 1,
            'Author': 1,
            'ISBN': 1,
            'Code_4': 1,
            'Code4': 1,
            'ItemType': 1,
            'Verfuegbar': 1,
            'Condition': 1,
            'HasDamage': 1,
            'User': 1,
            'Ort': 1,
            'Beschreibung': 1,
            'Image': 1
        }

        total_count = items_db.count_documents(query)
        
        library_items = list(
            items_db.find(query, projection)
            .sort([('Name', 1), ('_id', 1)])
            .skip(offset)
            .limit(limit)
        )

        item_ids = [str(item.get('_id')) for item in library_items if item.get('_id')]
        active_records = []
        if item_ids:
            active_records = list(ausleihungen_db.find(
                {'Item': {'$in': item_ids}, 'Status': 'active'},
                {'Item': 1, 'User': 1}
            ))

        active_item_ids = set()
        active_user_by_item = {}
        for rec in active_records:
            item_id = str(rec.get('Item') or '')
            if not item_id:
                continue
            active_item_ids.add(item_id)
            if item_id not in active_user_by_item:
                active_user_by_item[item_id] = rec.get('User', '')
        
        client.close()
        
        # Convert ObjectId to string for JSON serialization
        for item in library_items:
            item_id = str(item['_id'])
            item['_id'] = item_id
            if item.get('Code4') in (None, '') and item.get('Code_4') not in (None, ''):
                item['Code4'] = item.get('Code_4')

            condition_value = str(item.get('Condition', '')).strip().lower()
            has_damage = bool(item.get('HasDamage')) or condition_value == 'destroyed'
            has_active_borrow = item_id in active_item_ids

            if has_damage and not has_active_borrow:
                item['LibraryDisplayStatus'] = 'damaged'
            elif has_active_borrow or item.get('Verfuegbar') is False:
                item['LibraryDisplayStatus'] = 'borrowed'
            else:
                item['LibraryDisplayStatus'] = 'available'

            item['BorrowedBy'] = active_user_by_item.get(item_id) or item.get('User', '')

        count = len(library_items)
        return jsonify({
            'items': library_items,
            'offset': offset,
            'limit': limit,
            'count': count,
            'total': total_count,
            'has_more': (offset + count) < total_count
        }), 200
    except Exception as e:
        app.logger.error(f"Error fetching library items: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/library_scan_action', methods=['POST'])
def api_library_scan_action():
    """
    Scan-based library workflow:
    - scan student card id
    - scan media code (ISBN/Code_4)
    - borrow if available, otherwise return (toggle behavior)
    """
    if 'username' not in session:
        return jsonify({'ok': False, 'message': 'Nicht angemeldet.'}), 401
    if not cfg.MODULES.is_enabled('library'):
        return jsonify({'ok': False, 'message': 'Bibliotheks-Modul ist deaktiviert.'}), 403
    if not cfg.MODULES.is_enabled('student_cards'):
        return jsonify({'ok': False, 'message': 'Schülerausweis-Modul ist deaktiviert.'}), 403

    payload = request.get_json(silent=True) or {}
    student_card_id = us.normalize_student_card_id(payload.get('student_card_id') or payload.get('card_id'))
    item_code_raw = str(payload.get('item_code') or payload.get('code') or '').strip()
    duration_raw = str(payload.get('borrow_duration_days') or '').strip()

    if not student_card_id:
        return jsonify({'ok': False, 'message': 'Schülerausweis-ID fehlt.'}), 400
    if not item_code_raw:
        return jsonify({'ok': False, 'message': 'Mediencode fehlt.'}), 400

    normalized_isbn = normalize_and_validate_isbn(item_code_raw)
    normalized_code = item_code_raw.upper()

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        student_cards_col = db['student_cards']
        items_col = db['items']
        ausleihungen_col = db['ausleihungen']

        card_doc = student_cards_col.find_one({'AusweisId': student_card_id})
        if not card_doc:
            return jsonify({'ok': False, 'message': 'Ungültige Schülerausweis-ID.'}), 404
        card_doc = _decrypt_student_card_doc(card_doc)

        query_or = [
            {'Code_4': item_code_raw},
            {'Code_4': normalized_code},
        ]
        if normalized_isbn:
            query_or.append({'ISBN': normalized_isbn})

        item_doc = items_col.find_one({
            'ItemType': {'$in': LIBRARY_ITEM_TYPES},
            '$or': query_or
        })

        if not item_doc:
            return jsonify({'ok': False, 'message': 'Kein Bibliotheksmedium für diesen Code gefunden.'}), 404

        item_id = str(item_doc['_id'])
        borrower_name = card_doc.get('SchülerName') or f"Ausweis {student_card_id}"
        now = datetime.datetime.now()

        if item_doc.get('Verfuegbar', True):
            borrow_duration_days = None
            if duration_raw:
                try:
                    parsed_duration = int(duration_raw)
                    if 1 <= parsed_duration <= cfg.STUDENT_MAX_BORROW_DAYS:
                        borrow_duration_days = parsed_duration
                    else:
                        return jsonify({
                            'ok': False,
                            'message': f'Ausleihdauer muss zwischen 1 und {cfg.STUDENT_MAX_BORROW_DAYS} Tagen liegen.'
                        }), 400
                except ValueError:
                    return jsonify({'ok': False, 'message': 'Ungültige Ausleihdauer.'}), 400
            else:
                try:
                    card_default = int(card_doc.get('StandardAusleihdauer', cfg.STUDENT_DEFAULT_BORROW_DAYS))
                except (TypeError, ValueError):
                    card_default = cfg.STUDENT_DEFAULT_BORROW_DAYS
                borrow_duration_days = max(1, min(card_default, cfg.STUDENT_MAX_BORROW_DAYS))

            end_date = now + datetime.timedelta(days=borrow_duration_days) if borrow_duration_days else None
            it.update_item_status(item_id, False, borrower_name)
            au.add_ausleihung(item_id, borrower_name, now, end_date=end_date)

            _append_audit_event_standalone(
                event_type='ausleihung_borrowed',
                payload={
                    'channel': 'library_scan',
                    'item_id': item_id,
                    'item_name': item_doc.get('Name', ''),
                    'borrower': borrower_name,
                    'student_card_id': student_card_id,
                    'borrow_duration_days': borrow_duration_days,
                }
            )

            return jsonify({
                'ok': True,
                'action': 'borrowed',
                'item_id': item_id,
                'item_name': item_doc.get('Name', ''),
                'student_card_id': student_card_id,
                'borrower': borrower_name,
                'borrow_duration_days': borrow_duration_days,
                'message': f"{item_doc.get('Name', 'Medium')} wurde ausgeliehen."
            }), 200

        # Toggle back: item is currently borrowed -> return
        current_borrower = str(item_doc.get('User') or '').strip()
        if current_borrower and current_borrower != borrower_name and not us.check_admin(session['username']):
            return jsonify({
                'ok': False,
                'message': f"Medium ist aktuell an '{current_borrower}' ausgeliehen und kann mit diesem Ausweis nicht zurückgegeben werden."
            }), 409

        update_result = ausleihungen_col.update_many(
            {'Item': item_id, 'Status': 'active'},
            {'$set': {
                'Status': 'completed',
                'End': now,
                'LastUpdated': now
            }}
        )
        it.update_item_status(item_id, True, borrower_name)

        _append_audit_event_standalone(
            event_type='ausleihung_returned',
            payload={
                'channel': 'library_scan',
                'item_id': item_id,
                'item_name': item_doc.get('Name', ''),
                'borrower': borrower_name,
                'student_card_id': student_card_id,
                'completed_records': update_result.modified_count,
            }
        )

        return jsonify({
            'ok': True,
            'action': 'returned',
            'item_id': item_id,
            'item_name': item_doc.get('Name', ''),
            'student_card_id': student_card_id,
            'borrower': borrower_name,
            'completed_records': update_result.modified_count,
            'message': f"{item_doc.get('Name', 'Medium')} wurde zurückgegeben."
        }), 200
    except Exception as e:
        app.logger.error(f"Error in library scan action: {e}")
        return jsonify({'ok': False, 'message': 'Fehler beim Verarbeiten des Scan-Vorgangs.'}), 500
    finally:
        if client:
            client.close()


@app.route('/api/item_detail/<item_id>')
def api_item_detail(item_id):
    """
    API endpoint to fetch detail view HTML for a library item.
    """
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        item = it.get_item(item_id)
        if not item:
            return jsonify({'error': 'Item not found'}), 404

        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen_col = db['ausleihungen']

        active_borrow = ausleihungen_col.find_one(
            {'Item': str(item.get('_id')), 'Status': 'active'},
            {'User': 1}
        )
        client.close()

        condition_value = str(item.get('Condition', '')).strip().lower()
        has_damage = bool(item.get('HasDamage')) or condition_value == 'destroyed'
        if has_damage and not active_borrow:
            status_label = 'Defekt/Zerstört'
        elif item.get('Verfuegbar') is False or active_borrow:
            status_label = 'Ausgeliehen'
        else:
            status_label = 'Verfügbar'

        borrower_value = ''
        if active_borrow:
            borrower_value = active_borrow.get('User', '')
        elif item.get('User'):
            borrower_value = item.get('User')
        
        # Basic detail HTML
        detail_html = f"""
        <h2>{html.escape(item.get('Name', 'Untitled'))}</h2>
        <p><strong>Autor/Künstler:</strong> {html.escape(item.get('Autor', item.get('Author', '-')))}</p>
        <p><strong>ISBN:</strong> {html.escape(item.get('ISBN', item.get('Code4', '-')))}</p>
        <p><strong>Beschreibung:</strong> {html.escape(item.get('Beschreibung', '-'))}</p>
        <p><strong>Status:</strong> {html.escape(status_label)}</p>
        {f'<p><strong>Ausgeliehen von:</strong> {html.escape(str(borrower_value))}</p>' if borrower_value and status_label == 'Ausgeliehen' else ''}
        """
        return detail_html, 200
    except Exception as e:
        app.logger.error(f"Error fetching item detail: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/library_item/<item_id>/update', methods=['POST'])
def api_library_item_update(item_id):
    """Admin-only API to edit library item core fields from the library table view."""
    if 'username' not in session:
        return jsonify({'ok': False, 'message': 'Nicht angemeldet.'}), 401
    if not us.check_admin(session['username']):
        return jsonify({'ok': False, 'message': 'Administratorrechte erforderlich.'}), 403
    if not cfg.MODULES.is_enabled('library'):
        return jsonify({'ok': False, 'message': 'Bibliotheks-Modul ist deaktiviert.'}), 403

    payload = request.get_json(silent=True) or {}

    name = sanitize_form_value(payload.get('name'))
    description = sanitize_form_value(payload.get('beschreibung'))
    location = sanitize_form_value(payload.get('ort'))
    author = sanitize_form_value(payload.get('autor'))
    media_type = sanitize_form_value(payload.get('item_type')).lower() or 'book'
    code_4 = sanitize_form_value(payload.get('code_4'))
    isbn_raw = sanitize_form_value(payload.get('isbn'))

    if not name or not description or not location:
        return jsonify({'ok': False, 'message': 'Name, Ort und Beschreibung sind erforderlich.'}), 400

    if media_type not in LIBRARY_ITEM_TYPES:
        return jsonify({'ok': False, 'message': 'Ungültiger Medientyp.'}), 400

    normalized_isbn = ''
    if isbn_raw:
        normalized_isbn = normalize_and_validate_isbn(isbn_raw)
        if not normalized_isbn:
            return jsonify({'ok': False, 'message': 'Ungültige ISBN (nur ISBN-10/13).'}), 400

    if code_4 and not it.is_code_unique(code_4, exclude_id=item_id):
        return jsonify({'ok': False, 'message': 'Der Code wird bereits verwendet.'}), 409

    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']

        current = items_col.find_one({'_id': ObjectId(item_id)})
        if not current:
            client.close()
            return jsonify({'ok': False, 'message': 'Element nicht gefunden.'}), 404
        if current.get('ItemType') not in LIBRARY_ITEM_TYPES:
            client.close()
            return jsonify({'ok': False, 'message': 'Element ist kein Bibliotheksmedium.'}), 400

        update_doc = {
            'Name': name,
            'Beschreibung': description,
            'Ort': location,
            'Autor': author,
            'Code_4': code_4,
            'ItemType': media_type,
            'LastUpdated': datetime.datetime.now()
        }

        if normalized_isbn:
            update_doc['ISBN'] = normalized_isbn
        elif 'ISBN' in current:
            update_doc['ISBN'] = ''

        items_col.update_one({'_id': ObjectId(item_id)}, {'$set': update_doc})
        client.close()

        return jsonify({'ok': True, 'message': 'Bibliotheksmedium aktualisiert.'}), 200
    except Exception as e:
        app.logger.error(f"Error updating library item {item_id}: {e}")
        return jsonify({'ok': False, 'message': 'Fehler beim Aktualisieren des Bibliotheksmediums.'}), 500


@app.route('/upload_admin')
def upload_admin():
    """
    Upload page route for inventory items.
    Accessible to users with insert permission.
    Supports duplication by passing duplicate_from parameter.
    
    Returns:
        flask.Response: Rendered template or redirect
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    permissions = _get_current_user_permissions() or us.build_default_permission_payload('standard_user')
    if not _action_access_allowed(permissions, 'can_insert'):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    # Check if this is a duplication request
    duplicate_from = request.args.get('duplicate_from')
    duplicate_flag = request.args.get('duplicate')  # Check for sessionStorage-based duplication
    duplicate_data = None
    
    # Handle the old method (duplicate_from parameter with item ID)
    if duplicate_from:
        try:
            original_item = it.get_item(duplicate_from)
            if original_item:
                duplicate_data = {
                    'name': original_item.get('Name', ''),
                    'description': original_item.get('Beschreibung', ''),
                    'location': original_item.get('Ort', ''),
                    'room': original_item.get('Raum', ''),
                    'category': original_item.get('Kategorie', ''),
                    'year': original_item.get('Anschaffungsjahr', ''),
                    'cost': original_item.get('Anschaffungskosten', ''),
                    'filter1': original_item.get('Filter1', ''),
                    'filter2': original_item.get('Filter2', ''),
                    'filter3': original_item.get('Filter3', ''),
                    'images': original_item.get('Images', []),
                    'original_id': duplicate_from
                }
                # Copy all filter fields (Filter1_1 through Filter3_5)
                for i in range(1, 4):  # Filter1, Filter2, Filter3
                    for j in range(1, 6):  # _1 through _5
                        filter_key = f'Filter{i}_{j}'
                        if filter_key in original_item:
                            duplicate_data[f'filter{i}_{j}'] = original_item[filter_key]
                
                flash('Element wird dupliziert. Bitte überprüfen Sie die Daten und passen Sie sie bei Bedarf an.', 'info')
            else:
                flash('Ursprungs-Element für Duplizierung nicht gefunden.', 'error')
        except Exception as e:
            app.logger.warning(f"Error loading item for duplication: {e}")
            flash('Fehler beim Laden der Duplizierungsdaten.', 'error')
    
    # Handle the new method (sessionStorage-based duplication)
    elif duplicate_flag == 'true':
        # No server-side processing needed - JavaScript will handle sessionStorage data
        # Just indicate that duplication mode is active
        flash('Element wird dupliziert. Die Daten werden aus dem Session-Speicher geladen.', 'info')
    
    return render_template(
        'upload_admin.html',
        username=session['username'],
        duplicate_data=duplicate_data,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        show_library_features=False,
        upload_mode='item',
        page_title='Artikel hochladen',
        back_target='home_admin'
    )


@app.route('/library_admin')
def library_admin():
    """
    Dedicated page for library/book uploads with ISBN scanning.
    Accessible to users with insert permission when the library module is enabled.
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    permissions = _get_current_user_permissions() or us.build_default_permission_payload('standard_user')
    if not _action_access_allowed(permissions, 'can_insert'):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('library'):
        flash('Bibliotheks-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    duplicate_from = request.args.get('duplicate_from')
    duplicate_flag = request.args.get('duplicate')
    duplicate_data = None

    if duplicate_from:
        try:
            original_item = it.get_item(duplicate_from)
            if original_item:
                duplicate_data = {
                    'name': original_item.get('Name', ''),
                    'description': original_item.get('Beschreibung', ''),
                    'location': original_item.get('Ort', ''),
                    'room': original_item.get('Raum', ''),
                    'category': original_item.get('Kategorie', ''),
                    'year': original_item.get('Anschaffungsjahr', ''),
                    'cost': original_item.get('Anschaffungskosten', ''),
                    'filter1': original_item.get('Filter1', ''),
                    'filter2': original_item.get('Filter2', ''),
                    'filter3': original_item.get('Filter3', ''),
                    'images': original_item.get('Images', []),
                    'original_id': duplicate_from
                }
                for i in range(1, 4):
                    for j in range(1, 6):
                        filter_key = f'Filter{i}_{j}'
                        if filter_key in original_item:
                            duplicate_data[f'filter{i}_{j}'] = original_item[filter_key]
                flash('Buch wird dupliziert. Bitte überprüfen Sie die Daten und passen Sie sie bei Bedarf an.', 'info')
            else:
                flash('Ursprungs-Element für Duplizierung nicht gefunden.', 'error')
        except Exception as e:
            app.logger.warning(f"Error loading item for duplication: {e}")
            flash('Fehler beim Laden der Duplizierungsdaten.', 'error')
    elif duplicate_flag == 'true':
        flash('Buch wird dupliziert. Die Daten werden aus dem Session-Speicher geladen.', 'info')

    return render_template(
        'upload_admin.html',
        username=session['username'],
        duplicate_data=duplicate_data,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        show_library_features=True,
        upload_mode='library',
        page_title='Bücher hochladen',
        back_target='home_admin'
    )


@app.route('/student_cards_admin', methods=['GET', 'POST'])
def student_cards_admin():
    """
    Admin page for managing student library cards (Bibliotheksausweis).
    Only accessible by admins and only when the student cards module is enabled.
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('student_cards'):
        flash('Schülerausweis-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
    db = client[cfg.MONGODB_DB]
    student_cards = db['student_cards']

    edit_mode = False
    form_data = {}

    # Handle GET request to edit a card
    edit_id = request.args.get('edit')
    if edit_id:
        try:
            card = student_cards.find_one({'_id': ObjectId(edit_id)})
            if card:
                card = _decrypt_student_card_doc(card)
                edit_mode = True
                form_data = {
                    'card_id': str(card['_id']),
                    'ausweis_id': card.get('AusweisId', ''),
                    'student_name': card.get('SchülerName', ''),
                    'default_borrow_days': card.get('StandardAusleihdauer', 14),
                    'class_name': card.get('Klasse', ''),
                    'notes': card.get('Notizen', '')
                }
        except Exception as e:
            app.logger.error(f"Error loading student card for edit: {e}")
            flash('Fehler beim Laden des Ausweises.', 'error')

    # Handle POST request (add or edit)
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        ausweis_id = request.form.get('ausweis_id', '').strip().upper()
        student_name = request.form.get('student_name', '').strip()
        student_name_alias = student_name 
        default_borrow_days = request.form.get('default_borrow_days', 14)
        class_name = request.form.get('class_name', '').strip()
        notes = request.form.get('notes', '').strip()

        if action == 'delete':
            try:
                card_id = request.form.get('card_id')
                student_cards.delete_one({'_id': ObjectId(card_id)})
                flash('Ausweis wurde gelöscht.', 'success')
            except Exception as e:
                app.logger.error(f"Error deleting student card: {e}")
                flash('Fehler beim Löschen des Ausweises.', 'error')

        elif action == 'edit':
            if not ausweis_id or not student_name:
                flash('Bitte Ausweis-ID und Schülername angeben.', 'error')
            else:
                try:
                    card_id = request.form.get('card_id')
                    # Check if new ID already exists (and it's not the same card)
                    existing = student_cards.find_one({'AusweisId': ausweis_id, '_id': {'$ne': ObjectId(card_id)}})
                    if existing:
                        flash('Diese Ausweis-ID existiert bereits.', 'error')
                    else:
                        encrypted_payload = encrypt_document_fields(
                            {
                                'SchülerName': student_name_alias,
                                'Klasse': class_name,
                                'Notizen': notes,
                            },
                            STUDENT_CARD_ENCRYPTED_FIELDS
                        )
                        student_cards.update_one(
                            {'_id': ObjectId(card_id)},
                            {'$set': {
                                'AusweisId': ausweis_id,
                                'StandardAusleihdauer': int(default_borrow_days),
                                'Aktualisiert': datetime.datetime.now(),
                                **encrypted_payload
                            }}
                        )
                        flash('Ausweis wurde aktualisiert.', 'success')
                        return redirect(url_for('student_cards_admin'))
                except Exception as e:
                    app.logger.error(f"Error updating student card: {e}")
                    flash('Fehler beim Aktualisieren des Ausweises.', 'error')

        elif action == 'add':
            if not ausweis_id or not student_name:
                flash('Bitte Ausweis-ID und Schülername angeben.', 'error')
            else:
                # Check if ID already exists
                existing = student_cards.find_one({'AusweisId': ausweis_id})
                if existing:
                    flash('Diese Ausweis-ID existiert bereits.', 'error')
                else:
                    try:
                        encrypted_payload = encrypt_document_fields(
                            {
                                'SchülerName': student_name_alias,
                                'Klasse': class_name,
                                'Notizen': notes,
                            },
                            STUDENT_CARD_ENCRYPTED_FIELDS
                        )
                        student_cards.insert_one({
                            'AusweisId': ausweis_id,
                            'StandardAusleihdauer': int(default_borrow_days),
                            'Erstellt': datetime.datetime.now(),
                            **encrypted_payload,
                        })
                        flash('Neuer Ausweis wurde hinzugefügt.', 'success')
                        return redirect(url_for('student_cards_admin'))
                    except Exception as e:
                        app.logger.error(f"Error adding student card: {e}")
                        flash('Fehler beim Hinzufügen des Ausweises.', 'error')

    # Get all student cards
    all_cards = list(student_cards.find().sort('AusweisId', 1))
    all_cards = [_decrypt_student_card_doc(card) for card in all_cards]
    client.close()

    return render_template(
        'student_cards_admin.html',
        username=session['username'],
        student_cards=all_cards,
        edit_mode=edit_mode,
        form_data=form_data,
        config={'default': cfg.STUDENT_DEFAULT_BORROW_DAYS},
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards')
    )


@app.route('/student_cards_print', methods=['GET'])
def student_cards_print():
    """
    Generate a printable template for all student library cards (Bibliotheksausweis).
    Only accessible by admins and only when the student cards module is enabled.
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('student_cards'):
        flash('Schülerausweis-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
    db = client[cfg.MONGODB_DB]
    student_cards = db['student_cards']

    # Get all student cards sorted by ID
    all_cards = list(student_cards.find().sort('AusweisId', 1))
    all_cards = [_decrypt_student_card_doc(card) for card in all_cards]
    client.close()

    return render_template(
        'student_cards_print.html',
        student_cards=all_cards,
        current_datetime=datetime.datetime.now()
    )


@app.route('/student_card_barcode_print', methods=['GET'])
def student_card_barcode_print():
    """
    Generate a barcode-based print template for student library cards.
    Only accessible by admins and only when the student cards module is enabled.
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('student_cards'):
        flash('Schülerausweis-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
    db = client[cfg.MONGODB_DB]
    student_cards = db['student_cards']

    # Get all student cards sorted by ID
    all_cards = list(student_cards.find().sort('AusweisId', 1))
    all_cards = [_decrypt_student_card_doc(card) for card in all_cards]
    client.close()

    return render_template(
        'student_card_barcode_print.html',
        student_cards=all_cards,
        current_datetime=datetime.datetime.now(),
        download_link=url_for('student_card_barcode_download')
    )


@app.route('/student_card_barcode_download', methods=['GET'])
def student_card_barcode_download():
    """
    Download PDF with all student card barcodes (simplified version).
    """
    if 'username' not in session:
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Zugriff verweigert.', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('student_cards'):
        flash('Schülerausweis-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, white, black
        from io import BytesIO
        import barcode
        from barcode.writer import ImageWriter
        import tempfile
        import os
        
        client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
        db = client[cfg.MONGODB_DB]
        student_cards = db['student_cards']
        all_cards = list(student_cards.find().sort('AusweisId', 1))
        all_cards = [_decrypt_student_card_doc(card) for card in all_cards]
        client.close()

        pdf_buffer = BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=A4)
        
        page_width, page_height = A4
        margin = 10 * mm
        card_width = 88 * mm
        card_height = 56 * mm
        cols = 2
        gap_x = 8 * mm
        gap_y = 8 * mm
        x_positions = [margin, margin + card_width + gap_x]
        y_start = page_height - margin
        y_pos = y_start
        col_idx = 0
        
        # Professional color palette
        header_color = HexColor("#0F172A")  # Sehr dunkles blau
        accent_color = HexColor("#2563EB")  # Helles blau
        light_bg = HexColor("#F8FAFC")      # Sehr heller grau
        card_bg_color = HexColor("#FFFFFF") # Weiß
        text_dark = HexColor("#1E293B")     # Dunkler text
        text_gray = HexColor("#64748B")     # Grauer text
        
        for i, card in enumerate(all_cards):
            if col_idx == 0 and i > 0:
                y_pos -= (card_height + gap_y)
            
            # New page if needed
            if y_pos - card_height < margin:
                c.showPage()
                y_pos = y_start
                col_idx = 0
            
            x_pos = x_positions[col_idx]
            
            # Card shadow effect (unter border)
            c.setFillColor(HexColor("#E2E8F0"))
            c.rect(x_pos + 0.5*mm, y_pos - card_height - 0.5*mm, card_width, card_height, fill=1, stroke=0)
            
            # Card background
            c.setLineWidth(1)
            c.setFillColor(card_bg_color)
            c.setStrokeColor(HexColor("#CBD5E1"))
            c.rect(x_pos, y_pos - card_height, card_width, card_height, fill=1, stroke=1)
            
            # Left info section (38mm)
            info_width = 38 * mm
            c.setFillColor(light_bg)
            c.rect(x_pos, y_pos - card_height, info_width, card_height, fill=1, stroke=0)
            
            # Header bar
            c.setFillColor(header_color)
            c.rect(x_pos, y_pos - 10*mm, card_width, 10*mm, fill=1, stroke=0)
            
            # Header accent line
            c.setStrokeColor(accent_color)
            c.setLineWidth(2)
            c.line(x_pos, y_pos - 10*mm, x_pos + card_width, y_pos - 10*mm)
            
            # "SCHÜLERAUSWEIS" text in header
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(white)
            c.drawString(x_pos + 3*mm, y_pos - 6.5*mm, "SCHÜLERAUSWEIS")
            
            # Student name - large and bold
            c.setFillColor(text_dark)
            c.setFont("Helvetica-Bold", 10)
            name = card['SchülerName'][:20]
            c.drawString(x_pos + 3*mm, y_pos - 14*mm, name)
            
            # ID with label
            c.setFont("Helvetica", 8)
            c.setFillColor(text_gray)
            c.drawString(x_pos + 3*mm, y_pos - 18*mm, "Ausweis ID:")
            c.setFillColor(text_dark)
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_pos + 3*mm, y_pos - 21*mm, str(card['AusweisId']))
            
            # Class with label
            if card.get('Klasse'):
                c.setFont("Helvetica", 8)
                c.setFillColor(text_gray)
                c.drawString(x_pos + 3*mm, y_pos - 25*mm, "Klasse:")
                c.setFillColor(text_dark)
                c.setFont("Helvetica-Bold", 9)
                c.drawString(x_pos + 3*mm, y_pos - 28*mm, card['Klasse'])
            
            # Right barcode section with border highlight
            barcode_x_start = x_pos + info_width + 1*mm
            c.setFillColor(accent_color)
            c.setLineWidth(0)
            c.rect(barcode_x_start - 1*mm, y_pos - card_height, 
                   card_width - info_width + 1*mm, 2*mm, fill=1, stroke=0)
            
            # Generate and add larger barcode
            try:
                temp_dir = tempfile.gettempdir()
                barcode_path = os.path.join(temp_dir, f"barcode_{card['AusweisId']}")
                
                # Generate barcode with higher module width for better scanning
                barcode_obj = barcode.get('code128', str(card['AusweisId']), writer=ImageWriter())
                barcode_obj.save(barcode_path)
                barcode_file = f"{barcode_path}.png"
                
                if os.path.exists(barcode_file):
                    # Larger barcode taking up most of right section
                    barcode_width = (card_width - info_width - 4*mm)
                    barcode_height = 16*mm
                    barcode_y = y_pos - card_height + (card_height - barcode_height) / 2 + 2*mm
                    
                    c.drawImage(barcode_file, 
                               barcode_x_start + 1*mm, 
                               barcode_y, 
                               width=barcode_width, 
                               height=barcode_height, 
                               preserveAspectRatio=True)
                    os.remove(barcode_file)
                else:
                    raise Exception("Barcode file not created")
            except Exception as e:
                c.setFont("Helvetica", 7)
                c.setFillColor(HexColor("#DC2626"))
                c.drawString(barcode_x_start + 2*mm, y_pos - card_height + 20*mm, "⚠️ Barcode Error")
            
            col_idx += 1
            if col_idx >= cols:
                col_idx = 0
        
        c.save()
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'schuelerausweise_all_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
    except Exception as e:
        flash(f'Fehler beim PDF-Download: {str(e)}', 'error')
        return redirect(url_for('student_cards_admin'))


@app.route('/student_card_single_barcode_download/<card_id>', methods=['GET'])
def student_card_single_barcode_download(card_id):
    """
    Download PDF with single student card barcode.
    """
    if 'username' not in session:
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Zugriff verweigert.', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('student_cards'):
        flash('Schülerausweis-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor, white, black
        from io import BytesIO
        from bson.objectid import ObjectId
        import barcode
        from barcode.writer import ImageWriter
        import tempfile
        import os
        
        client = MongoClient(cfg.MONGODB_HOST, cfg.MONGODB_PORT)
        db = client[cfg.MONGODB_DB]
        student_cards = db['student_cards']
        card = student_cards.find_one({'_id': ObjectId(card_id)})
        card = _decrypt_student_card_doc(card)
        client.close()
        
        if not card:
            flash('Ausweis nicht gefunden.', 'error')
            return redirect(url_for('student_cards_admin'))

        pdf_buffer = BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=A4)
        
        page_width, page_height = A4
        margin = 20 * mm
        card_width = 105 * mm
        card_height = 70 * mm
        x_pos = (page_width - card_width) / 2
        y_pos = (page_height - card_height) / 2
        
        # Professional color palette
        header_color = HexColor("#0F172A")  # Sehr dunkles blau
        accent_color = HexColor("#2563EB")  # Helles blau
        light_bg = HexColor("#F8FAFC")      # Sehr heller grau
        card_bg_color = HexColor("#FFFFFF") # Weiß
        text_dark = HexColor("#1E293B")     # Dunkler text
        text_gray = HexColor("#64748B")     # Grauer text
        
        # Card shadow effect
        c.setFillColor(HexColor("#E2E8F0"))
        c.rect(x_pos + 1*mm, y_pos - card_height - 1*mm, card_width, card_height, fill=1, stroke=0)
        
        # Card background
        c.setLineWidth(1.5)
        c.setFillColor(card_bg_color)
        c.setStrokeColor(HexColor("#CBD5E1"))
        c.rect(x_pos, y_pos, card_width, card_height, fill=1, stroke=1)
        
        # Left info section (50mm)
        info_width = 50 * mm
        c.setFillColor(light_bg)
        c.rect(x_pos, y_pos - card_height, info_width, card_height, fill=1, stroke=0)
        
        # Vertical divider line
        c.setStrokeColor(accent_color)
        c.setLineWidth(3)
        c.line(x_pos + info_width, y_pos - card_height, x_pos + info_width, y_pos)
        
        # Header bar
        c.setFillColor(header_color)
        c.rect(x_pos, y_pos - 10*mm, card_width, 10*mm, fill=1, stroke=0)
        
        # Header accent line
        c.setStrokeColor(accent_color)
        c.setLineWidth(2.5)
        c.line(x_pos, y_pos - 10*mm, x_pos + card_width, y_pos - 10*mm)
        
        # "SCHÜLERAUSWEIS" text in header
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(white)
        c.drawString(x_pos + 4*mm, y_pos - 6.5*mm, "SCHÜLERAUSWEIS")
        
        # Student name - large and prominent
        c.setFillColor(text_dark)
        c.setFont("Helvetica-Bold", 12)
        name = card['SchülerName'][:25]
        c.drawString(x_pos + 4*mm, y_pos - 16*mm, name)
        
        # ID section with label
        c.setFont("Helvetica", 9)
        c.setFillColor(text_gray)
        c.drawString(x_pos + 4*mm, y_pos - 21*mm, "Ausweis-ID:")
        
        c.setFillColor(text_dark)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x_pos + 4*mm, y_pos - 25*mm, str(card['AusweisId']))
        
        # Class section
        if card.get('Klasse'):
            c.setFont("Helvetica", 9)
            c.setFillColor(text_gray)
            c.drawString(x_pos + 4*mm, y_pos - 30*mm, "Klasse:")
            
            c.setFillColor(text_dark)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(x_pos + 4*mm, y_pos - 34*mm, card['Klasse'])
        
        # Barcode section - right side with blue accent
        barcode_x_start = x_pos + info_width + 3*mm
        c.setFillColor(accent_color)
        c.setLineWidth(0)
        c.rect(x_pos + info_width, y_pos - card_height, 
               card_width - info_width, 3*mm, fill=1, stroke=0)
        
        # Generate and add large barcode
        try:
            temp_dir = tempfile.gettempdir()
            barcode_path = os.path.join(temp_dir, f"barcode_{card['AusweisId']}")
            
            # Generate barcode with better sizing
            barcode_obj = barcode.get('code128', str(card['AusweisId']), writer=ImageWriter())
            barcode_obj.save(barcode_path)
            barcode_file = f"{barcode_path}.png"
            
            if os.path.exists(barcode_file):
                # Large barcode on right side
                barcode_width = (card_width - info_width - 8*mm)
                barcode_height = 20*mm
                barcode_y_offset = (card_height - 10*mm - barcode_height) / 2
                
                c.drawImage(barcode_file, 
                           barcode_x_start, 
                           y_pos - card_height + barcode_y_offset + 5*mm, 
                           width=barcode_width, 
                           height=barcode_height, 
                           preserveAspectRatio=True)
                os.remove(barcode_file)
            else:
                raise Exception("Barcode file not created")
        except Exception as e:
            c.setFont("Helvetica", 9)
            c.setFillColor(HexColor("#DC2626"))
            c.drawString(barcode_x_start + 2*mm, y_pos - card_height + 25*mm, "⚠️ Barcode Error")
        
        # Bottom info line
        c.setStrokeColor(HexColor("#CBD5E1"))
        c.setLineWidth(0.5)
        c.line(x_pos, y_pos - card_height + 3*mm, x_pos + card_width, y_pos - card_height + 3*mm)
        
        c.save()
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'ausweis_{card["AusweisId"]}.pdf'
        )
    except Exception as e:
        flash(f'Fehler beim PDF-Download: {str(e)}', 'error')
        return redirect(url_for('student_cards_admin'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    User login route.
    Authenticates users and redirects to appropriate homepage based on role.
    
    Returns:
        flask.Response: Rendered template or redirect
    """
    if 'username' in session:
        return redirect(url_for('home'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        ctx = get_tenant_context()
        current_tenant_id = ctx.tenant_id if ctx else None
        current_tenant_db = ctx.db_name if ctx else cfg.MONGODB_DB
        app.logger.info(f"Login attempt: username={username!r} tenant={current_tenant_id or 'default'} db={current_tenant_db} host={request.host} ip={request.remote_addr}")
        app.logger.info(f"Debug login context: headers={dict(request.headers)} tenant_config={ctx.config if ctx else None} remote_addr={request.remote_addr} host={request.host}")
        app.logger.info(f"Raw login payload: username={username!r} password={password!r}")
        app.logger.info(f"Active MongoDB config: uri={getattr(cfg, 'MONGODB_URI', None)!r} host={cfg.MONGODB_HOST!r} port={cfg.MONGODB_PORT!r} default_db={cfg.MONGODB_DB!r}")
        if not username or not password:
            app.logger.warning(f"Login blocked: missing credentials tenant={current_tenant_id or 'default'} host={request.host} ip={request.remote_addr}")
            flash('Bitte alle Felder ausfüllen', 'error')
            return redirect(url_for('login'))
        
        user = us.check_nm_pwd(username, password)

        if user:
            app.logger.info(f"Login success: username={username!r} tenant={current_tenant_id or 'default'} db={current_tenant_db} host={request.host} ip={request.remote_addr}")
            session['username'] = username
            is_admin_user = bool(user.get('Admin', False))
            session['admin'] = is_admin_user
            session['is_admin'] = is_admin_user
            # Bind session favorites to the authenticated user to avoid cross-user leakage.
            try:
                session['favorites_owner'] = username
                session['favorites'] = list(dict.fromkeys([str(f) for f in us.get_favorites(username)]))
            except Exception:
                session['favorites_owner'] = username
                session['favorites'] = []
            if is_admin_user:
                permissions = us.get_effective_permissions(username)
                if _page_access_allowed(permissions, 'home_admin') and _action_access_allowed(permissions, 'can_manage_settings'):
                    return redirect(url_for('home_admin'))
                fallback_endpoint = _permission_denied_fallback_endpoint(permissions, current_endpoint='login')
                return redirect(url_for(fallback_endpoint))
            else:
                return redirect(url_for('home'))
        else:
            app.logger.warning(f"Login failed: username={username!r} tenant={current_tenant_id or 'default'} db={current_tenant_db} host={request.host} ip={request.remote_addr}")
            flash('Ungültige Anmeldedaten', 'error')
            get_flashed_messages()
    return render_template('login.html')


@app.route('/impressum')
def impressum():
    """
    Impressum route.

    Returns:
        flask.Response: Redirect to impressum
    """
    return render_template('impressum.html')

@app.route('/license')
def license():
    """
    License information route.
    Displays the Apache 2.0 license information.

    Returns:
        flask.Response: Rendered license template
    """
    return render_template('license.html')

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    """
    Change password route.
    Allows users to change their password if logged in.
    
    Returns:
        flask.Response: Rendered form or redirect after password change
    """
    if 'username' not in session:
        flash('Sie müssen angemeldet sein, um Ihr Passwort zu ändern.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        current_password = html.escape(request.form.get('current_password'))
        new_password = html.escape(request.form.get('new_password'))
        confirm_password = html.escape(request.form.get('confirm_password'))
        
        # Validate inputs
        if not all([current_password, new_password, confirm_password]):
            flash('Bitte füllen Sie alle Felder aus.', 'error')
            return render_template('change_password.html')
            
        if new_password != confirm_password:
            flash('Die neuen Passwörter stimmen nicht überein.', 'error')
            return render_template('change_password.html')
            
        # Verify current password
        user = us.check_nm_pwd(session['username'], current_password)
        if not user:
            flash('Das aktuelle Passwort ist nicht korrekt.', 'error')
            return render_template('change_password.html')
            
        # Check password strength
        if not us.check_password_strength(new_password):
            flash('Das neue Passwort ist zu schwach. Es sollte mindestens 6 Zeichen lang sein.', 'error')
            return render_template('change_password.html')
            
        # Update the password
        if us.update_password(session['username'], new_password):
            flash('Ihr Passwort wurde erfolgreich geändert.', 'success')
            return redirect(url_for('home'))
        else:
            flash('Fehler beim Ändern des Passworts. Bitte versuchen Sie es später erneut.', 'error')
            
    return render_template('change_password.html')

@app.route('/logout')
def logout():
    """
    User logout route.
    Removes user session data and redirects to login.
    
    Returns:
        flask.Response: Redirect to login page
    """
    session.pop('username', None)
    session.pop('admin', None)
    session.pop('is_admin', None)
    session.pop('favorites', None)
    session.pop('favorites_owner', None)
    return redirect(url_for('login'))


@app.route('/get_items', methods=['GET'])
def get_items():
    """Return items plus merged favorites (session + DB) and per-item favorite flag."""
    client = None
    try:
        _ensure_session_favs()
        username = session.get('username')
        # Merge DB favorites into session if logged in
        if username:
            try:
                db_favs = set(us.get_favorites(username))
                session_favs = set(session.get('favorites', []))
                merged = list(db_favs.union(session_favs))
                session['favorites'] = merged
            except Exception as fav_err:
                app.logger.warning(f"Could not merge DB favorites: {fav_err}")
        favorites = set(session.get('favorites', []))

        available_only = str(request.args.get('available_only', '')).strip().lower() in ('1', 'true', 'yes', 'on')
        offset_raw = request.args.get('offset')
        limit_raw = request.args.get('limit')
        light_mode_param = str(request.args.get('light_mode', '')).strip().lower()
        pagination_requested = offset_raw is not None or limit_raw is not None

        if pagination_requested:
            try:
                offset = max(0, int(offset_raw or '0'))
            except (TypeError, ValueError):
                offset = 0
            try:
                limit = int(limit_raw or '120')
            except (TypeError, ValueError):
                limit = 120
            limit = min(max(limit, 1), 500)
        else:
            offset = 0
            limit = None

        if light_mode_param in ('1', 'true', 'yes', 'on'):
            light_mode = True
        elif light_mode_param in ('0', 'false', 'no', 'off'):
            light_mode = False
        else:
            light_mode = (offset == 0)

        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        base_query = {
            'IsGroupedSubItem': {'$ne': True},
            'ItemType': {'$nin': LIBRARY_ITEM_TYPES},
            'Deleted': {'$ne': True},
        }

        total_count = items_col.count_documents(base_query)

        light_projection = {
            'Name': 1,
            'Code_4': 1,
            'Images': 1,
            'ThumbnailInfo': 1,
            'Verfuegbar': 1,
            'Filter': 1,
            'Filter2': 1,
            'Filter3': 1,
            'Ort': 1,
            'User': 1,
            'BlockedNow': 1,
            'Reservierbar': 1,
            'HasDamage': 1,
            'ItemType': 1,
        }

        # Full projection: all details for detailed view
        full_projection = {
            'Name': 1,
            'Ort': 1,
            'Beschreibung': 1,
            'Filter': 1,
            'Filter2': 1,
            'Filter3': 1,
            'Code_4': 1,
            'Images': 1,
            'ThumbnailInfo': 1,
            'Verfuegbar': 1,
            'User': 1,
            'BorrowerInfo': 1,
            'appointments': 1,
            'BlockedNow': 1,
            'Reservierbar': 1,
            'DamageReports': 1,
            'ISBN': 1,
            'Author': 1,
            'Autor': 1,
            'Anschaffungsjahr': 1,
            'Anschaffungskosten': 1,
            'Condition': 1,
            'HasDamage': 1,
            'ItemType': 1,
            'SeriesGroupId': 1,
            'LastUpdated': 1,
        }

        parent_projection = light_projection if light_mode else full_projection

        items_cur = items_col.find(base_query, parent_projection).sort([('Name', 1), ('_id', 1)])
        if pagination_requested:
            items_cur = items_cur.skip(offset).limit(limit)

        parent_items = list(items_cur)
        parent_ids = [str(item.get('_id')) for item in parent_items if item.get('_id') is not None]

        children_by_parent = {}
        if parent_ids:
            # Light mode: minimal child data for counting only
            light_child_projection = {
                '_id': 1,
                'ParentItemId': 1,
                'Code_4': 1,
                'Verfuegbar': 1,
                'Name': 1,
            }
            # Full mode: complete child data with all details
            full_child_projection = {
                '_id': 1,
                'ParentItemId': 1,
                'Code_4': 1,
                'Verfuegbar': 1,
                'Name': 1,
                'Images': 1,
                'ThumbnailInfo': 1,
                'Beschreibung': 1,
            }
            child_projection = light_child_projection if light_mode else full_child_projection
            
            child_cursor = items_col.find({
                'ParentItemId': {'$in': parent_ids},
                'IsGroupedSubItem': True,
                'Deleted': {'$ne': True},
            }, child_projection)
            for child in child_cursor:
                parent_id = str(child.get('ParentItemId') or '')
                if not parent_id:
                    continue
                children_by_parent.setdefault(parent_id, []).append(child)

        items = []
        for itm in parent_items:
            item_id_str = str(itm['_id'])
            grouped_children = children_by_parent.get(item_id_str, [])
            grouped_count = 1 + len(grouped_children)

            grouped_units = [itm] + grouped_children
            available_units = []
            grouped_all_codes = []
            for unit in grouped_units:
                unit_code = unit.get('Code_4')
                if unit_code is not None and str(unit_code).strip() != '':
                    grouped_all_codes.append(str(unit_code).strip())

                if unit.get('Verfuegbar', True):
                    unit_id = str(unit['_id']) if not isinstance(unit['_id'], str) else unit['_id']
                    code = unit.get('Code_4') or '-'
                    available_units.append({
                        'id': unit_id,
                        'code': code,
                        'label': f"{code} ({unit.get('Name', 'Item')})"
                    })

            itm['_id'] = item_id_str
            itm['GroupedDisplayCount'] = grouped_count
            itm['AvailableGroupedCount'] = len(available_units)
            itm['GroupedAvailableUnits'] = available_units
            itm['GroupedAllCodes'] = grouped_all_codes
            if grouped_count > 1:
                itm['Verfuegbar'] = len(available_units) > 0
            if available_only and not itm.get('Verfuegbar', False):
                continue
            itm['is_favorite'] = item_id_str in favorites
            items.append(itm)

        count = len(items)
        return jsonify({
            'items': items,
            'favorites': list(favorites),
            'offset': offset,
            'limit': limit if limit is not None else count,
            'count': count,
            'total': total_count,
            'light_mode': light_mode,
            'has_more': pagination_requested and ((offset + count) < total_count)
        })
    except Exception as e:
        return jsonify({'items': [], 'error': str(e)}), 500
    finally:
        if client:
            client.close()


@app.route('/get_item/<id>')
def get_item_json(id):
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        item = db['items'].find_one({'_id': ObjectId(id), 'Deleted': {'$ne': True}})
        if not item:
            return jsonify({'error': 'not found'}), 404
        item['_id'] = str(item['_id'])
        return jsonify(item)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/get_bookings')
def get_bookings():
    """Return calendar bookings for the current user session."""
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401

    client = None
    try:
        username = session.get('username')
        start = request.args.get('start')
        end = request.args.get('end')

        bookings = au.get_ausleihungen(status=['planned', 'active', 'completed'], start=start, end=end)

        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']

        result = []
        for booking in bookings:
            start_dt = booking.get('Start')
            if not start_dt:
                continue

            end_dt = booking.get('End')
            if not end_dt and isinstance(start_dt, datetime.datetime):
                end_dt = start_dt + datetime.timedelta(minutes=45)
            elif not end_dt:
                end_dt = start_dt

            item_id = str(booking.get('Item') or '')
            item_doc = None
            if item_id:
                try:
                    item_doc = items_col.find_one({'_id': ObjectId(item_id)})
                except Exception:
                    item_doc = None

            item_name = item_id or 'Ausleihe'
            item_borrower = ''
            if item_doc:
                item_name = item_doc.get('Name') or item_doc.get('Code_4') or item_name
                borrower_info = item_doc.get('BorrowerInfo') or {}
                borrower_name = borrower_info.get('User') if isinstance(borrower_info, dict) else ''
                item_borrower = str(item_doc.get('User') or borrower_name or '')

            status = booking.get('Status') or 'unknown'
            if status == 'active':
                status = 'current'

            period = booking.get('Period')
            title = item_name
            if period:
                title = f"{title} - {period}. Std"

            result.append({
                'id': str(booking.get('_id')),
                'title': title,
                'start': start_dt.isoformat() if isinstance(start_dt, datetime.datetime) else str(start_dt),
                'end': end_dt.isoformat() if isinstance(end_dt, datetime.datetime) else str(end_dt),
                'status': status,
                'itemId': item_id,
                'userName': str(booking.get('User') or ''),
                'notes': str(booking.get('Notes') or ''),
                'period': period,
                'isCurrentUser': str(booking.get('User') or '') == username,
                'itemBorrower': item_borrower,
            })

        return jsonify({'ok': True, 'bookings': result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'bookings': []}), 500
    finally:
        if client:
            client.close()


@app.route('/api/booking_conflicts')
def api_booking_conflicts():
    """
    Returns all active bookings that have a detected conflict
    (i.e. the item was already borrowed when the planned booking activated).
    Regular users see only their own conflicts; admins see all.
    """
    if 'username' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    try:
        is_admin = us.check_admin(session['username'])
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']

        query = {'ConflictDetected': True, 'Status': 'active'}
        if not is_admin:
            query['User'] = session['username']

        conflicts = list(ausleihungen.find(query))
        result = []
        for c in conflicts:
            item_doc = it.get_item(c.get('Item'))
            item_name = item_doc.get('Name', c.get('Item', '?')) if item_doc else c.get('Item', '?')
            conflict_at = c.get('ConflictAt')
            if isinstance(conflict_at, datetime.datetime):
                conflict_at = conflict_at.strftime('%Y-%m-%d %H:%M')
            result.append({
                'id': str(c['_id']),
                'Item': item_name,
                'User': c.get('User', '?'),
                'ConflictNote': c.get('ConflictNote', ''),
                'ConflictAt': conflict_at,
                'Status': c.get('Status', ''),
            })
        client.close()
        return jsonify({'conflicts': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e), 'conflicts': []}), 500

"""Favorites management endpoints (persistent + session cache)."""
def _ensure_session_favs():
    username = session.get('username')
    owner = session.get('favorites_owner')

    if not username:
        if 'favorites' not in session or not isinstance(session.get('favorites'), list):
            session['favorites'] = []
        return

    if owner != username:
        session['favorites_owner'] = username
        session['favorites'] = []
        session.modified = True
        return

    if 'favorites' not in session or not isinstance(session.get('favorites'), list):
        session['favorites'] = []
        session.modified = True

@app.route('/favorites', methods=['GET'])
def list_favorites():
    _ensure_session_favs()
    username = session.get('username')
    if username:
        try:
            db_favs = set(us.get_favorites(username))
            merged = list(db_favs.union(set(session['favorites'])))
            session['favorites'] = merged
        except Exception as e:
            app.logger.warning(f"Listing favorites merge failed: {e}")
    return jsonify({'ok': True, 'favorites': session['favorites']})

@app.route('/favorites/<item_id>', methods=['POST'])
def add_fav(item_id):
    _ensure_session_favs()
    if item_id not in session['favorites']:
        session['favorites'].append(item_id)
    username = session.get('username')
    if username:
        try:
            us.add_favorite(username, item_id)
        except Exception as e:
            app.logger.warning(f"Persist add favorite failed: {e}")
    session.modified = True
    return jsonify({'ok': True, 'favorites': session['favorites']})

@app.route('/favorites/<item_id>', methods=['DELETE'])
def remove_fav(item_id):
    _ensure_session_favs()
    session['favorites'] = [f for f in session['favorites'] if f != item_id]
    username = session.get('username')
    if username:
        try:
            us.remove_favorite(username, item_id)
        except Exception as e:
            app.logger.warning(f"Persist remove favorite failed: {e}")
    session.modified = True
    return jsonify({'ok': True, 'favorites': session['favorites']})

@app.route('/favorites/toggle/<item_id>', methods=['POST'])
def toggle_fav(item_id):
    _ensure_session_favs()

    session_favs = [str(f) for f in session.get('favorites', [])]
    item_id = str(item_id)
    is_favorite = item_id in session_favs

    username = session.get('username')

    if is_favorite:
        session_favs = [f for f in session_favs if f != item_id]
        if username:
            try:
                us.remove_favorite(username, item_id)
            except Exception as e:
                app.logger.warning(f"Persist toggle(remove) favorite failed: {e}")
    else:
        session_favs.append(item_id)
        if username:
            try:
                us.add_favorite(username, item_id)
            except Exception as e:
                app.logger.warning(f"Persist toggle(add) favorite failed: {e}")

    # Normalize and de-duplicate while preserving order.
    deduped = []
    seen = set()
    for fav in session_favs:
        if fav not in seen:
            seen.add(fav)
            deduped.append(fav)

    session['favorites'] = deduped
    session.modified = True

    return jsonify({
        'ok': True,
        'is_favorite': item_id in deduped,
        'favorites': deduped
    })

@app.route('/debug/favorites')
def debug_favorites():
    """Diagnostic endpoint: shows session favorites, DB favorites and merged output."""
    _ensure_session_favs()
    username = session.get('username')
    session_favs = list(session.get('favorites', []))
    db_favs = []
    if username:
        try:
            db_favs = us.get_favorites(username)
        except Exception as e:
            return jsonify({'ok': False, 'error': f'db_error: {e}', 'session': session_favs})
    merged = sorted(set(session_favs) | set(db_favs))
    return jsonify({'ok': True, 'user': username, 'session': session_favs, 'db': db_favs, 'merged': merged})


@app.route('/upload_item', methods=['POST'])
def upload_item():
    """
    Route for adding new items to the inventory.
    Handles file uploads and creates QR codes.
    Enhanced for mobile browser compatibility.
    
    Returns:
        flask.Response: Redirect to admin homepage or JSON response
    """
    # Check if the user is authenticated
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Nicht angemeldet'}), 401
        
    # Check if user may insert items
    username = session['username']
    permissions = _get_current_user_permissions() or us.build_default_permission_payload('standard_user')
    if not _action_access_allowed(permissions, 'can_insert'):
        return jsonify({'success': False, 'message': 'Einfüge-Rechte erforderlich'}), 403
        
    can_access_admin_home = _page_access_allowed(permissions, 'home_admin') and _action_access_allowed(permissions, 'can_manage_settings')
    success_redirect_endpoint = 'home_admin' if can_access_admin_home else 'home'

    # Detect if request is from mobile device
    is_mobile = 'Mobile' in request.headers.get('User-Agent', '')
    is_ios = 'iPhone' in request.headers.get('User-Agent', '') or 'iPad' in request.headers.get('User-Agent', '')
    
    # Log mobile request for debugging
    if is_mobile:
        app.logger.info(f"Mobile upload from {request.headers.get('User-Agent', 'unknown')} by {username}")
    
    try:
        # Strip whitespace from all text fields
        name = sanitize_form_value(request.form['name'])
        ort = sanitize_form_value(request.form['ort'])
        beschreibung = sanitize_form_value(request.form['beschreibung'])
        
        # Check both possible image field names
        images = request.files.getlist('images') or request.files.getlist('new_images')
        
        filter_upload = sanitize_form_value(request.form.getlist('filter'))
        filter_upload2 = sanitize_form_value(request.form.getlist('filter2'))
        filter_upload3 = sanitize_form_value(request.form.getlist('filter3'))
        anschaffungs_jahr = sanitize_form_value(request.form.getlist('anschaffungsjahr'))
        anschaffungs_kosten = sanitize_form_value(request.form.getlist('anschaffungskosten'))
        code_4 = sanitize_form_value(request.form.getlist('code_4'))
        isbn_raw = sanitize_form_value(request.form.get('isbn', ''))
        upload_mode = sanitize_form_value(request.form.get('upload_mode', 'item'))
        individual_codes_raw = sanitize_form_value(request.form.get('individual_codes', ''))
        item_count_raw = sanitize_form_value(request.form.get('item_count', '1'))

        try:
            item_count = int(item_count_raw) if item_count_raw else 1
        except (TypeError, ValueError):
            item_count = 1
        item_count = max(1, min(item_count, 100))

        # Optional list of per-item codes (one code per line)
        individual_codes = []
        if individual_codes_raw:
            individual_codes = [c.strip() for c in str(individual_codes_raw).replace(',', '\n').splitlines() if c.strip()]
        
        # Check if this is a duplication
        is_duplicating = request.form.get('is_duplicating') == 'true'
        
        # Get duplicate_images if duplicating
        duplicate_images = request.form.getlist('duplicate_images') if is_duplicating else []
        print(f"DEBUG: Duplicate images from form: {duplicate_images}, count: {len(duplicate_images)}")
        
        # Make sure duplicate_images is always a list, even if there's only one
        if is_duplicating and duplicate_images and not isinstance(duplicate_images, list):
            duplicate_images = [duplicate_images]
        
        # Log details about each image
        if is_duplicating and duplicate_images:
            for i, img in enumerate(duplicate_images):
                print(f"DEBUG: Duplicate image {i+1}/{len(duplicate_images)}: {img}")
        
        # Get book cover image if downloaded
        book_cover_image = request.form.get('book_cover_image')
        
        # Special handling for mobile browsers that might send data differently
        if is_mobile and 'mobile_data' in request.form:
            try:
                mobile_data = json.loads(request.form['mobile_data'])
                # Override values with mobile data if available
                if 'filters' in mobile_data:
                    filter_upload = mobile_data.get('filters', [])
                if 'filters2' in mobile_data:
                    filter_upload2 = mobile_data.get('filters2', [])
                if 'filters3' in mobile_data:
                    filter_upload3 = mobile_data.get('filters3', [])
                if 'duplicate_images' in mobile_data and mobile_data['duplicate_images']:
                    duplicate_images = mobile_data.get('duplicate_images', [])
            except json.JSONDecodeError as e:
                app.logger.error(f"Error parsing mobile data: {str(e)}")
    except Exception as e:
        error_msg = f"Fehler beim Verarbeiten der Formulardaten: {str(e)}"
        app.logger.error(error_msg)
        if is_mobile:
            return jsonify({'success': False, 'message': error_msg}), 400
        else:
            flash('Fehler beim Verarbeiten der Formulardaten. Bitte versuchen Sie es erneut.', 'error')
            return redirect(url_for(success_redirect_endpoint))

    # Expand special "all values" selections for predefined filters.
    filter_upload = expand_filter_selection(filter_upload, 1)
    filter_upload2 = expand_filter_selection(filter_upload2, 2)

    # Validation
    if not name or not ort or not beschreibung:
        error_msg = 'Bitte füllen Sie alle erforderlichen Felder aus'
        if is_mobile:
            return jsonify({'success': False, 'message': error_msg}), 400
        else:
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))

    item_isbn = ''
    item_type = 'general'
    if cfg.MODULES.is_enabled('library'):
        item_isbn = normalize_and_validate_isbn(isbn_raw)
        if isbn_raw and not item_isbn:
            error_msg = 'Ungültige ISBN. Bitte ISBN-10 oder ISBN-13 verwenden.'
            if is_mobile:
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))
        if item_isbn:
            item_type = 'book'

    if upload_mode == 'library':
        if not cfg.MODULES.is_enabled('library'):
            error_msg = 'Bibliotheks-Modul ist deaktiviert.'
            if is_mobile:
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))
        if not item_isbn:
            error_msg = 'Für Bücher ist eine gültige ISBN erforderlich.'
            if is_mobile:
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for('library_admin'))
        item_type = 'book'

    # Only check for images if not duplicating and no duplicate images provided and no book cover
    # For library mode, skip this check as images come only from ISBN fetch
    if upload_mode != 'library' and not is_duplicating and not images and not duplicate_images and not book_cover_image:
        error_msg = 'Bitte laden Sie mindestens ein Bild hoch'
        if is_mobile:
            return jsonify({'success': False, 'message': error_msg}), 400
        else:
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))

    # Check if base code is unique for single-item uploads
    if code_4 and item_count == 1 and not it.is_code_unique(code_4[0]):
        error_msg = 'Der Code wird bereits verwendet. Bitte wählen Sie einen anderen Code.'
        if is_mobile:
            return jsonify({'success': False, 'message': error_msg}), 400
        else:
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))

    # Validate optional per-item codes
    if individual_codes:
        if len(individual_codes) > item_count:
            error_msg = f'Zu viele Einzelcodes angegeben ({len(individual_codes)}), erlaubt sind maximal {item_count}.'
            if is_mobile:
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))

        if len(set(individual_codes)) != len(individual_codes):
            error_msg = 'Doppelte Einzelcodes erkannt. Bitte alle Codes eindeutig eintragen.'
            if is_mobile:
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))

        for specific_code in individual_codes:
            if not it.is_code_unique(specific_code):
                error_msg = f'Der Einzelcode "{specific_code}" wird bereits verwendet.'
                if is_mobile:
                    return jsonify({'success': False, 'message': error_msg}), 400
                flash(error_msg, 'error')
                return redirect(url_for(success_redirect_endpoint))

    def generate_unique_batch_code(base_code, position):
        """Generate a unique code for every item in a batch."""
        if not base_code:
            return None

        candidate = base_code if position == 1 else f"{base_code}-{position}"
        if it.is_code_unique(candidate):
            return candidate

        suffix = 1
        while suffix <= 1000:
            alternative = f"{candidate}-{suffix}"
            if it.is_code_unique(alternative):
                return alternative
            suffix += 1
        return None

    # Process any new uploaded images with robust error handling
    image_filenames = []
    processed_count = 0
    error_count = 0
    skipped_count = 0
    
    # Create a structured log entry for upload session
    upload_session_id = str(uuid.uuid4())[:8]
    app.logger.info(f"Starting image upload session {upload_session_id} - Files: {len(images)}, User: {username}")
    
    # Ensure all required directories exist
    for directory in [app.config['UPLOAD_FOLDER']]:
        try:
            os.makedirs(directory, exist_ok=True)
        except Exception as e:
            app.logger.error(f"Failed to create directory {directory}: {str(e)}")
    
    # Process each image independently
    for index, image in enumerate(images):
        # In library mode, skip manual image uploads (use only book_cover_image from ISBN fetch)
        if upload_mode == 'library':
            app.logger.info(f"[Library Mode] Skipping manual image upload {index+1}/{len(images)}")
            skipped_count += 1
            continue
        
        image_log_prefix = f"[Upload {upload_session_id}][Image {index+1}/{len(images)}]"
        
        if not image or not image.filename or image.filename == '':
            app.logger.warning(f"{image_log_prefix} Empty file or filename")
            skipped_count += 1
            continue
            
        # Get file extension for special handling
        _, file_ext = os.path.splitext(image.filename.lower())
        is_png = file_ext.lower() == '.png'
        
        if is_png:
            app.logger.info(f"PNG DEBUG: {image_log_prefix} Detected PNG file: {image.filename}")
            # Check file size
            image.seek(0, os.SEEK_END)
            file_size = image.tell() / (1024 * 1024)  # Size in MB
            image.seek(0)  # Reset file pointer
            app.logger.info(f"PNG DEBUG: {image_log_prefix} PNG file size: {file_size:.2f}MB")
            
            # Check first few bytes for PNG signature and analyze header
            header_bytes = image.read(64)  # Read more for thorough analysis
            image.seek(0)  # Reset pointer
            png_signature = b'\x89PNG\r\n\x1a\n'
            is_valid_signature = header_bytes.startswith(png_signature)
            
            # Create a hex dump of header for debugging
            hex_dump = ' '.join([f"{b:02x}" for b in header_bytes[:32]])
            app.logger.info(f"PNG DEBUG: {image_log_prefix} PNG header hex: {hex_dump}")
            app.logger.info(f"PNG DEBUG: {image_log_prefix} PNG signature valid: {is_valid_signature}, bytes: {header_bytes[:8]!r}")
            
            # Analyze PNG chunks if signature is valid
            if is_valid_signature:
                try:
                    # Look for IHDR chunk that should follow the signature
                    if header_bytes[8:12] == b'IHDR':
                        # Extract width and height from IHDR chunk (bytes 16-23)
                        import struct
                        width = struct.unpack('>I', header_bytes[16:20])[0]
                        height = struct.unpack('>I', header_bytes[20:24])[0]
                        bit_depth = header_bytes[24]
                        color_type = header_bytes[25]
                        app.logger.info(f"PNG DEBUG: {image_log_prefix} PNG dimensions from header: {width}x{height}, bit depth: {bit_depth}, color type: {color_type}")
                    else:
                        app.logger.warning(f"PNG DEBUG: {image_log_prefix} Expected IHDR chunk not found. Found: {header_bytes[8:12]!r}")
                except Exception as chunk_err:
                    app.logger.error(f"PNG DEBUG: {image_log_prefix} Error analyzing PNG chunks: {str(chunk_err)}")
            else:
                app.logger.error(f"PNG DEBUG: {image_log_prefix} Invalid PNG signature!")
        
        app.logger.info(f"{image_log_prefix} Processing: {image.filename}")
        
        try:
            # Comprehensive file validation with detailed logging
            is_allowed, error_message = allowed_file(image.filename, image, max_size_mb=cfg.IMAGE_MAX_UPLOAD_MB)
            
            if not is_allowed:
                app.logger.warning(f"{image_log_prefix} Validation failed: {error_message}")
                if is_png:
                    app.logger.error(f"PNG DEBUG: {image_log_prefix} PNG validation failed: {error_message}")
                skipped_count += 1
                if not is_mobile:
                    flash(error_message, 'error')
                continue
                
            # Get the file extension for content type determination
            secure_name = secure_filename(image.filename)
            _, ext_part = os.path.splitext(secure_name)
            is_png = ext_part.lower() == '.png'
            
            # Generate a completely unique filename using UUID
            unique_id = str(uuid.uuid4())
            timestamp = time.strftime("%Y%m%d%H%M%S")
            
            # New filename format with UUID to ensure uniqueness
            saved_filename = f"{unique_id}_{timestamp}{ext_part}"
            app.logger.info(f"{image_log_prefix} Assigned unique filename: {saved_filename}")
            
            if is_png:
                app.logger.info(f"PNG DEBUG: {image_log_prefix} Creating PNG with filename: {saved_filename}")
            
            # For iOS devices, we need special handling for the file save
            if is_ios:
                app.logger.info(f"{image_log_prefix} Using iOS-specific file handling")
                # Save to a temporary file first to avoid iOS stream issues
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f"temp_{saved_filename}")
                
                try:
                    if is_png:
                        app.logger.info(f"PNG DEBUG: {image_log_prefix} Using iOS PNG save method")
                        # Before saving, verify the file content again
                        try:
                            image.seek(0)
                            pre_save_data = image.read(16)
                            image.seek(0)
                            pre_save_hex = ' '.join([f"{b:02x}" for b in pre_save_data])
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Pre-save data: {pre_save_hex}")
                        except Exception as pre_err:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Error checking pre-save data: {str(pre_err)}")
                    
                    # For PNGs, try a direct binary save first
                    if is_png:
                        try:
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Attempting binary save for iOS PNG")
                            image.seek(0)
                            png_data = image.read()
                            with open(temp_path, 'wb') as f:
                                f.write(png_data)
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Binary write complete, size: {len(png_data)} bytes")
                        except Exception as bin_err:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Binary write failed: {str(bin_err)}")
                            # Fall back to normal save
                            image.seek(0)
                            image.save(temp_path)
                    else:
                        image.save(temp_path)
                    
                    # Validate the saved file
                    if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
                        if is_png:
                            file_size = os.path.getsize(temp_path)
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Temp PNG file saved successfully: {file_size/1024:.1f}KB")
                            
                            # Verify it's a valid PNG
                            try:
                                with open(temp_path, 'rb') as f:
                                    png_header = f.read(16)
                                    png_signature = b'\x89PNG\r\n\x1a\n'
                                    is_valid = png_header.startswith(png_signature)
                                    
                                    header_hex = ' '.join([f"{b:02x}" for b in png_header])
                                    app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved file header: {header_hex}")
                                    
                                    if not is_valid:
                                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Invalid PNG signature in saved file!")
                                    else:
                                        app.logger.info(f"PNG DEBUG: {image_log_prefix} Valid PNG signature confirmed in saved file")
                                        
                                    # Try opening with PIL to confirm it's valid
                                    try:
                                        with Image.open(temp_path) as img:
                                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved PNG validates with PIL: {img.format} {img.size}")
                                    except Exception as pil_err:
                                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Saved PNG fails PIL validation: {str(pil_err)}")
                                        
                            except Exception as verify_err:
                                app.logger.error(f"PNG DEBUG: {image_log_prefix} Error verifying PNG: {str(verify_err)}")
                        
                        # Rename to the final filename
                        final_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
                        os.rename(temp_path, final_path)
                        app.logger.info(f"{image_log_prefix} Successfully saved via iOS handler: {os.path.getsize(final_path)/1024:.1f}KB")
                    else:
                        if is_png:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Failed to save temp PNG file (zero size or missing)")
                        raise Exception("Failed to save image file (zero size or missing)")
                except Exception as e:
                    app.logger.error(f"{image_log_prefix} iOS save failed: {str(e)}")
                    if is_png:
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} iOS PNG save failed: {str(e)}")
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Error type: {type(e).__name__}")
                        # Log full traceback for PNG errors
                        import io
                        tb_output = io.StringIO()
                        traceback.print_exc(file=tb_output)
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Full traceback:\n{tb_output.getvalue()}")
                    
                    # Try regular save as fallback
                    try:
                        image.seek(0)  # Reset file pointer
                        if is_png:
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Attempting fallback PNG save method")
                        
                        image.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_filename))
                        app.logger.info(f"{image_log_prefix} Fallback save successful")
                        
                        if is_png:
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Fallback PNG save successful")
                    except Exception as fallback_err:
                        app.logger.error(f"{image_log_prefix} Fallback save also failed: {str(fallback_err)}")
                        if is_png:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Fallback PNG save also failed: {str(fallback_err)}")
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Error type: {type(fallback_err).__name__}")
                            traceback.print_exc()
                        error_count += 1
                        continue
            else:
                # Regular file save for non-iOS devices
                try:
                    if is_png:
                        app.logger.info(f"PNG DEBUG: {image_log_prefix} Using standard PNG save method")
                        
                        # Check file content before saving
                        try:
                            image.seek(0)
                            pre_save_data = image.read(16)
                            image.seek(0)
                            pre_save_hex = ' '.join([f"{b:02x}" for b in pre_save_data])
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Pre-save data: {pre_save_hex}")
                            
                            # Check if it's really a PNG
                            png_signature = b'\x89PNG\r\n\x1a\n'
                            if not pre_save_data.startswith(png_signature):
                                app.logger.error(f"PNG DEBUG: {image_log_prefix} File does not have valid PNG signature before save!")
                        except Exception as pre_err:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Error checking pre-save data: {str(pre_err)}")
                        
                        # Try an alternative saving method for PNGs with detailed error tracking
                        try:
                            # Read the image data directly
                            image.seek(0)
                            image_data = image.read()
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Read {len(image_data)} bytes of PNG data")
                            
                            # Write it manually to file
                            save_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
                            with open(save_path, 'wb') as f:
                                f.write(image_data)
                            
                            file_size = os.path.getsize(save_path)
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Direct binary PNG write successful: {file_size/1024:.1f}KB")
                            
                            # Verify the saved file
                            try:
                                with open(save_path, 'rb') as f:
                                    saved_header = f.read(16)
                                    header_hex = ' '.join([f"{b:02x}" for b in saved_header])
                                    app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved PNG header: {header_hex}")
                                    
                                    is_valid = saved_header.startswith(png_signature)
                                    app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved PNG has valid signature: {is_valid}")
                                    
                                    if is_valid:
                                        # Additional validation with PIL
                                        try:
                                            with Image.open(save_path) as img:
                                                app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved PNG validates with PIL: {img.format} {img.size}")
                                        except Exception as pil_err:
                                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Saved PNG fails PIL validation: {str(pil_err)}")
                            except Exception as verify_err:
                                app.logger.error(f"PNG DEBUG: {image_log_prefix} Error verifying saved PNG: {str(verify_err)}")
                            
                        except Exception as png_write_err:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Direct PNG write failed: {str(png_write_err)}")
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Error type: {type(png_write_err).__name__}")
                            
                            # Log traceback for PNG errors
                            import io
                            tb_output = io.StringIO()
                            traceback.print_exc(file=tb_output)
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Binary write traceback:\n{tb_output.getvalue()}")
                            
                            # Fall back to standard method
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Attempting standard PIL save method")
                            image.seek(0)
                            image.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_filename))
                    else:
                        # Standard save for non-PNG files
                        image.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_filename))
                    
                    app.logger.info(f"{image_log_prefix} Standard save successful")
                except Exception as save_err:
                    app.logger.error(f"{image_log_prefix} Failed to save file: {str(save_err)}")
                    if is_png:
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Standard PNG save failed: {str(save_err)}")
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Error type: {type(save_err).__name__}")
                        
                        # Log traceback for PNG errors
                        import io
                        tb_output = io.StringIO()
                        traceback.print_exc(file=tb_output)
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Standard save traceback:\n{tb_output.getvalue()}")
                        
                        # Try one last method - save as a different format
                        try:
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Attempting last resort conversion to WebP")
                            image.seek(0)
                            
                            # Try to convert PNG to WebP as a last resort
                            with Image.open(image) as img:
                                # Ensure RGBA for transparency
                                if img.mode != 'RGBA':
                                    img = img.convert('RGBA')
                                webp_path = os.path.splitext(os.path.join(app.config['UPLOAD_FOLDER'], saved_filename))[0] + '.webp'
                                img.save(webp_path, 'WEBP')
                                app.logger.info(f"PNG DEBUG: {image_log_prefix} Successfully saved as WebP instead: {webp_path}")
                                # Update the saved_filename to reflect the new extension
                                saved_filename = os.path.basename(webp_path)
                        except Exception as webp_err:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} Final WebP conversion failed: {str(webp_err)}")
                            
                        traceback.print_exc()
                    error_count += 1
                    continue
            
            # Verify the file was saved correctly
            saved_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
            if not os.path.exists(saved_path) or os.path.getsize(saved_path) == 0:
                app.logger.error(f"{image_log_prefix} Saved file is missing or empty: {saved_path}")
                if is_png:
                    app.logger.error(f"PNG DEBUG: {image_log_prefix} Saved PNG is missing or empty: {saved_path}")
                error_count += 1
                continue
            
            # Special verification for PNG files
            if is_png:
                try:
                    app.logger.info(f"PNG DEBUG: {image_log_prefix} Verifying saved PNG: {saved_path}")
                    saved_size = os.path.getsize(saved_path) / 1024.0  # in KB
                    app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved PNG size: {saved_size:.1f}KB")
                    
                    # Check file integrity by trying to open it
                    try:
                        with Image.open(saved_path) as img:
                            png_width, png_height = img.size
                            png_mode = img.mode
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Saved PNG valid: {png_width}x{png_height}, mode: {png_mode}")
                    except Exception as png_verify_err:
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} PNG verification failed: {str(png_verify_err)}")
                        
                        # Try to fix by copying from the original
                        try:
                            image.seek(0)
                            with open(saved_path, 'wb') as f:
                                f.write(image.read())
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Attempted to fix PNG by direct copy")
                        except Exception as png_fix_err:
                            app.logger.error(f"PNG DEBUG: {image_log_prefix} PNG fix attempt failed: {str(png_fix_err)}")
                except Exception as e:
                    app.logger.error(f"PNG DEBUG: {image_log_prefix} PNG verification error: {str(e)}")
            
            # Generate optimized versions (thumbnails and previews)
            optimization_success = False
            try:
                # Log original file size before optimization
                original_path = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
                original_size = os.path.getsize(original_path)
                
                # Get original image dimensions
                original_dimensions = "unknown"
                try:
                    with Image.open(original_path) as img:
                        original_dimensions = f"{img.width}x{img.height}"
                        if is_png:
                            app.logger.info(f"PNG DEBUG: {image_log_prefix} Original PNG dimensions: {original_dimensions}, mode: {img.mode}")
                except Exception as dim_err:
                    app.logger.warning(f"{image_log_prefix} Could not get image dimensions: {str(dim_err)}")
                    if is_png:
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Could not get PNG dimensions: {str(dim_err)}")
                        app.logger.error(f"PNG DEBUG: {image_log_prefix} Error type: {type(dim_err).__name__}")
                        traceback.print_exc()
                
                app.logger.info(f"{image_log_prefix} Starting optimization for {saved_filename} ({original_size/1024:.1f}KB, {original_dimensions})")
                
                # PNG-specific optimization options
                if is_png:
                    app.logger.info(f"PNG DEBUG: {image_log_prefix} Starting PNG optimization")
                    # For PNGs, we might need different parameters
                    optimization_result = generate_optimized_versions(
                        saved_filename, 
                        max_original_width=500, 
                        target_size_kb=100,  # Higher target for PNGs to maintain transparency
                        debug_prefix=f"PNG DEBUG: {image_log_prefix}"
                    )
                else:
                    # Standard optimization for non-PNG images
                    optimization_result = generate_optimized_versions(saved_filename, max_original_width=500, target_size_kb=80)
                
                # Log file size after optimization
                if optimization_result['success'] and optimization_result['original']:
                    optimized_name = optimization_result['original']
                    optimized_path = os.path.join(app.config['UPLOAD_FOLDER'], optimized_name)
                    
                    if os.path.exists(optimized_path):
                        optimized_size = os.path.getsize(optimized_path)
                        reduction = (1 - (optimized_size / original_size)) * 100 if original_size > 0 else 0
                        
                        # Get optimized dimensions
                        optimized_dimensions = "unknown"
                        try:
                            with Image.open(optimized_path) as img:
                                optimized_dimensions = f"{img.width}x{img.height}"
                        except Exception as dim_err:
                            app.logger.warning(f"{image_log_prefix} Could not get optimized dimensions: {str(dim_err)}")
                            
                        app.logger.info(
                            f"{image_log_prefix} Optimization results:\n"
                            f"  File: {saved_filename} → {optimized_name}\n"
                            f"  Size: {original_size/1024:.1f}KB → {optimized_size/1024:.1f}KB ({reduction:.1f}% reduction)\n"
                            f"  Dimensions: {original_dimensions} → {optimized_dimensions}"
                        )
                        
                        # Use the optimized filename
                        saved_filename = optimized_name
                    else:
                        app.logger.warning(f"{image_log_prefix} Optimized file reported success but not found: {optimized_path}")
                else:
                    app.logger.warning(f"{image_log_prefix} Optimization failed or returned no file")
            except Exception as e:
                app.logger.error(f"{image_log_prefix} Optimization failed: {str(e)}")
                traceback.print_exc()
                
                # No fallback thumbnail generation needed as we only use the main image
            
            # Always add the filename to our list even if optimization failed
            # We'll use the original in that case
            image_filenames.append(saved_filename)
            processed_count += 1
            app.logger.info(f"{image_log_prefix} Successfully processed")
            
        except Exception as e:
            app.logger.error(f"{image_log_prefix} Unexpected error: {str(e)}")
            traceback.print_exc()
            error_count += 1
            # Continue with the next image
    
    # Log summary of upload session
    app.logger.info(f"Upload session {upload_session_id} completed: {processed_count} processed, {error_count} errors, {skipped_count} skipped")

    # Handle duplicate images if duplicating
    if duplicate_images:
        app.logger.info(f"Processing {len(duplicate_images)} duplicate images: {duplicate_images}")
        
        # For mobile browsers, we need to verify the duplicate images exist first
        verified_duplicates = []
        for dup_img in duplicate_images:
            # Try looking in different paths
            # Add all possible paths where images might be stored
            dev_upload_path = app.config['UPLOAD_FOLDER']
            prod_upload_path = '/var/Inventarsystem/Web/uploads'
            
            # Also look for image variations with suffixes that might be in the path
            name_part, ext_part = os.path.splitext(dup_img)
            possible_filenames = [
                dup_img,
                f"{name_part}.webp", # Check WebP first
                f"{name_part}.jpg",  # In case it was converted to JPG
                f"{name_part}.png",  # In case it was saved as PNG
            ]
            
            possible_paths = []
            for filename in possible_filenames:
                possible_paths.extend([
                    os.path.join(dev_upload_path, filename),  # Development upload path
                    os.path.join(prod_upload_path, filename),  # Production upload path
                ])
            
            app.logger.info(f"Looking for duplicate image {dup_img} in paths: {possible_paths}")
            
            # Try to find the original image
            found = False
            for path in possible_paths:
                if os.path.exists(path) and os.path.isfile(path):
                    verified_duplicates.append((dup_img, path))
                    app.logger.info(f"Found duplicate image at: {path}")
                    found = True
                    break
            
            if not found:
                app.logger.warning(f"Duplicate image not found: {dup_img}")
                # Try to find any image with a similar filename (removing size or resolution parts)
                # This handles cases where the filename may have variations like "_800" suffix
                base_name = os.path.splitext(dup_img)[0]
                base_name = re.sub(r'_\d+$', '', base_name)  # Remove trailing _NUMBER
                
                if len(base_name) > 5:  # Only if we have a meaningful base name
                    app.logger.info(f"Trying to find similar images with base name: {base_name}")
                    
                    # Search in development directory
                    dev_files = os.listdir(app.config['UPLOAD_FOLDER']) if os.path.exists(app.config['UPLOAD_FOLDER']) else []
                    # Search in production directory
                    prod_path = "/var/Inventarsystem/Web/uploads"
                    prod_files = os.listdir(prod_path) if os.path.exists(prod_path) else []
                    
                    # Combine all files
                    all_files = dev_files + prod_files
                    
                    # Find similar files
                    for f in all_files:
                        if base_name in f:
                            img_path = os.path.join(app.config['UPLOAD_FOLDER'], f)
                            if not os.path.exists(img_path):
                                img_path = os.path.join(prod_path, f)
                            
                            if os.path.exists(img_path) and os.path.isfile(img_path):
                                app.logger.info(f"Found similar image: {f} at {img_path}")
                                verified_duplicates.append((f, img_path))
                                found = True
                                break
                
                # If we still can't find anything, just use a placeholder
                if not found:
                    app.logger.warning(f"Could not find any similar image for: {dup_img}, will use placeholder")
        
        # Create copies of verified images with new unique filenames
        duplicate_image_copies = []
        
        # Create a placeholder name for each image that wasn't found
        placeholder_used = False
        original_count = len(duplicate_images)
        
        # Process each original image - either use found file or placeholder
        for i, dup_img in enumerate(duplicate_images):
            # Look for corresponding verified image
            found_image = None
            for verified_img, src_path in verified_duplicates:
                if verified_img == dup_img:
                    found_image = (verified_img, src_path)
                    break
            
            try:
                # Generate a new unique filename (same for real or placeholder)
                unique_id = str(uuid.uuid4())
                timestamp = time.strftime("%Y%m%d%H%M%S")
                _, ext_part = os.path.splitext(dup_img) if dup_img else '.jpg'
                new_filename = f"{unique_id}_{timestamp}{ext_part}"
                
                # If we found the image, copy it
                if found_image:
                    dup_img, src_path = found_image
                    app.logger.info(f"Copying image {i+1}/{original_count} from {src_path} to {new_filename}")
                    
                    # Copy the image file to the new name
                    dst_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
                    
                    # Make sure the target directory exists
                    os.makedirs(os.path.dirname(dst_path), exist_ok=True)
                    
                    # Copy the file
                    shutil.copy2(src_path, dst_path)
                    
                    # Verify the file was copied successfully
                    if os.path.exists(dst_path):
                        app.logger.info(f"Successfully copied image to {dst_path}")
                    else:
                        app.logger.error(f"Failed to copy image to {dst_path}")
                        # If copy fails, use placeholder
                        raise Exception("Copy failed - will use placeholder")
                    
                    # Generate optimized versions (thumbnails and previews) for the new copy
                    try:
                        result = generate_optimized_versions(new_filename, max_original_width=500, target_size_kb=80)
                        app.logger.info(f"Generated optimized versions: {result}")
                        if result['success'] and result['original']:
                            new_filename = result['original']
                    except Exception as e:
                        app.logger.error(f"Error generating optimized versions for {new_filename}: {e}")
                        # If optimization fails, at least keep the original file
                        result = {'original': new_filename}
                        traceback.print_exc()
                
                # If we didn't find the image, use a placeholder
                else:
                    app.logger.warning(f"Using placeholder for image {i+1}/{original_count} (original: {dup_img})")
                    
                    # Copy placeholder to uploads directory with the new filename
                    placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
                    if not os.path.exists(placeholder_path):
                        placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
                    
                    if os.path.exists(placeholder_path):
                        dst_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
                        shutil.copy2(placeholder_path, dst_path)
                        app.logger.info(f"Copied placeholder image to {dst_path}")
                        placeholder_used = True
                        
                        # Skip the optimization step for placeholder images
                        # Just add directly to the list of image filenames
                        continue
                    else:
                        app.logger.error(f"Placeholder image not found at {placeholder_path}")
                        # Create a simple placeholder file
                        with open(os.path.join(app.config['UPLOAD_FOLDER'], new_filename), 'w') as f:
                            f.write("Placeholder")
                        placeholder_used = True
                        # Skip the optimization step
                        continue
                
                # Add the new filename to our list (either copied or placeholder)
                duplicate_image_copies.append(new_filename)
                processed_count += 1
                
                app.logger.info(f"Processed image {i+1}/{original_count}: {new_filename}")
            except Exception as e:
                app.logger.error(f"Error processing image {i+1}/{original_count} ({dup_img}): {str(e)}")
                traceback.print_exc()
                error_count += 1
        
        # Log placeholder usage
        if placeholder_used:
            app.logger.warning(f"Used placeholders for some missing images during duplication")
        
        # Log if no images were processed
        if not duplicate_image_copies:
            app.logger.warning(f"No duplicate images were processed")
            if duplicate_images:
                app.logger.warning(f"Original had {len(duplicate_images)} images, but none were copied")
        
        # Add the new image copies to our list of filenames
        image_filenames.extend(duplicate_image_copies)

    # Handle book cover image if provided
    if book_cover_image:
        # Verify the book cover image exists
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], book_cover_image)
        if os.path.exists(full_path) and os.path.isfile(full_path):
            # Create a unique filename for the book cover
            unique_id = str(uuid.uuid4())
            timestamp = time.strftime("%Y%m%d%H%M%S")
            _, ext_part = os.path.splitext(book_cover_image)
            
            new_filename = f"{unique_id}_{timestamp}_book_cover{ext_part}"
            new_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
            
            # Copy the file to the new unique name
            shutil.copy2(full_path, new_path)
            
            # Use the new filename instead
            image_filenames.append(new_filename)
            app.logger.info(f"Copied book cover from {book_cover_image} to {new_filename}")
        else:
            app.logger.warning(f"Book cover image not found: {book_cover_image}")
    
    # Log image processing stats
    app.logger.info(f"Upload stats: processed={processed_count}, errors={error_count}, skipped={skipped_count}, duplicates={len(duplicate_images) if duplicate_images else 0}")
    
    # If location is not in the predefined list, add it
    predefined_locations = it.get_predefined_locations()
    if ort and ort not in predefined_locations:
        it.add_predefined_location(ort)
    
    reservierbar = 'reservierbar' in request.form

    # Add one or more own items; sub-items are hidden in overview and counted on parent.
    created_item_ids = []
    series_group_id = str(uuid.uuid4()) if item_count > 1 else None
    base_code = code_4[0] if code_4 else None

    for position in range(1, item_count + 1):
        unique_code = None
        if position <= len(individual_codes):
            unique_code = individual_codes[position - 1]
        else:
            unique_code = generate_unique_batch_code(base_code, position)

        if (base_code or individual_codes) and not unique_code:
            error_msg = 'Fehler bei der Code-Erzeugung für mehrere Artikel.'
            if is_mobile:
                return jsonify({'success': False, 'message': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))

        parent_item_id = str(created_item_ids[0]) if created_item_ids else None
        item_id = it.add_item(
            name, ort, beschreibung, image_filenames, filter_upload,
            filter_upload2, filter_upload3,
            anschaffungs_jahr[0] if anschaffungs_jahr else None,
            anschaffungs_kosten[0] if anschaffungs_kosten else None,
            unique_code,
            reservierbar=reservierbar,
            series_group_id=series_group_id,
            series_count=item_count,
            series_position=position,
            is_grouped_sub_item=(position > 1),
            parent_item_id=parent_item_id,
            isbn=item_isbn,
            item_type=item_type
        )
        if not item_id:
            break
        created_item_ids.append(item_id)

    if len(created_item_ids) != item_count:
        error_msg = f'Nur {len(created_item_ids)} von {item_count} Artikeln konnten erstellt werden.'
        if is_mobile:
            return jsonify({'success': False, 'message': error_msg}), 500
        flash(error_msg, 'error')
        return redirect(url_for(success_redirect_endpoint))

    item_id = created_item_ids[0] if created_item_ids else None
    
    if item_id:
    # Create QR code for the item (deactivated)
    # create_qr_code(str(item_id))
        success_msg = f'Element wurde erfolgreich hinzugefügt ({len(created_item_ids)} erstellt)'
        
        if is_mobile:
            return jsonify({
                'success': True, 
                'message': success_msg,
                'itemId': str(item_id),
                'stats': {
                    'processed': processed_count,
                    'errors': error_count,
                    'skipped': skipped_count,
                    'duplicates': len(duplicate_images) if duplicate_images else 0,
                    'totalImages': len(image_filenames)
                }
            })
        else:
            flash(success_msg, 'success')
            return redirect(url_for(success_redirect_endpoint, highlight_item=str(item_id)))
    else:
        error_msg = 'Fehler beim Hinzufügen des Elements'
        if is_mobile:
            return jsonify({'success': False, 'message': error_msg}), 500
        else:
            flash(error_msg, 'error')
            return redirect(url_for(success_redirect_endpoint))


@app.route('/duplicate_item', methods=['POST'])
def duplicate_item():
    """
    Route for duplicating an existing item.
    Returns JSON response with success status.
    Enhanced for mobile browser compatibility.
    
    Returns:
        flask.Response: JSON response with success status and data
    """
    try:
        # Check authentication
        if 'username' not in session:
            return jsonify({'success': False, 'message': 'Nicht angemeldet'}), 401
        
        # Check if user is admin
        username = session['username']
        if not us.check_admin(username):
            return jsonify({'success': False, 'message': 'Keine Administratorrechte'}), 403
        
        # Detect if request is from mobile device
        is_mobile = 'Mobile' in request.headers.get('User-Agent', '')
        is_ios = 'iPhone' in request.headers.get('User-Agent', '') or 'iPad' in request.headers.get('User-Agent', '')
        
        # Log mobile duplication for debugging
        if is_mobile:
            app.logger.info(f"Mobile duplication from {request.headers.get('User-Agent', 'unknown')} by {username}")
        
        # Get original item ID
        original_item_id = request.form.get('original_item_id')
        if not original_item_id:
            return jsonify({'success': False, 'message': 'Ursprungs-Element-ID fehlt'}), 400
        
        # Fetch original item data
        original_item = it.get_item(original_item_id)
        if not original_item:
            return jsonify({'success': False, 'message': 'Ursprungs-Element nicht gefunden'}), 404
        
        # Process filters as arrays (same as stored in database)
        filter1_array = original_item.get('Filter', [])
        filter2_array = original_item.get('Filter2', [])
        filter3_array = original_item.get('Filter3', [])
        
        # Ensure filters are arrays
        if not isinstance(filter1_array, list):
            filter1_array = [filter1_array] if filter1_array else []
        if not isinstance(filter2_array, list):
            filter2_array = [filter2_array] if filter2_array else []
        if not isinstance(filter3_array, list):
            filter3_array = [filter3_array] if filter3_array else []
            
        # Verify image paths for mobile devices to avoid issues with non-existent images
        images = original_item.get('Images', [])
        verified_images = []
        
        if is_mobile:
            for img in images:
                img_path = os.path.join(app.config['UPLOAD_FOLDER'], img)
                if os.path.exists(img_path) and os.path.isfile(img_path):
                    verified_images.append(img)
                else:
                    app.logger.warning(f"Image not found for duplication: {img}")
            
            # If we lost images in verification, log it
            if len(verified_images) < len(images):
                app.logger.warning(f"Only {len(verified_images)} of {len(images)} images verified for mobile duplication")
        else:
            verified_images = images

        # For iOS devices, add more diagnostics and reduce data size if needed
        if is_ios:
            # Check if images exist (we now use main images as thumbnails)
            images_exist = []
            for img in verified_images[:5]:  # Only check first 5 to save time
                img_path = os.path.join(app.config['UPLOAD_FOLDER'], img)
                if os.path.exists(img_path):
                    images_exist.append(True)
                else:
                    images_exist.append(False)
            
            # Log detailed diagnostics
            app.logger.info(f"iOS duplication details: {len(verified_images)} images, "
                           f"images available: {all(images_exist)}, "
                           f"filter sizes: {len(filter1_array)}, {len(filter2_array)}, {len(filter3_array)}")
                           
        return jsonify({
            'success': True, 
            'message': 'Duplizierungsdaten erfolgreich vorbereitet',
            'item_data': {
                'name': original_item.get('Name', ''),
                'description': original_item.get('Beschreibung', ''),
                'location': original_item.get('Ort', ''),
                'room': original_item.get('Raum', ''),
                'category': original_item.get('Kategorie', ''),
                'year': original_item.get('Anschaffungsjahr', ''),
                'cost': original_item.get('Anschaffungskosten', ''),
                'filter1': filter1_array,
                'filter2': filter2_array,
                'filter3': filter3_array,
                'images': verified_images,  # Using verified images instead of original
                'isMobile': is_mobile,
                'isIOS': is_ios
            }
        })
        
    except Exception as e:
        print(f"Error in duplicate_item: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Serverfehler beim Duplizieren'}), 500


def _soft_delete_item_groups(db, root_item_ids, username):
    """Soft-delete one or more item groups and their borrow records."""
    now = datetime.datetime.now()
    unique_group_item_ids = []
    seen_ids = set()

    for root_item_id in root_item_ids:
        for group_item_id in it.get_group_item_ids(root_item_id):
            if group_item_id not in seen_ids:
                seen_ids.add(group_item_id)
                unique_group_item_ids.append(group_item_id)

    if not unique_group_item_ids:
        return {
            'success': False,
            'message': 'Element nicht gefunden.',
            'soft_deleted_items': 0,
            'soft_deleted_borrows': 0,
            'group_item_ids': [],
            'archive': {
                'archive_created': False,
                'archived_files': 0,
                'deleted_files': 0,
                'archive_path': None,
            },
        }

    object_ids = []
    for group_item_id in unique_group_item_ids:
        try:
            object_ids.append(ObjectId(group_item_id))
        except Exception:
            continue

    item_docs_for_archive = []
    if object_ids:
        item_docs_for_archive = list(db['items'].find(
            {'_id': {'$in': object_ids}},
            {'_id': 1, 'Images': 1}
        ))

    delete_success = True
    soft_deleted_items = 0

    for group_item_id in unique_group_item_ids:
        result = db['items'].update_one(
            {'_id': ObjectId(group_item_id), 'Deleted': {'$ne': True}},
            {'$set': {
                'Deleted': True,
                'DeletedAt': now,
                'DeletedBy': username,
                'LastUpdated': now,
                'Verfuegbar': False,
            }}
        )
        if result.modified_count > 0:
            soft_deleted_items += 1
        elif result.matched_count == 0:
            delete_success = False

    borrow_result = db['ausleihungen'].update_many(
        {
            'Item': {'$in': unique_group_item_ids},
            'Status': {'$ne': 'deleted'}
        },
        {'$set': {
            'Status': 'deleted',
            'DeletedAt': now,
            'DeletedBy': username,
            'LastUpdated': now,
        }}
    )

    archive_result = {
        'archive_created': False,
        'archived_files': 0,
        'deleted_files': 0,
        'archive_path': None,
    }
    try:
        archive_result = encrypt_soft_deleted_media_pack(item_docs_for_archive, actor=username)
    except Exception as archive_err:
        app.logger.warning(f"Soft-delete media archive failed: {archive_err}")

    _append_audit_event_standalone(
        event_type='inventory_item_soft_deleted',
        payload={
            'root_item_ids': root_item_ids,
            'group_item_ids': unique_group_item_ids,
            'soft_deleted_items': soft_deleted_items,
            'soft_deleted_borrow_records': borrow_result.modified_count,
            'soft_delete_archive': archive_result,
        }
    )

    return {
        'success': delete_success,
        'soft_deleted_items': soft_deleted_items,
        'soft_deleted_borrows': borrow_result.modified_count,
        'group_item_ids': unique_group_item_ids,
        'archive': archive_result,
        'message': 'OK' if delete_success else 'Fehler beim revisionssicheren Deaktivieren des Elements.',
    }


@app.route('/delete_item/<id>', methods=['POST'])
def delete_item(id):
    """
    Route for deleting inventory items.
    
    Args:
        id (str): ID of the item to delete
        
    Returns:
        flask.Response: Redirect to admin homepage
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    # Resolve all related item ids (grouped variants) and load their data
    group_item_ids = it.get_group_item_ids(id)
    if not group_item_ids:
        flash('Element nicht gefunden.', 'error')
        return redirect(url_for('home_admin'))

    group_items = []
    for group_item_id in group_item_ids:
        group_item = it.get_item(group_item_id)
        if group_item:
            group_items.append(group_item)

    if not group_items:
        flash('Element nicht gefunden.', 'error')
        return redirect(url_for('home_admin'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        delete_result = _soft_delete_item_groups(db, [id], session.get('username', ''))
        delete_success = bool(delete_result.get('success'))
        soft_deleted_items = int(delete_result.get('soft_deleted_items', 0))
        soft_deleted_borrows = int(delete_result.get('soft_deleted_borrows', 0))
        archived_files = int((delete_result.get('archive') or {}).get('archived_files', 0))
    except Exception as e:
        app.logger.error(f"Error during soft-delete for item group {id}: {str(e)}")
        delete_success = False
        archived_files = 0
    finally:
        if client:
            client.close()

    if delete_success:
        flash(
            f'Elementgruppe revisionssicher deaktiviert ({soft_deleted_items}/{len(group_item_ids)} Versionen, {archived_files} Mediendateien verschlüsselt archiviert).',
            'success'
        )
    else:
        flash('Fehler beim revisionssicheren Deaktivieren des Elements.', 'error')
        
    return redirect(url_for('home_admin'))


@app.route('/bulk_delete_items', methods=['POST'])
def bulk_delete_items():
    """Soft-delete multiple selected item groups in one request."""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Nicht angemeldet'}), 401
    if not us.check_admin(session['username']):
        return jsonify({'success': False, 'message': 'Administratorrechte erforderlich'}), 403

    payload = request.get_json(silent=True) or {}
    item_ids = payload.get('item_ids') or request.form.getlist('item_ids')
    if isinstance(item_ids, str):
        item_ids = [item_ids]
    item_ids = [str(item_id).strip() for item_id in item_ids if str(item_id).strip()]

    if not item_ids:
        return jsonify({'success': False, 'message': 'Keine Elemente ausgewählt.'}), 400

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]

        result = _soft_delete_item_groups(db, item_ids, session.get('username', ''))
        if not result.get('success'):
            return jsonify({
                'success': False,
                'message': result.get('message', 'Fehler beim Löschen.'),
                'deleted_items': result.get('soft_deleted_items', 0),
                'deleted_borrows': result.get('soft_deleted_borrows', 0),
            }), 400

        return jsonify({
            'success': True,
            'message': f"{result.get('soft_deleted_items', 0)} Elemente revisionssicher deaktiviert.",
            'deleted_items': result.get('soft_deleted_items', 0),
            'deleted_borrows': result.get('soft_deleted_borrows', 0),
            'archived_files': (result.get('archive') or {}).get('archived_files', 0),
            'archive_created': (result.get('archive') or {}).get('archive_created', False),
            'group_item_ids': result.get('group_item_ids', []),
        })
    except Exception as e:
        app.logger.error(f"Error during bulk delete: {str(e)}")
        return jsonify({'success': False, 'message': 'Fehler beim Sammellöschen.'}), 500
    finally:
        if client:
            client.close()


@app.route('/edit_item/<id>', methods=['POST'])
def edit_item(id):
    """
    Route for editing an existing inventory item.
    
    Args:
        id (str): ID of the item to edit
        
    Returns:
        flask.Response: Redirect to admin homepage with status message
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    # Strip whitespace from all text fields
    name = sanitize_form_value(request.form.get('name'))
    ort = sanitize_form_value(request.form.get('ort'))
    beschreibung = sanitize_form_value(request.form.get('beschreibung'))
    
    # Strip whitespace from all filter values
    filter1 = sanitize_form_value(request.form.getlist('filter'))
    filter2 = sanitize_form_value(request.form.getlist('filter2'))
    filter3 = sanitize_form_value(request.form.getlist('filter3'))

    # Expand special "all values" selections for predefined filters.
    filter1 = expand_filter_selection(filter1, 1)
    filter2 = expand_filter_selection(filter2, 2)
    
    anschaffungs_jahr = sanitize_form_value(request.form.get('anschaffungsjahr'))
    anschaffungs_kosten = sanitize_form_value(request.form.get('anschaffungskosten'))
    code_4 = sanitize_form_value(request.form.get('code_4'))
    isbn_raw = sanitize_form_value(request.form.get('isbn', ''))
    reservierbar = 'reservierbar' in request.form

    item_isbn = ''
    item_type = 'general'
    if cfg.MODULES.is_enabled('library'):
        item_isbn = normalize_and_validate_isbn(isbn_raw)
        if isbn_raw and not item_isbn:
            flash('Ungültige ISBN. Bitte ISBN-10 oder ISBN-13 verwenden.', 'error')
            return redirect(url_for('home_admin'))
        if item_isbn:
            item_type = 'book'
    
    # Check if code is unique (excluding the current item)
    if code_4 and not it.is_code_unique(code_4, exclude_id=id):
        flash('Der Code wird bereits verwendet. Bitte wählen Sie einen anderen Code.', 'error')
        return redirect(url_for('home_admin'))
    
    # Get current item to check availability status
    current_item = it.get_item(id)
    if not current_item:
        flash('Element nicht gefunden', 'error')
        return redirect(url_for('home_admin'))
    
    # Preserve current availability status
    verfuegbar = current_item.get('Verfuegbar', True)
    
    # Handle existing images - get list of images to keep
    images_to_keep = request.form.getlist('existing_images')
    
    # Get the original list of images from the item
    original_images = current_item.get('Images', [])
    
    # Keep only the images that weren't marked for deletion
    images = [img for img in original_images if img in images_to_keep]
    
    # Handle new image uploads
    new_images = request.files.getlist('new_images')
    
    # Process any new image uploads
    for image in new_images:
        if image and image.filename:
            is_allowed, error_message = allowed_file(image.filename)
            if is_allowed:
                # Get the file extension
                _, ext_part = os.path.splitext(secure_filename(image.filename))
                
                # Generate a completely unique filename using UUID
                unique_id = str(uuid.uuid4())
                timestamp = time.strftime("%Y%m%d%H%M%S")
                
                # New filename format with UUID to ensure uniqueness
                filename = f"{unique_id}_{timestamp}{ext_part}"
                
                image.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                
                # Optimize the image
                try:
                    opt_result = generate_optimized_versions(filename, max_original_width=500, target_size_kb=80)
                    if opt_result['success'] and opt_result['original']:
                        filename = opt_result['original']
                except Exception as e:
                    app.logger.error(f"Error optimizing image in edit_item: {e}")
                
                images.append(filename)
            else:
                flash(error_message, 'error')
                return redirect(url_for('home_admin'))

    # If location is not in the predefined list, maybe add it (depending on policy)
    predefined_locations = it.get_predefined_locations()
    if ort and ort not in predefined_locations:
        it.add_predefined_location(ort)
    
    # Update the item
    result = it.update_item(
        id, name, ort, beschreibung, 
        images, verfuegbar, filter1, filter2, filter3,
        anschaffungs_jahr, anschaffungs_kosten, code_4, reservierbar,
        isbn=item_isbn,
        item_type=item_type
    )
    
    if result:
        flash('Element erfolgreich aktualisiert', 'success')
    else:
        flash('Fehler beim Aktualisieren des Elements', 'error')
    
    return redirect(url_for('home_admin'))


@app.route('/report_damage/<id>', methods=['POST'])
def report_damage(id):
    """Register a damage report entry for an item (admin only)."""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Nicht angemeldet'}), 401
    if not us.check_admin(session['username']):
        return jsonify({'success': False, 'message': 'Administratorrechte erforderlich'}), 403

    payload = request.get_json(silent=True) or {}
    description = str(payload.get('description', '')).strip() or 'Schaden gemeldet'

    if len(description) > 1000:
        return jsonify({'success': False, 'message': 'Die Schadensbeschreibung ist zu lang (max. 1000 Zeichen).'}), 400

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        item_doc = items_col.find_one({'_id': ObjectId(id)}, {'Name': 1})
        if not item_doc:
            return jsonify({'success': False, 'message': 'Objekt nicht gefunden.'}), 404

        now = datetime.datetime.now()
        damage_entry = {
            'description': description,
            'reported_by': session['username'],
            'reported_at': now,
        }

        result = items_col.update_one(
            {'_id': ObjectId(id)},
            {
                '$push': {'DamageReports': {'$each': [damage_entry], '$position': 0}},
                '$set': {'HasDamage': True, 'LastUpdated': now}
            }
        )

        if result.matched_count == 0:
            return jsonify({'success': False, 'message': 'Objekt nicht gefunden.'}), 404

        updated_item = items_col.find_one({'_id': ObjectId(id)}, {'DamageReports': 1})
        damage_count = len(updated_item.get('DamageReports', [])) if updated_item else 0

        _create_notification(
            db,
            audience='admin',
            notif_type='damage_reported',
            title='Defekt gemeldet',
            message=(
                f"Fuer das Item '{item_doc.get('Name', id)}' wurde ein Defekt gemeldet. "
                f"Meldung von {session.get('username', '-')}: {description}"
            ),
            reference={'item_id': id, 'damage_count': damage_count, 'url': '/admin/damaged_items'},
            severity='danger',
        )

        # Best-effort system log entry for auditability
        try:
            logs_collection = db['system_logs']
            logs_collection.insert_one({
                'type': 'damage_report',
                'timestamp': now.isoformat(),
                'user': session.get('username'),
                'item_id': id,
                'item_name': item_doc.get('Name', ''),
                'note': description,
                'damage_count': damage_count,
                'ip': request.remote_addr
            })
        except Exception as log_err:
            app.logger.warning(f"Damage report log write failed for item {id}: {log_err}")

        return jsonify({
            'success': True,
            'message': 'Schaden erfolgreich erfasst.',
            'damage_count': damage_count
        })
    except Exception as e:
        app.logger.error(f"Error reporting damage for item {id}: {e}")
        return jsonify({'success': False, 'message': 'Fehler beim Speichern der Schadensmeldung.'}), 500
    finally:
        if client:
            client.close()


@app.route('/mark_damage_repaired/<id>', methods=['POST'])
def mark_damage_repaired(id):
    """Mark all currently open damage reports of an item as repaired (admin only)."""
    if 'username' not in session:
        return jsonify({'success': False, 'message': 'Nicht angemeldet'}), 401
    if not us.check_admin(session['username']):
        return jsonify({'success': False, 'message': 'Administratorrechte erforderlich'}), 403

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        ausleihungen_col = db['ausleihungen']

        item = items_col.find_one({'_id': ObjectId(id)}, {'DamageReports': 1, 'DamageRepairs': 1})
        if not item:
            return jsonify({'success': False, 'message': 'Objekt nicht gefunden.'}), 404

        open_reports = item.get('DamageReports', [])
        if not open_reports:
            return jsonify({'success': False, 'message': 'Keine offenen Schäden vorhanden.'}), 400

        active_borrow = ausleihungen_col.find_one({'Item': str(id), 'Status': 'active'}, {'_id': 1})
        now = datetime.datetime.now()
        repair_entry = {
            'repaired_by': session['username'],
            'repaired_at': now,
            'resolved_reports': open_reports
        }

        result = items_col.update_one(
            {'_id': ObjectId(id)},
            {
                '$push': {'DamageRepairs': {'$each': [repair_entry], '$position': 0}},
                '$set': {
                    'DamageReports': [],
                    'HasDamage': False,
                    'LastUpdated': now,
                    **({} if active_borrow else {'Verfuegbar': True}),
                }
            }
        )

        if not active_borrow:
            items_col.update_one(
                {'_id': ObjectId(id)},
                {
                    '$unset': {'User': '', 'Condition': ''}
                }
            )

        if result.matched_count == 0:
            return jsonify({'success': False, 'message': 'Objekt nicht gefunden.'}), 404

        # Best-effort system log entry for repair action
        try:
            logs_collection = db['system_logs']
            logs_collection.insert_one({
                'type': 'damage_repair',
                'timestamp': now.isoformat(),
                'user': session.get('username'),
                'item_id': id,
                'resolved_count': len(open_reports),
                'ip': request.remote_addr
            })
        except Exception as log_err:
            app.logger.warning(f"Damage repair log write failed for item {id}: {log_err}")

        return jsonify({
            'success': True,
            'message': 'Schäden als repariert markiert.',
            'resolved_count': len(open_reports)
        })
    except Exception as e:
        app.logger.error(f"Error marking damages repaired for item {id}: {e}")
        return jsonify({'success': False, 'message': 'Fehler beim Markieren als repariert.'}), 500
    finally:
        if client:
            client.close()


@app.route('/get_ausleihungen', methods=['GET'])
def get_ausleihungen():
    """
    API endpoint to retrieve all borrowing records.
    
    Returns:
        dict: Dictionary containing all borrowing records
    """
    ausleihungen = au.get_ausleihungen()
    return {'ausleihungen': ausleihungen}


@app.route('/ausleihen/<id>', methods=['POST'])
def ausleihen(id):
    """
    Route for borrowing an item from inventory.
    
    Args:
        id (str): ID of the item to borrow
        
    Returns:
        flask.Response: Redirect to appropriate homepage
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    username = session['username']
    requested_return_to = (request.form.get('return_to') or '').strip().lower()
    if requested_return_to == 'library' and cfg.MODULES.is_enabled('library'):
        redirect_target = 'library_view'
    else:
        redirect_target = 'home_admin' if us.check_admin(username) else 'home'

    item = it.get_item(id)
    if not item:
        flash('Element nicht gefunden', 'error')
        return redirect(url_for(redirect_target))

    effective_borrower = username
    borrow_duration_days = None
    student_card_id = us.normalize_student_card_id(request.form.get('borrower_card_id'))
    item_type = str(item.get('ItemType', '')).strip().lower()
    is_library_item = item_type in LIBRARY_ITEM_TYPES

    # Library media can only be borrowed with a valid student card.
    if is_library_item:
        if not cfg.MODULES.is_enabled('student_cards'):
            flash('Bibliotheksmedien können nur mit aktivem Schülerausweis-Modul ausgeliehen werden.', 'error')
            return redirect(url_for(redirect_target))
        if not student_card_id:
            flash('Für Bibliotheksmedien ist eine gültige Schülerausweis-ID erforderlich.', 'error')
            return redirect(url_for(redirect_target))

        client = None
        try:
            client = MongoClient(MONGODB_HOST, MONGODB_PORT)
            db = client[MONGODB_DB]
            student_cards_col = db['student_cards']
            card_doc = student_cards_col.find_one({'AusweisId': student_card_id})
            if not card_doc:
                flash('Ungültige Schülerausweis-ID. Bibliotheksmedien können nur mit gültigem Ausweis ausgeliehen werden.', 'error')
                return redirect(url_for(redirect_target))

            effective_borrower = card_doc.get('SchülerName') or f"Ausweis {student_card_id}"

            if not (request.form.get('borrow_duration_days') or '').strip():
                try:
                    card_default = int(card_doc.get('StandardAusleihdauer', cfg.STUDENT_DEFAULT_BORROW_DAYS))
                except (TypeError, ValueError):
                    card_default = cfg.STUDENT_DEFAULT_BORROW_DAYS
                borrow_duration_days = max(1, min(card_default, cfg.STUDENT_MAX_BORROW_DAYS))
        finally:
            if client:
                client.close()

    if cfg.MODULES.is_enabled('student_cards'):
        duration_raw = (request.form.get('borrow_duration_days') or '').strip()
        if duration_raw:
            try:
                parsed_duration = int(duration_raw)
                if 1 <= parsed_duration <= cfg.STUDENT_MAX_BORROW_DAYS:
                    borrow_duration_days = parsed_duration
                else:
                    flash(f'Ausleihdauer muss zwischen 1 und {cfg.STUDENT_MAX_BORROW_DAYS} Tagen liegen.', 'error')
                    return redirect(url_for(redirect_target))
            except ValueError:
                flash('Ungültige Ausleihdauer angegeben.', 'error')
                return redirect(url_for(redirect_target))

        # Admins can borrow on behalf of students via student card id.
        if us.check_admin(username) and not is_library_item:
            if student_card_id:
                student_user = us.get_user_by_student_card(student_card_id)
                if not student_user:
                    flash('Keine Schülerin/kein Schüler mit dieser Ausweis-ID gefunden.', 'error')
                    return redirect(url_for('home_admin'))
                effective_borrower = student_user.get('Username') or student_user.get('username') or username
                if borrow_duration_days is None:
                    try:
                        card_default = int(student_user.get('MaxBorrowDays', cfg.STUDENT_DEFAULT_BORROW_DAYS))
                    except (TypeError, ValueError):
                        card_default = cfg.STUDENT_DEFAULT_BORROW_DAYS
                    borrow_duration_days = max(1, min(card_default, cfg.STUDENT_MAX_BORROW_DAYS))

    start_date = datetime.datetime.now()
    end_date = None
    if borrow_duration_days:
        end_date = start_date + datetime.timedelta(days=borrow_duration_days)

    # Grouped inventory mode: parent item + hidden sub-items are handled as separate physical units.
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        grouped_children = list(items_col.find({'ParentItemId': id, 'IsGroupedSubItem': True}))
        client.close()
    except Exception:
        grouped_children = []

    if grouped_children:
        exemplare_count = request.form.get('exemplare_count', 1)
        specific_item_id = (request.form.get('specific_item_id') or '').strip()
        try:
            exemplare_count = int(exemplare_count)
            if exemplare_count < 1:
                exemplare_count = 1
        except (ValueError, TypeError):
            exemplare_count = 1

        grouped_units = [item] + [{**child, '_id': str(child.get('_id'))} for child in grouped_children]
        available_units = [unit for unit in grouped_units if unit.get('Verfuegbar', True)]

        planned_counts = {}
        try:
            client = MongoClient(MONGODB_HOST, MONGODB_PORT)
            db = client[MONGODB_DB]
            ausleihungen_col = db['ausleihungen']
            unit_ids = [str(unit.get('_id')) for unit in grouped_units if unit.get('_id')]
            planned_cursor = ausleihungen_col.find({'Item': {'$in': unit_ids}, 'Status': 'planned'})
            for planned_booking in planned_cursor:
                planned_item_id = str(planned_booking.get('Item') or '')
                if not planned_item_id:
                    continue
                planned_counts[planned_item_id] = planned_counts.get(planned_item_id, 0) + 1
            client.close()
        except Exception:
            planned_counts = {}

        if not available_units:
            flash('Keine Exemplare verfügbar.', 'error')
            return redirect(url_for(redirect_target))

        selected_units = []
        if specific_item_id:
            chosen = next((unit for unit in available_units if str(unit.get('_id')) == specific_item_id), None)
            if not chosen:
                flash('Das ausgewählte Exemplar ist nicht verfügbar.', 'error')
                return redirect(url_for(redirect_target))
            selected_units = [chosen]
        else:
            available_units.sort(key=lambda unit: (planned_counts.get(str(unit.get('_id')) or '', 0), str(unit.get('Code_4') or ''), str(unit.get('_id') or '')))
            if exemplare_count > len(available_units):
                flash(f'Nicht genügend Exemplare verfügbar. Angefordert: {exemplare_count}, Verfügbar: {len(available_units)}', 'error')
                return redirect(url_for(redirect_target))
            selected_units = available_units[:exemplare_count]

        for unit in selected_units:
            unit_id = str(unit.get('_id'))
            it.update_item_status(unit_id, False, effective_borrower)
            au.add_ausleihung(unit_id, effective_borrower, start_date, end_date=end_date)

        _append_audit_event_standalone(
            event_type='ausleihung_returned',
            payload={
                'channel': 'inventory_route',
                'item_id': id,
                'borrower': effective_borrower,
                'borrow_duration_days': borrow_duration_days,
                'selected_unit_ids': [str(unit.get('_id')) for unit in selected_units],
                'selected_unit_codes': [str(unit.get('Code_4') or '') for unit in selected_units],
            }
        )

        if len(selected_units) == 1:
            selected_code = selected_units[0].get('Code_4') or '-'
            flash(f'Exemplar {selected_code} erfolgreich ausgeliehen', 'success')
        else:
            flash(f'{len(selected_units)} Exemplare erfolgreich ausgeliehen', 'success')

        return redirect(url_for(redirect_target))
    
    # Before borrowing, block if there's a conflicting planned booking
    try:
        now = datetime.datetime.now()
        # Fetch planned bookings for this item from DB
        planned = au.get_planned_ausleihungen()
        # Count relevant upcoming planned bookings for today or ongoing
        upcoming_planned_today = []
        for appt in planned:
            appt_item = str(appt.get('Item')) if appt.get('Item') is not None else None
            if appt_item != id:
                continue
            appt_start = appt.get('Start')
            appt_end = appt.get('End') or appt_start
            if not appt_start:
                continue
            # Consider conflict if appointment ends in the future and is today
            try:
                if appt_end >= now and appt_start.date() == now.date():
                    upcoming_planned_today.append(appt)
            except Exception:
                # Fallback simple check
                if appt_start.date() == now.date():
                    upcoming_planned_today.append(appt)
        if upcoming_planned_today:
            # For single-instance items, block outright; for multi-exemplar, allow only if capacity suffices
            item_doc = it.get_item(id)
            total_exemplare = item_doc.get('Exemplare', 1) if item_doc else 1
            if total_exemplare <= 1:
                flash('Dieses Objekt hat heute eine geplante Reservierung und kann aktuell nicht ausgeliehen werden.', 'error')
                return redirect(url_for(redirect_target))
            else:
                # If planned count equals or exceeds remaining capacity, block
                current_borrowed = len(item_doc.get('ExemplareStatus', [])) if item_doc else 0
                if current_borrowed + len(upcoming_planned_today) >= total_exemplare:
                    flash('Alle Exemplare sind aufgrund geplanter Reservierungen heute belegt.', 'error')
                    return redirect(url_for(redirect_target))
    except Exception as e:
        app.logger.warning(f"Could not enforce planned booking guard: {e}")

    # Get number of exemplars to borrow (default to 1)
    exemplare_count = request.form.get('exemplare_count', 1)
    try:
        exemplare_count = int(exemplare_count)
        if exemplare_count < 1:
            exemplare_count = 1
    except (ValueError, TypeError):
        exemplare_count = 1
        
    # Check if the item has exemplars defined
    total_exemplare = item.get('Exemplare', 1)
    
    # Get current exemplar status
    exemplare_status = item.get('ExemplareStatus', [])
    
    # Count how many exemplars are currently available
    borrowed_count = len(exemplare_status)
    available_count = total_exemplare - borrowed_count
    
    if available_count < exemplare_count:
        flash(f'Nicht genügend Exemplare verfügbar. Angefordert: {exemplare_count}, Verfügbar: {available_count}', 'error')
        return redirect(url_for(redirect_target))
    
    # If we reach here, we can borrow the requested number of exemplars
    current_date = datetime.datetime.now().strftime('%d.%m.%Y %H:%M')
    
    # If the item doesn't use exemplars (single item)
    if total_exemplare <= 1:
        it.update_item_status(id, False, effective_borrower)
        start_date = datetime.datetime.now()
        au.add_ausleihung(id, effective_borrower, start_date, end_date=end_date)
        _append_audit_event_standalone(
            event_type='ausleihung_returned',
            payload={
                'channel': 'inventory_route',
                'item_id': id,
                'borrower': effective_borrower,
                'borrow_duration_days': borrow_duration_days,
                'single_item': True,
            }
        )
        flash('Element erfolgreich ausgeliehen', 'success')
    else:
        # Handle multi-exemplar item
        new_borrowed_exemplars = []
        
        # Create new entries for borrowed exemplars
        for i in range(exemplare_count):
            # Find the next available exemplar number
            exemplar_number = 1
            used_numbers = [ex.get('number') for ex in exemplare_status]
            
            while exemplar_number in used_numbers:
                exemplar_number += 1
                
            new_borrowed_exemplars.append({
                'number': exemplar_number,
                'user': effective_borrower,
                'date': current_date
            })
        
        # Add new borrowed exemplars to the status
        updated_status = exemplare_status + new_borrowed_exemplars
        
        # Update the item with the new status
        it.update_item_exemplare_status(id, updated_status)
        
        # Update the item's availability if all exemplars are borrowed
        if len(updated_status) >= total_exemplare:
            it.update_item_status(id, False, username)
        
        # Create ausleihung records for each borrowed exemplar
        start_date = datetime.datetime.now()
        for exemplar in new_borrowed_exemplars:
            exemplar_id = f"{id}_{exemplar['number']}"
            au.add_ausleihung(exemplar_id, effective_borrower, start_date, end_date=end_date, exemplar_data={
                'parent_id': id,
                'exemplar_number': exemplar['number']
            })

        _append_audit_event_standalone(
            event_type='ausleihung_returned',
            payload={
                'channel': 'inventory_route',
                'item_id': id,
                'borrower': effective_borrower,
                'borrow_duration_days': borrow_duration_days,
                'exemplar_numbers': [ex.get('number') for ex in new_borrowed_exemplars],
            }
        )
        
        flash(f'{exemplare_count} Exemplare erfolgreich ausgeliehen', 'success')
    
    return redirect(url_for(redirect_target))


@app.route('/zurueckgeben/<id>', methods=['POST'])
def zurueckgeben(id): 
    """
    Route for returning a borrowed item.
    Creates or updates a record of the borrowing session.
    
    Args:
        id (str): ID of the item to return
        
    Returns:
        flask.Response: Redirect to appropriate homepage
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    item = it.get_item(id)
    if not item:
        flash('Element nicht gefunden', 'error')
        return redirect(url_for('home'))
        
    username = session['username']
    
    if not item.get('Verfuegbar', True) and (us.check_admin(session['username']) or item.get('User') == username):
        try:
            # Get ALL active borrowing records for this item and complete them
            client = MongoClient(MONGODB_HOST, MONGODB_PORT)
            db = client[MONGODB_DB]
            ausleihungen = db['ausleihungen']
            
            # Find all active records for this item
            active_records = ausleihungen.find({
                'Item': id,
                'Status': 'active'
            })
            
            end_date = datetime.datetime.now()
            original_user = item.get('User', username)
            
            updated_count = 0
            for record in active_records:
                ausleihung_id = str(record['_id'])
                # Update each active record
                result = ausleihungen.update_one(
                    {'_id': ObjectId(ausleihung_id)},
                    {'$set': {
                        'Status': 'completed',
                        'End': end_date,
                        'LastUpdated': datetime.datetime.now()
                    }}
                )
                
                if result.modified_count > 0:
                    updated_count += 1
            
            client.close()
            
            # Update the item status
            it.update_item_status(id, True, original_user)
            
            if updated_count > 0:
                flash(f'Element erfolgreich zurückgegeben ({updated_count} Datensätze abgeschlossen)', 'success')
            else:
                flash('Element erfolgreich zurückgegeben', 'success')

            _append_audit_event_standalone(
                event_type='ausleihung_returned',
                payload={
                    'channel': 'inventory_route',
                    'item_id': id,
                    'returned_by': username,
                    'original_borrower': original_user,
                    'completed_records': updated_count,
                }
            )
                
        except Exception as e:
            print(f"Error in return process: {e}")
            it.update_item_status(id, True)
            flash(f'Element zurückgegeben, aber ein Fehler ist aufgetreten: {str(e)}', 'warning')
    else:
        flash('Sie sind nicht berechtigt, dieses Element zurückzugeben, oder es ist bereits verfügbar', 'error')

    # Check if request came from my_borrowed_items page
    source_page = request.form.get('source_page')
    referrer = request.headers.get('Referer', '')
    if source_page == 'my_borrowed_items' or '/my_borrowed_items' in referrer:
        return redirect(url_for('my_borrowed_items'))
    
    if 'username' in session and not us.check_admin(session['username']):
        return redirect(url_for('home'))
    return redirect(url_for('home_admin'))

@app.route('/get_filter', methods=['GET'])
def get_filter():
    """
    API endpoint to retrieve available item filters/categories.

    Returns:
        dict: Dictionary of available filters by filter field
    """
    return jsonify({
        'filter1': it.get_primary_filters(),
        'filter2': it.get_secondary_filters(),
        'filter3': it.get_tertiary_filters(),
    })
    

@app.route('/get_ausleihung_by_item/<id>')
def get_ausleihung_by_item_route(id):
    """
    API endpoint to retrieve borrowing details for a specific item.
    
    Args:
        id (str): ID of the item to retrieve
        
    Returns:
        dict: Borrowing details for the item
    """
    if 'username' not in session:
        return {'error': 'Not authorized', 'status': 'forbidden'}, 403
    
    # Get the borrowing record
    ausleihung = au.get_ausleihung_by_item(id, include_history=False)        # Add client-side status verification if a borrowing record exists
    if ausleihung:
        # Add verified status to each borrowing record with logging
        current_status = au.get_current_status(
            ausleihung,
            log_changes=True, 
            user=session.get('username', None)
        )
        ausleihung['VerifiedStatus'] = current_status
    
    # Admin users can see all borrowing details
    # Regular users can only see their own borrowings
    if ausleihung and (us.check_admin(session['username']) or ausleihung.get('User') == session['username']):
        return {'ausleihung': ausleihung, 'status': 'success'}
    
    # Get item name for better error message
    item = it.get_item(id)
    item_name = item.get('Name', 'Unknown') if item else 'Unknown'
    
    # Return a more informative error
    return {
        'error': 'No active borrowing record found for this item',
        'item_name': item_name,
        'status': 'not_found'
    }, 200  # Return 200 instead of 404 to allow processing of the error message


@app.route('/get_planned_bookings/<item_id>')
def get_planned_bookings(item_id):
    """
    Return all planned bookings for a given item (admin only).
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 403

    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        cursor = ausleihungen.find({'Item': item_id, 'Status': 'planned'}).sort('Start', 1)
        bookings = []
        for r in cursor:
            bookings.append({
                'id': str(r.get('_id')),
                'user': r.get('User', ''),
                'period': r.get('Period'),
                'start': r.get('Start').isoformat() if r.get('Start') else None,
                'end': r.get('End').isoformat() if r.get('End') else None,
                'notes': r.get('Notes', '')
            })
        client.close()
        return jsonify({'ok': True, 'bookings': bookings})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/get_planned_bookings_public/<item_id>')
def get_planned_bookings_public(item_id):
    """
    Return planned bookings for a given item (normal users; limited fields, no notes).
    """
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401

    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        cursor = ausleihungen.find({'Item': item_id, 'Status': 'planned'}).sort('Start', 1)
        bookings = []
        for r in cursor:
            bookings.append({
                'period': r.get('Period'),
                'start': r.get('Start').isoformat() if r.get('Start') else None,
                'end': r.get('End').isoformat() if r.get('End') else None
            })
        client.close()
        return jsonify({'ok': True, 'bookings': bookings})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/check_availability')
def check_availability():
    """
    Check if a given item is available for the specified date and period range.
    Query params: item_id, date=YYYY-MM-DD, start=<1-10>, end=<1-10>
    Returns: { ok, available, conflicts:[...] }
    """
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401

    item_id = request.args.get('item_id')
    date_str = request.args.get('date')
    start_p = request.args.get('start')
    end_p = request.args.get('end') or start_p
    if not item_id or not date_str or not start_p:
        return jsonify({'ok': False, 'error': 'missing parameters'}), 400

    try:
        # Parse date
        booking_date = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        start_num = int(start_p)
        end_num = int(end_p)
        if end_num < start_num:
            start_num, end_num = end_num, start_num

        # Compute requested time window
        start_times = get_period_times(booking_date, start_num)
        end_times = get_period_times(booking_date, end_num)
        if not start_times or not end_times:
            return jsonify({'ok': False, 'error': 'invalid period(s)'}), 400
        req_start = start_times['start']
        req_end = end_times['end']

        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        items_col = db['items']

        # Collect potential conflicts (planned and active) for this day
        same_day_start = datetime.datetime.combine(booking_date.date(), datetime.time.min)
        same_day_end = datetime.datetime.combine(booking_date.date(), datetime.time.max)
        candidates = list(ausleihungen.find({
            'Item': item_id,
            'Status': {'$in': ['planned', 'active']},
            'Start': {'$lte': same_day_end},
            'End': {'$gte': same_day_start}
        }))

        conflicts = []
        for r in candidates:
            r_start = r.get('Start')
            r_end = r.get('End')
            # If end missing for active, assume lasts through the day
            if r_end is None:
                r_end = same_day_end
            if r_start is None:
                r_start = same_day_start
            # Overlap check: req_start < r_end and req_end > r_start
            if req_start < r_end and req_end > r_start:
                conflicts.append({
                    'id': str(r.get('_id')),
                    'status': r.get('Status'),
                    'user': r.get('User', ''),
                    'start': r_start.isoformat() if r_start else None,
                    'end': r_end.isoformat() if r_end else None,
                    'period': r.get('Period')
                })

        # Also include current availability if checking today and item is borrowed now
        item_doc = items_col.find_one({'_id': ObjectId(item_id)})
        if item_doc and not item_doc.get('Verfuegbar', True):
            now = datetime.datetime.now()
            if req_start.date() == now.date():
                conflicts.append({'status': 'active', 'user': item_doc.get('User'), 'start': None, 'end': None, 'period': None, 'id': None})

        client.close()
        return jsonify({'ok': True, 'available': len(conflicts) == 0, 'conflicts': conflicts})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# def create_qr_code(id):
#     """
#     Generate a QR code for an item.
#     The QR code contains a URL that points to the item details.
#     
#     Args:
#         id (str): ID of the item to generate QR code for
#         
#     Returns:
#         str: Filename of the generated QR code, or None if item not found
#     """
#     qr = qrcode.QRCode(
#         version=1,
#         error_correction=ERROR_CORRECT_L,  # Use imported constant
#         box_size=10,
#         border=4,
#     )
#     
#     # Parse and reconstruct the URL properly
#     parsed_url = urlparse(request.url_root)
#     
#     # Force HTTPS if needed
#     scheme = 'https' if parsed_url.scheme == 'http' else parsed_url.scheme
#     
#     # Properly reconstruct the base URL
#     base_url = urlunparse((scheme, parsed_url.netloc, '', '', '', ''))
#     
#     # URL that will open this item directly
#     item_url = f"{base_url}:{Port}/item/{id}"
#     qr.add_data(item_url)
#     qr.make(fit=True)
#
#     item = it.get_item(id)
#     if not item:
#         return None
#     
#     img = qr.make_image(fill_color="black", back_color="white")
#     
#     # Create a unique filename using UUID
#     unique_id = str(uuid.uuid4())
#     timestamp = time.strftime("%Y%m%d%H%M%S")
#     
#     # Still include the original name for readability but ensure uniqueness with UUID
#     safe_name = secure_filename(item['Name'])
#     filename = f"{safe_name}_{unique_id}_{timestamp}.png"
#     qr_path = os.path.join(app.config['QR_CODE_FOLDER'], filename)
#
#     
#     # Fix the file handling - save to file object, not string
#     with open(qr_path, 'wb') as f:
#         img.save(f)
#     
#     return filename

# Fix fromisoformat None value checks
@app.route('/plan_booking', methods=['POST'])
def plan_booking():
    """
    Create a new planned booking or a range of bookings
    """
    if 'username' not in session:
        return {"success": False, "error": "Not authenticated"}, 401
        
    try:
        # Extract form data
        item_id = (request.form.get('item_id') or '').strip()
        start_date_str = (request.form.get('booking_date') or request.form.get('start_date') or '').strip()
        end_date_str = (request.form.get('booking_end_date') or request.form.get('end_date') or '').strip()
        period_start = (request.form.get('period_start') or '').strip()
        period_end = (request.form.get('period_end') or '').strip()
        notes = html.escape(request.form.get('notes', '') or '')
        booking_type = (request.form.get('booking_type', 'single') or 'single').strip().lower()
        
        # Validate inputs
        if not all([item_id, start_date_str, period_start]):
            return {"success": False, "error": "Missing required fields"}, 400
        if not end_date_str:
            end_date_str = start_date_str
            
        # Parse dates
        try:
            if start_date_str:
                start_date = datetime.datetime.fromisoformat(start_date_str)
            else:
                return {"success": False, "error": "Missing start date"}, 400
            
            if end_date_str:
                end_date = datetime.datetime.fromisoformat(end_date_str)
            else:
                return {"success": False, "error": "Missing end date"}, 400
            
            # For single day bookings, use the start date as the end date
            if booking_type == 'single':
                end_date = start_date
        except ValueError as e:
            return {"success": False, "error": f"Invalid date format: {e}"}, 400
            
        # Check if item exists
        item = it.get_item(item_id)
        if not item:
            return {"success": False, "error": "Item not found"}, 404
            
        # Check if item is reservable
        if not item.get('Reservierbar', True):
             return {"success": False, "error": "Dieses Item kann nicht reserviert werden."}, 400
        
        # Handle period range
        periods = []
        if period_start:
            period_start_num = int(period_start)
        else:
            period_start_num = 1  # Default if None
    
        # If period_end is provided, it's a range of periods
        if period_end:
            period_end_num = int(period_end)
            
            # Validate period range
            if period_end_num < period_start_num:
                return {"success": False, "error": "End period cannot be before start period"}, 400
                
            # Create list of all periods in the range
            periods = list(range(period_start_num, period_end_num + 1))
        else:
            # Single period booking
            periods = [period_start_num]
            
        # For date range bookings, we'll process each date separately
        booking_ids = []
        errors = []
        
        # If it's a range of days
        if booking_type == 'range' and start_date != end_date:
            current_date = start_date
            while current_date <= end_date:
                # For each day in the range
                day_booking_ids, day_errors = process_day_bookings(
                    item_id, 
                    current_date,
                    periods,
                    notes
                )
                booking_ids.extend(day_booking_ids)
                errors.extend(day_errors)
                
                # Move to next day
                current_date += datetime.timedelta(days=1)
        else:
            # Single day with multiple periods
            booking_ids, errors = process_day_bookings(
                item_id,
                start_date,
                periods,
                notes
            )
            
        # Return results
        if errors:
            if booking_ids:
                # Some succeeded, some failed
                return {
                    "success": True, 
                    "partial": True,
                    "booking_ids": booking_ids,
                    "errors": errors
                }
            else:
                # All failed
                return {"success": False, "errors": errors}, 400
        else:
            # All succeeded
            return {"success": True, "booking_ids": booking_ids}
            
    except Exception as e:
        import traceback
        print(f"Error in plan_booking: {e}")
        traceback.print_exc()
        return {"success": False, "error": f"Serverfehler: {str(e)}"}, 500

def process_day_bookings(item_id, booking_date, periods, notes):
    """
    Helper function to process bookings for a single day across multiple periods
    
    Args:
        item_id: The item to book
        booking_date: The date for the booking
        periods: List of period numbers to book
        notes: Booking notes
        
    Returns:
        tuple: (list of booking_ids, list of errors)
    """
    booking_ids = []
    errors = []
    
    for period in periods:
        # Get period times
        period_times = get_period_times(booking_date, period)
        if not period_times:
            errors.append(f"Invalid period {period}")
            continue
            
        # Create the start and end times for this period
        start_time = period_times.get('start')
        end_time = period_times.get('end')
        
        # Check for conflicts
        if au.check_booking_conflict(item_id, start_time, end_time, period):
            errors.append(f"Conflict for period {period} on {booking_date.strftime('%Y-%m-%d')}")
            continue
            
        # Create the booking
        booking_id = au.add_planned_booking(
            item_id,
            session['username'],
            start_time,
            end_time,
            notes,
            period=period
        )
        
        if booking_id:
            booking_ids.append(str(booking_id))
        else:
            errors.append(f"Failed to create booking for period {period}")
            
    return booking_ids, errors
@app.route('/add_booking', methods=['POST'])
def add_booking():
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Not logged in'})
    
    item_id = html.escape(request.form.get('item_id'))
    
    # Check if item exists and is reservable
    item = it.get_item(item_id)
    if not item:
        return jsonify({'success': False, 'error': 'Item not found'})
        
    if not item.get('Reservierbar', True):
        return jsonify({'success': False, 'error': 'Dieses Item kann nicht reserviert werden.'})

    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    period = request.form.get('period')
    notes = request.form.get('notes', '')
    
    # Parse dates as naive datetime objects
    try:
        # Simple datetime parsing without timezone
        if start_date_str:
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d %H:%M:%S')
        else:
            return jsonify({'success': False, 'error': 'Missing start date'})
        
        if end_date_str:
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S')
        else:
            end_date = None
        
        # Continue with adding the booking
        booking_id = au.add_planned_booking(
            item_id=item_id,
            user=session['username'],
            start_date=start_date,
            end_date=end_date,
            notes=notes,
            period=period
        )
        
        return jsonify({'success': True, 'booking_id': str(booking_id)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/cancel_booking/<id>', methods=['POST'])
def cancel_booking(id):
    """
    Cancel a planned booking
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    # Get the booking
    booking = au.get_booking(id)
    if not booking:
        return {"success": False, "error": "Booking not found"}, 404
        
    # Check if user owns this booking
    if booking.get('User') != session['username'] and not us.check_admin(session['username']):
               return {"success": False, "error": "Not authorized to cancel this booking"}, 403
    
    # Cancel the booking
    result = au.cancel_booking(id)
    
    if result:
        return {"success": True}
    else:
        return {"success": False, "error": "Failed to cancel booking"}

@app.route('/terminplan', methods=['GET'])
def terminplan():
    """
    Route to display the booking calendar
    """
    try:
        if 'username' not in session:
            flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
            return redirect(url_for('login'))
        
        # Make sure the template exists
        template_path = os.path.join(BASE_DIR, 'templates', 'terminplan.html')
        if not os.path.exists(template_path):
            print(f"Template file not found: {template_path}")
            flash('Vorlage nicht gefunden. Bitte kontaktieren Sie den Administrator.', 'error')
            return redirect(url_for('home'))
            
        return render_template('terminplan.html', school_periods=SCHOOL_PERIODS)
    except Exception as e:
        import traceback
        print(f"Error rendering terminplan: {e}")
        traceback.print_exc()
        flash('Ein Fehler ist beim Anzeigen des Kalenders aufgetreten.', 'error')
        return redirect(url_for('home'))


'''-------------------------------------------------------------------------------------------------------------ADMIN ROUTES------------------------------------------------------------------------------------------------------------------'''

@app.route('/register', methods=['GET', 'POST'])
def register():
    """
    User registration route.false
    Returns:
        flask.Response: Rendered template or redirect
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if 'username' in session:
        if request.method == 'POST':
            password = request.form['password']
            name = (request.form.get('name') or '').strip()
            last_name = (request.form.get('last-name') or '').strip()

            # Generate a username from the first 2 letters of first and last name.
            username = us.build_unique_username_from_name(name, last_name)
            
            permission_preset = (request.form.get('permission_preset') or 'standard_user').strip()
            use_custom_permissions = request.form.get('use_custom_permissions') == 'on'
            
            if not username or not password or not name or not last_name:
                flash('Bitte füllen Sie alle Felder aus', 'error')
                return redirect(url_for('register'))
            if not us.check_password_strength(password):
                flash('Passwort ist zu schwach', 'error')
                return redirect(url_for('register'))

            action_permissions = None
            page_permissions = None
            if use_custom_permissions:
                action_permissions = {}
                for action_key, _ in PERMISSION_ACTION_OPTIONS:
                    action_permissions[action_key] = request.form.get(f'action_{action_key}') == 'on'

                page_permissions = {}
                for endpoint_name, _ in PERMISSION_PAGE_OPTIONS:
                    page_permissions[endpoint_name] = request.form.get(f'page_{endpoint_name}') == 'on'

            us.add_user(
                username,
                password,
                name,
                last_name,
                is_student=False,
                student_card_id=None,
                max_borrow_days=None,
                permission_preset=permission_preset,
                action_permissions=action_permissions,
                page_permissions=page_permissions,
            )
            return redirect(url_for('home'))
        return render_template(
            'register.html',
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
            student_default_borrow_days=cfg.STUDENT_DEFAULT_BORROW_DAYS,
            student_max_borrow_days=cfg.STUDENT_MAX_BORROW_DAYS
        )
    flash('Sie sind nicht berechtigt, diese Seite anzuzeigen', 'error')
    return redirect(url_for('login'))


@app.route('/user_del', methods=['GET'])
def user_del():
    """
    User deletion interface.
    Displays a list of users that can be deleted by an administrator.
    Prevents self-deletion by hiding the current user from the list.
    
    Returns:
        flask.Response: Rendered template with user list or redirect
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    all_users = us.get_all_users()

    users_list = []
    for user in all_users:
        username = None
        for field in ['Username']:
            if field in user:
                username = user[field]
                break
                
        if username and username != session['username']:
            try:
                permissions_payload = us.get_effective_permissions(username)
            except Exception:
                permissions_payload = us.build_default_permission_payload('standard_user')
            try:
                name = us.get_name(username)
                last_name = us.get_last_name(username)
                if name and last_name:
                    fullname = f"{name} {last_name}"
                elif name:
                    fullname = name
                elif last_name:
                    fullname = last_name
                else:
                    fullname = None
            except:
                name = ""
                last_name = ""
                fullname = None
            users_list.append({
                'username': username,
                'admin': user.get('Admin', False),
                'fullname': fullname,
                'name': name,
                'last_name': last_name,
                'permission_preset': permissions_payload.get('preset', 'standard_user'),
                'action_permissions': permissions_payload.get('actions', {}),
                'page_permissions': permissions_payload.get('pages', {}),
            })
    
    return render_template(
        'user_del.html',
        users=users_list,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards')
    )


@app.route('/delete_user', methods=['POST'])
def delete_user():
    """
    Process user deletion request.
    Deletes a specified user from the system.
    Includes safety checks to prevent self-deletion.
    
    Returns:
        flask.Response: Redirect to the user deletion interface with status
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    username = request.form.get('username')
    if not username:
        flash('Kein Benutzer ausgewählt', 'error')
        return redirect(url_for('user_del'))
    
    # Prevent self-deletion
    if username == session['username']:
        flash('Sie können Ihr eigenes Konto nicht löschen', 'error')
        return redirect(url_for('user_del'))
    
    # Reset this user's borrowings and free items before deleting the user
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        items_col = db['items']
        now = datetime.datetime.now()

        # Complete all active borrowings of this user
        ausleihungen.update_many(
            {'User': username, 'Status': 'active'},
            {'$set': {'Status': 'completed', 'End': now, 'LastUpdated': now}}
        )

        # Cancel all planned borrowings of this user
        ausleihungen.update_many(
            {'User': username, 'Status': 'planned'},
            {'$set': {'Status': 'cancelled', 'LastUpdated': now}}
        )

        # Free all items currently associated with this user
        items_col.update_many(
            {'User': username},
            {'$set': {'Verfuegbar': True, 'LastUpdated': now}, '$unset': {'User': ""}}
        )

        client.close()
    except Exception as e:
        flash(f'Warnung: Ausleihungen/Reservierungen für {username} konnten nicht vollständig zurückgesetzt werden: {str(e)}', 'warning')

    # Delete the user
    try:
        us.delete_user(username)
        flash(f'Benutzer {username} erfolgreich gelöscht', 'success')
    except Exception as e:
        flash(f'Fehler beim Löschen des Benutzers: {str(e)}', 'error')
    
    return redirect(url_for('user_del'))


@app.route('/admin/borrowings')
def admin_borrowings():
    """
    Admin view: list all active and planned borrowings with ability to reset.
    """
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    _ensure_audit_indexes_once()

    client = MongoClient(MONGODB_HOST, MONGODB_PORT)
    db = client[MONGODB_DB]
    ausleihungen = db['ausleihungen']
    items_col = db['items']

    # Load active and planned borrowings
    records = list(ausleihungen.find({'Status': {'$in': ['active', 'planned']}}).sort('Start', -1))

    def fmt_dt(dt):
        try:
            return dt.strftime('%d.%m.%Y %H:%M') if dt else ''
        except Exception:
            return str(dt) if dt else ''

    def fmt_money(value):
        return _format_money_value(value)

    entries = []
    for r in records:
        it_id = r.get('Item')
        item_doc = it.get_item(it_id)
        invoice_data = r.get('InvoiceData') or {}
        try:
            item_code = item_doc.get('Code_4')
            item_name = item_doc.get('Name')
            item_cost = item_doc.get('Anschaffungskosten')
            condition_value = str(item_doc.get('Condition', '')).strip().lower()
            has_damage = bool(item_doc.get('HasDamage')) or condition_value == 'destroyed' or bool(item_doc.get('DamageReports'))
        except:
            item_code = None
            item_name = None
            item_cost = None
            has_damage = False
        entries.append({
            'id': str(r.get('_id')),
            'item_id': str(item_doc.get('_id')) if item_doc and item_doc.get('_id') else str(it_id or ''),
            'item_code': str(item_code) if item_code is not None else '',
            'item_name': str(item_name or ''),
            'item_cost': fmt_money(item_cost),
            'item_cost_raw': item_cost if item_cost is not None else '',
            'user': r.get('User', ''),
            'status': r.get('Status', ''),
            'start': fmt_dt(r.get('Start')),
            'end': fmt_dt(r.get('End')),
            'period': r.get('Period') if r.get('Period') is not None else '',
            'notes': r.get('Notes', ''),
            'invoice_number': invoice_data.get('invoice_number', ''),
            'invoice_amount': fmt_money(invoice_data.get('amount')) if invoice_data.get('amount') is not None else fmt_money(item_cost),
            'invoice_reason': invoice_data.get('damage_reason', ''),
            'invoice_created_at': fmt_dt(invoice_data.get('created_at')) if isinstance(invoice_data.get('created_at'), datetime.datetime) else '',
            'invoice_paid': bool(invoice_data.get('paid', False)),
            'invoice_paid_at': fmt_dt(invoice_data.get('paid_at')) if isinstance(invoice_data.get('paid_at'), datetime.datetime) else '',
            'invoice_corrections_count': len(r.get('InvoiceCorrections', []) or []),
            'has_damage': has_damage,
        })

    client.close()

    return render_template(
        'admin_borrowings.html',
        entries=entries,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards')
    )

"""-----------------------------------------------------------Audit Routes-------------------------------------------------------"""

@app.route('/admin/audit/verify', methods=['GET'])
def admin_verify_audit_chain():
    """Admin endpoint to verify audit chain integrity."""
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        al.ensure_audit_indexes(db)
        result = al.verify_audit_chain(db)
        status_code = 200 if result.get('ok') else 409
        return jsonify(result), status_code
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 500
    finally:
        if client:
            client.close()

@app.route('/admin/audit', methods=['GET'])
def admin_audit_dashboard():
    """Admin dashboard for audit chain status and recent events."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        al.ensure_audit_indexes(db)
        verify_result = al.verify_audit_chain(db)

        audit_rows = list(
            db['audit_log'].find(
                {},
                {
                    'chain_index': 1,
                    'event_type': 1,
                    'actor': 1,
                    'source': 1,
                    'ip': 1,
                    'timestamp': 1,
                    'created_at': 1,
                    'entry_hash': 1,
                    'prev_hash': 1,
                    'payload': 1,
                }
            ).sort('chain_index', -1).limit(200)
        )

        return render_template(
            'admin_audit.html',
            verify_result=verify_result,
            audit_rows=audit_rows,
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        )
    except Exception as exc:
        app.logger.error(f"Error loading audit dashboard: {exc}")
        flash('Fehler beim Laden des Audit-Dashboards.', 'error')
        return redirect(url_for('home_admin'))
    finally:
        if client:
            client.close()

@app.route('/admin/audit/export/pdf/official', methods=['GET'])
def admin_audit_export_pdf_official():
    """Export audit report as professional PDF (Official Report - full DIN 5008 compliant)."""
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    try:
        limit = int((request.args.get('limit') or '1000').strip())
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        al.ensure_audit_indexes(db)
        verify_result = al.verify_audit_chain(db)

        event_counts = list(
            db['audit_log'].aggregate([
                {'$group': {'_id': '$event_type', 'count': {'$sum': 1}}},
                {'$project': {'_id': 0, 'event_type': {'$ifNull': ['$_id', 'unknown']}, 'count': 1}},
                {'$sort': {'count': -1, 'event_type': 1}}
            ])
        )

        audit_rows = list(
            db['audit_log'].find(
                {},
                {
                    'chain_index': 1,
                    'event_type': 1,
                    'actor': 1,
                    'source': 1,
                    'ip': 1,
                    'timestamp': 1,
                    'created_at': 1,
                    'entry_hash': 1,
                    'prev_hash': 1,
                    'payload': 1,
                }
            ).sort('chain_index', -1).limit(limit)
        )

        # Get school information from settings or use defaults
        school_info = _get_school_info_for_export()

        # Generate PDF
        pdf_content = pdf_export.generate_audit_pdf(
            verify_result=verify_result,
            event_counts=event_counts,
            audit_rows=audit_rows,
            export_type='official',
            school_info=school_info
        )

        response = make_response(pdf_content)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=audit-official-report-{datetime.datetime.utcnow().strftime("%Y%m%d-%H%M%S")}.pdf'
        return response

    except Exception as exc:
        app.logger.error(f"PDF Official Report export error: {str(exc)}\n{traceback.format_exc()}")
        return jsonify({'ok': False, 'error': str(exc)}), 500
    finally:
        if client:
            client.close()

"""-----------------------------------------------------------Image Cache Management Routes-------------------------------------------------------"""

@app.route('/admin/image_cache_stats', methods=['GET'])
def admin_image_cache_stats():
    """
    Get statistics about optimized image cache.
    Admin-only endpoint for monitoring and maintenance.
    
    Returns:
        JSON with cache statistics (file count, total size, creation dates)
    """
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    
    permissions = _get_current_user_permissions()
    if not _action_access_allowed(permissions, 'can_manage_settings'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    
    try:
        cache_dir = os.path.join(app.config['THUMBNAIL_FOLDER'], 'optimized_480p')
        
        if not os.path.exists(cache_dir):
            return jsonify({
                'ok': True,
                'cache_exists': False,
                'file_count': 0,
                'total_size_mb': 0
            })
        
        files = []
        total_size = 0
        
        for filename in os.listdir(cache_dir):
            file_path = os.path.join(cache_dir, filename)
            if not os.path.isfile(file_path):
                continue
            
            file_size = os.path.getsize(file_path)
            total_size += file_size
            mod_time = os.path.getmtime(file_path)
            
            files.append({
                'name': filename,
                'size_kb': round(file_size / 1024, 2),
                'modified': datetime.datetime.fromtimestamp(mod_time).isoformat()
            })
        
        files.sort(key=lambda x: x['modified'], reverse=True)
        
        return jsonify({
            'ok': True,
            'cache_exists': True,
            'file_count': len(files),
            'total_size_mb': round(total_size / (1024 * 1024), 2),
            'files': files[:20]  # Return only newest 20 files
        })
    except Exception as e:
        app.logger.error(f"Error getting cache stats: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/image_cache_cleanup', methods=['POST'])
def admin_image_cache_cleanup():
    """
    Trigger cleanup of old optimized images.
    Admin-only endpoint for maintenance.
    
    Args (via form):
        max_age_days: Delete images older than this many days (default 30)
        
    Returns:
        JSON with cleanup results (deleted count, freed space)
    """
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    
    permissions = _get_current_user_permissions()
    if not _action_access_allowed(permissions, 'can_manage_settings'):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    
    try:
        max_age_days = int(request.form.get('max_age_days', 30))
        max_age_days = max(1, min(max_age_days, 365))  # Clamp between 1 and 365 days
        
        result = cleanup_old_optimized_images(max_age_days)
        
        if result['error']:
            return jsonify({'ok': False, 'error': result['error']}), 500
        
        # Log the action
        _append_audit_event_standalone(
            event_type='admin_image_cache_cleanup',
            payload={
                'max_age_days': max_age_days,
                'deleted_count': result['deleted'],
                'freed_mb': result['freed_mb']
            }
        )
        
        return jsonify({
            'ok': True,
            'deleted': result['deleted'],
            'freed_mb': result['freed_mb'],
            'message': f"Cleaned up {result['deleted']} images, freed {result['freed_mb']} MB"
        })
    except Exception as e:
        app.logger.error(f"Error during image cache cleanup: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)}), 500

"""-----------------------------------------------------------Borrowing Management Routes-------------------------------------------------------"""

@app.route('/admin/reset_borrowing/<borrow_id>', methods=['POST'])
def admin_reset_borrowing(borrow_id):
    """
    Admin action: reset a single borrowing.
    - If active: complete it and free the item
    - If planned: cancel it
    """
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        items_col = db['items']

        rec = ausleihungen.find_one({'_id': ObjectId(borrow_id)})
        if not rec:
            client.close()
            flash('Ausleihung nicht gefunden', 'error')
            return redirect(url_for('admin_borrowings'))

        status = rec.get('Status')
        item_id = rec.get('Item')
        user = rec.get('User')

        now = datetime.datetime.now()
        if status == 'active':
            ausleihungen.update_one({'_id': rec['_id']}, {'$set': {'Status': 'completed', 'End': now, 'LastUpdated': now}})
            # Free the item
            if item_id:
                try:
                    items_col.update_one({'_id': ObjectId(item_id)}, {'$set': {'Verfuegbar': True, 'LastUpdated': now}, '$unset': {'User': ""}})
                except Exception:
                    pass
            flash('Aktive Ausleihe wurde zurückgesetzt (abgeschlossen).', 'success')
            _append_audit_event_standalone(
                event_type='ausleihung_admin_reset',
                payload={
                    'borrow_id': borrow_id,
                    'from_status': 'active',
                    'to_status': 'completed',
                    'item_id': str(item_id or ''),
                    'borrower': str(user or ''),
                }
            )
        elif status == 'planned':
            ausleihungen.update_one({'_id': rec['_id']}, {'$set': {'Status': 'cancelled', 'LastUpdated': now}})
            flash('Geplante Ausleihe wurde storniert.', 'success')
            _append_audit_event_standalone(
                event_type='ausleihung_admin_reset',
                payload={
                    'borrow_id': borrow_id,
                    'from_status': 'planned',
                    'to_status': 'cancelled',
                    'item_id': str(item_id or ''),
                    'borrower': str(user or ''),
                }
            )
        else:
            flash('Diese Ausleihe ist weder aktiv noch geplant.', 'warning')

        client.close()
    except Exception as e:
        flash(f'Fehler beim Zurücksetzen: {str(e)}', 'error')

    return redirect(url_for('admin_borrowings'))


@app.route('/admin/borrowings/<borrow_id>/invoice', methods=['POST'])
def admin_create_invoice(borrow_id):
    """Create a PDF invoice for a destroyed borrowed item."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        items_col = db['items']

        borrow_doc = ausleihungen.find_one({'_id': ObjectId(borrow_id)})
        if not borrow_doc:
            flash('Ausleihung nicht gefunden.', 'error')
            return redirect(url_for('admin_borrowings'))

        item_id = borrow_doc.get('Item')
        item_doc = None
        if item_id:
            try:
                item_doc = items_col.find_one({'_id': ObjectId(item_id)})
            except Exception:
                item_doc = items_col.find_one({'_id': item_id})

        if not item_doc:
            flash('Zugehöriges Element nicht gefunden.', 'error')
            return redirect(url_for('admin_borrowings'))

        if borrow_doc.get('Status') not in ['active', 'completed'] and not borrow_doc.get('InvoiceData'):
            flash('Eine Rechnung kann nur für aktive oder bereits dokumentierte Ausleihungen erstellt werden.', 'warning')
            return redirect(url_for('admin_borrowings'))

        amount_raw = request.form.get('invoice_amount') or request.form.get('amount') or item_doc.get('Anschaffungskosten')
        amount_value = _parse_money_value(amount_raw)
        if amount_value is None or amount_value <= 0:
            flash('Bitte einen gültigen Rechnungsbetrag angeben.', 'error')
            return redirect(url_for('admin_borrowings'))

        damage_reason = str(request.form.get('damage_reason', '')).strip() or 'Das ausgeliehene Element wurde beschädigt oder zerstört.'
        if len(damage_reason) > 2000:
            flash('Die Schadensbeschreibung ist zu lang.', 'error')
            return redirect(url_for('admin_borrowings'))

        mark_destroyed = request.form.get('mark_destroyed') == 'on'
        close_borrowing = request.form.get('close_borrowing') == 'on'
        now = datetime.datetime.now()

        existing_invoice = borrow_doc.get('InvoiceData') or {}
        if existing_invoice.get('invoice_number'):
            flash('Für diese Ausleihe existiert bereits eine Rechnung. Bitte Korrekturbuchung verwenden.', 'warning')
            return redirect(url_for('admin_borrowings'))

        invoice_number = existing_invoice.get('invoice_number') or _build_invoice_number(borrow_doc['_id'], now)
        borrower = borrow_doc.get('User', '')
        item_name = item_doc.get('Name', '')
        item_code = item_doc.get('Code_4', '')

        invoice_data = {
            'invoice_number': invoice_number,
            'amount': round(amount_value, 2),
            'amount_text': _format_money_value(amount_value),
            'damage_reason': damage_reason,
            'created_at': now,
            'created_by': session.get('username', ''),
            'borrower': borrower,
            'item_id': str(item_doc['_id']),
            'item_name': item_name,
            'item_code': item_code,
            'mark_destroyed': mark_destroyed,
            'status_before_invoice': borrow_doc.get('Status', ''),
            'paid': False,
            'paid_at': None,
            'paid_by': '',
        }

        update_fields = {
            'InvoiceData': invoice_data,
            'InvoiceLocked': True,
            'LastUpdated': now,
        }
        if close_borrowing:
            update_fields['Status'] = 'completed'
            update_fields['End'] = now

        ausleihungen.update_one({'_id': borrow_doc['_id']}, {'$set': update_fields})

        if mark_destroyed:
            damage_entry = {
                'description': damage_reason,
                'reported_by': session['username'],
                'reported_at': now,
                'invoice_number': invoice_number,
                'invoice_amount': round(amount_value, 2),
                'source': 'invoice'
            }
            items_col.update_one(
                {'_id': item_doc['_id']},
                {
                    '$push': {'DamageReports': {'$each': [damage_entry], '$position': 0}},
                    '$set': {
                        'HasDamage': True,
                        'Verfuegbar': False,
                        'LastUpdated': now,
                        'Condition': 'destroyed'
                    },
                    '$unset': {'User': ''}
                }
            )

        try:
            logs_collection = db['system_logs']
            logs_collection.insert_one({
                'type': 'damage_invoice',
                'timestamp': now.isoformat(),
                'user': session.get('username'),
                'borrow_id': borrow_id,
                'item_id': str(item_doc['_id']),
                'item_name': item_name,
                'invoice_number': invoice_number,
                'amount': round(amount_value, 2),
                'ip': request.remote_addr,
            })
        except Exception as log_err:
            app.logger.warning(f"Damage invoice log write failed for borrow {borrow_id}: {log_err}")

        _append_audit_event_standalone(
            event_type='invoice_created',
            payload={
                'borrow_id': borrow_id,
                'invoice_number': invoice_number,
                'amount': round(amount_value, 2),
                'mark_destroyed': mark_destroyed,
                'close_borrowing': close_borrowing,
                'item_id': str(item_doc.get('_id')),
            }
        )

        pdf_buffer = pdf_export._build_invoice_pdf(invoice_data)
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'rechnung_{invoice_number}.pdf'
        )
    except Exception as e:
        app.logger.error(f"Error creating damage invoice for borrow {borrow_id}: {e}")
        flash('Fehler beim Erstellen der Rechnung.', 'error')
        return redirect(url_for('admin_borrowings'))
    finally:
        if client:
            client.close()


@app.route('/admin/borrowings/<borrow_id>/invoice/mark-paid', methods=['POST'])
def admin_mark_invoice_paid(borrow_id):
    """Mark an existing invoice as paid."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']

        borrow_doc = ausleihungen.find_one({'_id': ObjectId(borrow_id)}, {'InvoiceData': 1})
        if not borrow_doc:
            flash('Ausleihung nicht gefunden.', 'error')
            return redirect(url_for('admin_borrowings'))

        invoice_data = borrow_doc.get('InvoiceData') or {}
        if not invoice_data:
            flash('Für diese Ausleihung existiert keine Rechnung.', 'warning')
            return redirect(url_for('admin_borrowings'))

        if invoice_data.get('paid') is True:
            flash('Rechnung ist bereits als bezahlt markiert.', 'info')
            return redirect(url_for('admin_borrowings'))

        now = datetime.datetime.now()
        result = ausleihungen.update_one(
            {'_id': borrow_doc['_id']},
            {
                '$set': {
                    'InvoiceData.paid': True,
                    'InvoiceData.paid_at': now,
                    'InvoiceData.paid_by': session.get('username', ''),
                    'LastUpdated': now,
                }
            }
        )

        if result.matched_count == 0:
            flash('Ausleihung nicht gefunden.', 'error')
            return redirect(url_for('admin_borrowings'))

        try:
            logs_collection = db['system_logs']
            logs_collection.insert_one({
                'type': 'damage_invoice_paid',
                'timestamp': now.isoformat(),
                'user': session.get('username'),
                'borrow_id': borrow_id,
                'invoice_number': invoice_data.get('invoice_number', ''),
                'amount': invoice_data.get('amount'),
                'ip': request.remote_addr,
            })
        except Exception as log_err:
            app.logger.warning(f"Damage invoice paid log write failed for borrow {borrow_id}: {log_err}")

        _append_audit_event_standalone(
            event_type='invoice_marked_paid',
            payload={
                'borrow_id': borrow_id,
                'invoice_number': invoice_data.get('invoice_number', ''),
                'amount': invoice_data.get('amount'),
            }
        )

        flash('Rechnung wurde als bezahlt markiert.', 'success')
        return redirect(url_for('admin_borrowings'))
    except Exception as e:
        app.logger.error(f"Error marking invoice paid for borrow {borrow_id}: {e}")
        flash('Fehler beim Markieren als bezahlt.', 'error')
        return redirect(url_for('admin_borrowings'))
    finally:
        if client:
            client.close()


@app.route('/admin/borrowings/<borrow_id>/invoice/finalize', methods=['POST'])
def admin_finalize_invoice_and_repair(borrow_id):
    """Mark invoice as paid and item as repaired in one action."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        items_col = db['items']

        borrow_doc = ausleihungen.find_one({'_id': ObjectId(borrow_id)})
        if not borrow_doc:
            flash('Ausleihung nicht gefunden.', 'error')
            return redirect(url_for('admin_borrowings'))

        invoice_data = borrow_doc.get('InvoiceData') or {}
        if not invoice_data:
            flash('Für diese Ausleihung existiert keine Rechnung.', 'warning')
            return redirect(url_for('admin_borrowings'))

        item_doc = None
        item_id = borrow_doc.get('Item')
        if item_id:
            try:
                item_doc = items_col.find_one({'_id': ObjectId(item_id)})
            except Exception:
                item_doc = items_col.find_one({'_id': item_id})

        now = datetime.datetime.now()

        update_fields = {
            'LastUpdated': now,
        }
        if invoice_data.get('paid') is not True:
            update_fields.update({
                'InvoiceData.paid': True,
                'InvoiceData.paid_at': now,
                'InvoiceData.paid_by': session.get('username', ''),
            })
        if borrow_doc.get('Status') == 'active':
            update_fields['Status'] = 'completed'
            update_fields['End'] = now

        ausleihungen.update_one({'_id': borrow_doc['_id']}, {'$set': update_fields})

        repaired = False
        resolved_count = 0
        if item_doc:
            open_reports = item_doc.get('DamageReports', []) or []
            resolved_count = len(open_reports)
            item_update = {
                '$set': {
                    'DamageReports': [],
                    'HasDamage': False,
                    'Verfuegbar': True,
                    'LastUpdated': now,
                },
                '$unset': {
                    'User': '',
                    'Condition': '',
                },
            }
            if open_reports:
                repair_entry = {
                    'repaired_by': session['username'],
                    'repaired_at': now,
                    'resolved_reports': open_reports,
                }
                item_update['$push'] = {'DamageRepairs': {'$each': [repair_entry], '$position': 0}}

            item_result = items_col.update_one({'_id': item_doc['_id']}, item_update)
            repaired = item_result.matched_count > 0

        try:
            logs_collection = db['system_logs']
            logs_collection.insert_one({
                'type': 'damage_invoice_finalize',
                'timestamp': now.isoformat(),
                'user': session.get('username'),
                'borrow_id': borrow_id,
                'item_id': str(item_doc.get('_id')) if item_doc else '',
                'invoice_number': invoice_data.get('invoice_number', ''),
                'amount': invoice_data.get('amount'),
                'repaired': repaired,
                'resolved_damage_reports': resolved_count,
                'ip': request.remote_addr,
            })
        except Exception as log_err:
            app.logger.warning(f"Damage invoice finalize log write failed for borrow {borrow_id}: {log_err}")

        _append_audit_event_standalone(
            event_type='invoice_finalized_and_repaired',
            payload={
                'borrow_id': borrow_id,
                'item_id': str(item_doc.get('_id')) if item_doc else '',
                'invoice_number': invoice_data.get('invoice_number', ''),
                'amount': invoice_data.get('amount'),
                'repaired': repaired,
                'resolved_damage_reports': resolved_count,
            }
        )

        if repaired:
            flash('Rechnung als bezahlt markiert und Element als repariert abgeschlossen.', 'success')
        else:
            flash('Rechnung als bezahlt markiert. Element konnte nicht repariert werden (nicht gefunden).', 'warning')
        return redirect(url_for('admin_borrowings'))
    except Exception as e:
        app.logger.error(f"Error finalizing invoice/repair for borrow {borrow_id}: {e}")
        flash('Fehler beim Kombinieren von bezahlt und repariert.', 'error')
        return redirect(url_for('admin_borrowings'))
    finally:
        if client:
            client.close()


@app.route('/admin/borrowings/<borrow_id>/invoice/pdf', methods=['GET'])
def admin_view_invoice_pdf(borrow_id):
    """View a previously created invoice PDF for a borrowing."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']
        items_col = db['items']

        borrow_doc = ausleihungen.find_one({'_id': ObjectId(borrow_id)}, {'InvoiceData': 1, 'Item': 1, 'User': 1})
        if not borrow_doc:
            flash('Ausleihung nicht gefunden.', 'error')
            return redirect(url_for('library_loans_admin'))

        invoice_data = borrow_doc.get('InvoiceData') or {}
        if not invoice_data:
            flash('Für diese Ausleihe wurde noch keine Rechnung erstellt.', 'warning')
            return redirect(url_for('library_loans_admin'))

        item_doc = None
        item_id = borrow_doc.get('Item')
        if item_id:
            try:
                item_doc = items_col.find_one({'_id': ObjectId(item_id)})
            except Exception:
                item_doc = items_col.find_one({'_id': item_id})

        pdf_payload = _prepare_invoice_pdf_payload(invoice_data, borrow_doc=borrow_doc, item_doc=item_doc)
        invoice_number = pdf_payload.get('invoice_number', 'rechnung')
        pdf_buffer = pdf_export._build_invoice_pdf(pdf_payload)
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'rechnung_{invoice_number}.pdf'
        )
    except Exception as e:
        app.logger.error(f"Error loading stored invoice PDF for borrow {borrow_id}: {e}")
        flash('Fehler beim Öffnen der Rechnung.', 'error')
        return redirect(url_for('library_loans_admin'))
    finally:
        if client:
            client.close()


@app.route('/admin/borrowings/<borrow_id>/invoice/correction', methods=['POST'])
def admin_add_invoice_correction(borrow_id):
    """Append an invoice correction entry without mutating the original invoice body."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        ausleihungen = db['ausleihungen']

        borrow_doc = ausleihungen.find_one({'_id': ObjectId(borrow_id)}, {'InvoiceData': 1})
        if not borrow_doc:
            flash('Ausleihung nicht gefunden.', 'error')
            return redirect(url_for('admin_borrowings'))

        invoice_data = borrow_doc.get('InvoiceData') or {}
        if not invoice_data:
            flash('Korrektur nicht möglich: Es existiert noch keine Rechnung.', 'warning')
            return redirect(url_for('admin_borrowings'))

        correction_reason = str(request.form.get('correction_reason', '')).strip()
        if not correction_reason:
            flash('Bitte eine Begründung für die Korrektur angeben.', 'error')
            return redirect(url_for('admin_borrowings'))

        delta_raw = request.form.get('amount_delta')
        delta_value = _parse_money_value(delta_raw) if delta_raw not in (None, '') else 0.0
        if delta_value is None:
            flash('Ungültiger Korrekturbetrag.', 'error')
            return redirect(url_for('admin_borrowings'))

        now = datetime.datetime.now()
        correction_number = f"CORR-{now.strftime('%Y%m%d-%H%M%S')}-{str(borrow_doc.get('_id'))[-6:].upper()}"
        correction_entry = {
            'correction_number': correction_number,
            'reason': correction_reason,
            'amount_delta': round(delta_value, 2),
            'amount_delta_text': _format_money_value(delta_value),
            'created_at': now,
            'created_by': session.get('username', ''),
            'invoice_number': invoice_data.get('invoice_number', ''),
        }

        ausleihungen.update_one(
            {'_id': borrow_doc['_id']},
            {
                '$push': {'InvoiceCorrections': {'$each': [correction_entry], '$position': 0}},
                '$set': {'LastUpdated': now, 'InvoiceLocked': True}
            }
        )

        _append_audit_event_standalone(
            event_type='invoice_correction_added',
            payload={
                'borrow_id': borrow_id,
                'invoice_number': invoice_data.get('invoice_number', ''),
                'correction_number': correction_number,
                'amount_delta': round(delta_value, 2),
                'reason': correction_reason,
            }
        )

        flash('Korrekturbuchung wurde revisionssicher ergänzt.', 'success')
        return redirect(url_for('admin_borrowings'))
    except Exception as e:
        app.logger.error(f"Error creating invoice correction for borrow {borrow_id}: {e}")
        flash('Fehler beim Anlegen der Korrekturbuchung.', 'error')
        return redirect(url_for('admin_borrowings'))
    finally:
        if client:
            client.close()


@app.route('/admin/library/items/<item_id>/invoices', methods=['GET'])
def library_item_invoices(item_id):
    """Show all stored invoices for one specific library item."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not cfg.MODULES.is_enabled('library'):
        flash('Bibliotheks-Modul ist deaktiviert.', 'error')
        return redirect(url_for('home_admin'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        ausleihungen = db['ausleihungen']

        try:
            item_doc = items_col.find_one({'_id': ObjectId(item_id)})
        except Exception:
            item_doc = items_col.find_one({'_id': item_id})

        if not item_doc:
            flash('Bibliotheksmedium nicht gefunden.', 'error')
            return redirect(url_for('library_loans_admin'))

        borrow_docs = list(ausleihungen.find(
            {
                'Item': str(item_doc.get('_id')),
                'InvoiceData': {'$exists': True, '$ne': {}}
            },
            {
                'Status': 1,
                'User': 1,
                'Start': 1,
                'End': 1,
                'InvoiceData': 1,
            }
        ).sort('InvoiceData.created_at', -1))

        entries = []
        for borrow_doc in borrow_docs:
            invoice_data = borrow_doc.get('InvoiceData') or {}
            created_at = invoice_data.get('created_at')
            created_at_display = created_at.strftime('%d.%m.%Y %H:%M') if isinstance(created_at, datetime.datetime) else (str(created_at) if created_at else '')
            paid_at = invoice_data.get('paid_at')
            paid_at_display = paid_at.strftime('%d.%m.%Y %H:%M') if isinstance(paid_at, datetime.datetime) else (str(paid_at) if paid_at else '')

            entries.append({
                'borrow_id': str(borrow_doc.get('_id')),
                'borrow_status': borrow_doc.get('Status', ''),
                'borrow_user': borrow_doc.get('User', ''),
                'borrow_start': borrow_doc.get('Start').strftime('%d.%m.%Y %H:%M') if isinstance(borrow_doc.get('Start'), datetime.datetime) else '',
                'borrow_end': borrow_doc.get('End').strftime('%d.%m.%Y %H:%M') if isinstance(borrow_doc.get('End'), datetime.datetime) else '',
                'invoice_number': invoice_data.get('invoice_number', ''),
                'invoice_amount': _format_money_value(invoice_data.get('amount')),
                'invoice_reason': invoice_data.get('damage_reason', ''),
                'invoice_created_at': created_at_display,
                'invoice_created_by': invoice_data.get('created_by', ''),
                'invoice_paid': bool(invoice_data.get('paid', False)),
                'invoice_paid_at': paid_at_display,
                'invoice_paid_by': invoice_data.get('paid_by', ''),
            })

        return render_template(
            'library_item_invoices.html',
            item={
                'id': str(item_doc.get('_id')),
                'name': item_doc.get('Name', ''),
                'code': item_doc.get('Code_4', ''),
                'author': item_doc.get('Author', ''),
                'isbn': item_doc.get('ISBN', ''),
            },
            invoices=entries,
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        )
    except Exception as e:
        app.logger.error(f"Error loading invoice history for item {item_id}: {e}")
        flash('Fehler beim Laden der Rechnungshistorie.', 'error')
        return redirect(url_for('library_loans_admin'))
    finally:
        if client:
            client.close()


@app.route('/admin_reset_user_password', methods=['POST'])
def admin_reset_user_password():
    """
    Admin route to reset a user's password.
    Resets the password for the specified user to a temporary password.
    Only accessible by administrators.
    
    Returns:
        flask.Response: Redirect to user management page with status message
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adresse zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    username = request.form.get('username')
    new_password = html.escape(request.form.get('new_password', 'Password123'))  # Default temporary password
    
    if not username:
        flash('Kein Benutzer ausgewählt', 'error')
        return redirect(url_for('user_del'))
    
    # Check if user exists
    user = us.get_user(username)
    if not user:
        flash(f'Benutzer {username} nicht gefunden', 'error')
        return redirect(url_for('user_del'))
    
    # Prevent changing own password through this route (use change_password instead)
    if username == session['username']:
        flash('Sie können Ihr eigenes Passwort nicht über diese Funktion ändern. Bitte verwenden Sie dafür die Option "Passwort ändern" im Profil-Menü.', 'error')
        return redirect(url_for('user_del'))
    
    # Reset the password
    try:
        us.update_password(username, new_password)
        flash(f'Passwort für {username} wurde erfolgreich zurückgesetzt auf: {new_password}', 'success')
    except Exception as e:
        flash(f'Fehler beim Zurücksetzen des Passworts: {str(e)}', 'error')
    
    return redirect(url_for('user_del'))


@app.route('/admin_update_user_name', methods=['POST'])
def admin_update_user_name():
    """
    Admin route to update a user's name details.
    
    Returns:
        flask.Response: Redirect to user management page
    """
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Nicht autorisierter Zugriff', 'error')
        return redirect(url_for('login'))
        
    username = request.form.get('username')
    name = html.escape(request.form.get('name'))
    last_name = html.escape(request.form.get('last_name'))
    
    if not username:
        flash('Kein Benutzer ausgewählt', 'error')
        return redirect(url_for('user_del'))
        
    if us.update_user_name(username, name, last_name):
        flash(f'Name für {username} wurde erfolgreich aktualisiert.', 'success')
    else:
        flash(f'Fehler beim Aktualisieren des Namens.', 'error')
        
    return redirect(url_for('user_del'))


@app.route('/admin_update_user_permissions', methods=['POST'])
def admin_update_user_permissions():
    """Admin route to update permission preset and per-endpoint overrides for a user."""
    if 'username' not in session:
        flash('Nicht autorisierter Zugriff', 'error')
        return redirect(url_for('login'))

    username = request.form.get('username', '').strip()
    preset_key = request.form.get('permission_preset', 'standard_user').strip()

    if not username:
        flash('Kein Benutzer ausgewählt', 'error')
        return redirect(url_for('user_del'))

    target_user = us.get_user(username)
    if not target_user:
        flash(f'Benutzer {username} nicht gefunden', 'error')
        return redirect(url_for('user_del'))

    action_permissions = {}
    for action_key, _ in PERMISSION_ACTION_OPTIONS:
        action_permissions[action_key] = request.form.get(f'action_{action_key}') == 'on'

    page_permissions = {}
    for endpoint_name, _ in PERMISSION_PAGE_OPTIONS:
        page_permissions[endpoint_name] = request.form.get(f'page_{endpoint_name}') == 'on'

    if us.update_user_permissions(username, preset_key, action_permissions, page_permissions):
        flash(f'Berechtigungen für {username} wurden aktualisiert.', 'success')
    else:
        flash('Fehler beim Aktualisieren der Berechtigungen.', 'error')

    return redirect(url_for('user_del'))


@app.route('/admin_anonymize_names', methods=['POST'])
def admin_anonymize_names():
    """Anonymize already stored personal names into short aliases."""
    if 'username' not in session:
        flash('Nicht autorisierter Zugriff', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        users_col = db['users']
        student_cards_col = db['student_cards']

        users_updated = 0
        cards_updated = 0

        for user_doc in users_col.find({}, {'name': 1, 'last_name': 1, 'Username': 1, 'username': 1}):
            first = str(user_doc.get('name') or '').strip()
            last = str(user_doc.get('last_name') or '').strip()
            fallback = str(user_doc.get('Username') or user_doc.get('username') or '').strip()

            alias = us.build_name_synonym(first or fallback, last)
            result = users_col.update_one(
                {'_id': user_doc['_id']},
                {'$set': {'name': alias, 'last_name': ''}}
            )
            if result.modified_count > 0:
                users_updated += 1

        for card_doc in student_cards_col.find({}, {'SchülerName': 1, 'Klasse': 1, 'Notizen': 1}):
            decrypted = _decrypt_student_card_doc(card_doc)
            class_name = sanitize_form_value(decrypted.get('Klasse', ''))
            notes = sanitize_form_value(decrypted.get('Notizen', ''))

            encrypted_payload = encrypt_document_fields(
                {
                    'SchülerName': decrypted.get('SchülerName', ''),
                    'Klasse': class_name,
                    'Notizen': notes,
                },
                STUDENT_CARD_ENCRYPTED_FIELDS,
            )

            result = student_cards_col.update_one(
                {'_id': card_doc['_id']},
                {'$set': {'Aktualisiert': datetime.datetime.now(), **encrypted_payload}}
            )
            if result.modified_count > 0:
                cards_updated += 1

        flash(
            f'Anonymisierung abgeschlossen: {users_updated} Benutzer und {cards_updated} Ausweise aktualisiert.',
            'success'
        )
    except Exception as exc:
        app.logger.error(f'Error anonymizing names: {exc}')
        flash('Fehler bei der Anonymisierung der Namen.', 'error')
    finally:
        if client:
            client.close()

    return redirect(url_for('user_del'))


@app.route('/logs')
def logs():
    """
    View system logs interface.
    Displays a history of all item borrowings with detailed information.
    
    Returns:
        flask.Response: Rendered template with logs or redirect if not authenticated
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
        
    # Get ausleihungen
    all_ausleihungen = au.get_ausleihungen()
    
    formatted_items = []
    for ausleihung in all_ausleihungen:
        try:
            # Get item details - from sample data, Item is an ID
            item = it.get_item(ausleihung.get('Item'))
            item_name = item.get('Name', 'Unknown Item') if item else 'Unknown Item'
            
            # Get user details - from sample data, User is a username string
           

            username = ausleihung.get('User', 'Unknown User')
            # Determine (verified) status for display
            try:
                display_status = au.get_current_status(ausleihung)
            except Exception:
                display_status = ausleihung.get('Status', 'unknown')
            
            # Format dates for display
            start_date = ausleihung.get('Start')
            if isinstance(start_date, datetime.datetime):
                start_date = start_date.strftime('%Y-%m-%d %H:%M')
                
            end_date = ausleihung.get('End', 'Not returned')
            if isinstance(end_date, datetime.datetime):
                end_date = end_date.strftime('%Y-%m-%d %H:%M')
            
            # Calculate duration
            duration = 'N/A'
            if isinstance(ausleihung.get('Start'), datetime.datetime) and isinstance(ausleihung.get('End'), datetime.datetime):
                duration_td = ausleihung['End'] - ausleihung['Start']
                hours, remainder = divmod(duration_td.seconds, 3600)
                minutes, _ = divmod(remainder, 60)
                if duration_td.days > 0:
                    duration = f"{duration_td.days}d {hours}h {minutes}m"
                else:
                    duration = f"{hours}h {minutes}m"
            
            formatted_items.append({
                'Item': item_name,
                'User': username,
                'Start': start_date,
                'End': end_date,
                'Duration': duration,
                'Status': display_status,
                'EventType': 'Ausleihe',
                'id': str(ausleihung['_id']),
                'ConflictDetected': ausleihung.get('ConflictDetected', False),
                'ConflictNote': ausleihung.get('ConflictNote', ''),
            })
        except Exception as e:
            continue

    # Add damage + repair entries from system_logs
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        logs_collection = db['system_logs']
        extra_logs = list(logs_collection.find({'type': {'$in': ['damage_report', 'damage_repair']}}))

        for log_item in extra_logs:
            log_type = log_item.get('type', '')
            item_id = log_item.get('item_id')
            item_obj = it.get_item(item_id) if item_id else None
            item_name = item_obj.get('Name', 'Unknown Item') if item_obj else 'Unknown Item'
            ts_raw = log_item.get('timestamp')
            ts_display = ts_raw or '-'
            if isinstance(ts_raw, datetime.datetime):
                ts_display = ts_raw.strftime('%Y-%m-%d %H:%M')
            elif isinstance(ts_raw, str):
                try:
                    ts_display = datetime.datetime.fromisoformat(ts_raw).strftime('%Y-%m-%d %H:%M')
                except Exception:
                    ts_display = ts_raw

            if log_type == 'damage_report':
                note = log_item.get('note', '')
                formatted_items.append({
                    'Item': item_name,
                    'User': log_item.get('user', 'Unknown User'),
                    'Start': ts_display,
                    'End': '-',
                    'Duration': '-',
                    'Status': 'gemeldet',
                    'EventType': 'Schaden',
                    'id': str(log_item.get('_id', '')),
                    'ConflictDetected': False,
                    'ConflictNote': note,
                })
            elif log_type == 'damage_repair':
                resolved_count = log_item.get('resolved_count', 0)
                formatted_items.append({
                    'Item': item_name,
                    'User': log_item.get('user', 'Unknown User'),
                    'Start': ts_display,
                    'End': '-',
                    'Duration': '-',
                    'Status': 'repariert',
                    'EventType': 'Reparatur',
                    'id': str(log_item.get('_id', '')),
                    'ConflictDetected': False,
                    'ConflictNote': f'Erledigte Schäden: {resolved_count}',
                })
        client.close()
    except Exception as e:
        app.logger.warning(f"Could not load damage/repair logs: {e}")

    def parse_sort_date(value):
        if isinstance(value, datetime.datetime):
            return value
        if isinstance(value, str):
            for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d'):
                try:
                    return datetime.datetime.strptime(value, fmt)
                except Exception:
                    continue
            try:
                return datetime.datetime.fromisoformat(value)
            except Exception:
                return datetime.datetime.min
        return datetime.datetime.min

    formatted_items.sort(key=lambda entry: parse_sort_date(entry.get('Start')), reverse=True)
    
    return render_template(
        'logs.html',
        items=formatted_items,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards')
    )


@app.route('/get_logs', methods=['GET'])
def get_logs():
    """
    API endpoint to retrieve all borrowing logs.
    
    Returns:
        dict: Dictionary containing all borrowing records or redirect if not authenticated
    """
    if not session.get('username'):
        return redirect(url_for('login'))
    logs = au.get_ausleihungen()
    return logs


@app.route('/get_usernames', methods=['GET'])
def get_usernames():
    """
    API endpoint to retrieve all usernames from the system.
    Requires administrator privileges.
    
    Returns:
        dict: Dictionary containing all users or redirect if not authenticated
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('logout'))
    elif 'username' in session and us.check_admin(session['username']):
        return jsonify(us.get_all_users())  # Fixed to use get_all_users
    else:
        flash('Bitte melden Sie sich an, um auf diese Funktion zuzugreifen', 'error')
        return redirect(url_for('login'))  # Added proper return

# New routes for filter management

@app.route('/manage_filters')
def manage_filters():
    """
"
    Admin page to manage predefined filter values.
    
    Returns:
        flask.Response: Rendered filter management template or redirect
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    # Get predefined filter values
    filter1_values = it.get_predefined_filter_values(1)
    filter2_values = it.get_predefined_filter_values(2)
    
    return render_template(
        'manage_filters.html',
        filter1_values=filter1_values,
        filter2_values=filter2_values,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards')
    )

@app.route('/add_filter_value/<int:filter_num>', methods=['POST'])
def add_filter_value(filter_num):
    """
    Add a new predefined value to the specified filter.
    
    Args:
        filter_num (int): Filter number (1 or 2)
        
    Returns:
        flask.Response: Redirect to filter management page
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'success': False, 'error': 'Not authorized'}), 403
    
    value = sanitize_form_value(request.form.get('value'))
    
    if not value:
        flash('Bitte geben Sie einen Wert ein', 'error')
        return redirect(url_for('manage_filters'))

    if value == FILTER_SELECT_ALL_TOKEN:
        flash('Dieser Wert ist reserviert und kann nicht als Filterwert gespeichert werden.', 'error')
        return redirect(url_for('manage_filters'))
    
    # Add the value to the filter
    success = it.add_predefined_filter_value(filter_num, value)
    
    if success:
        flash(f'Wert "{value}" wurde zu Filter {filter_num} hinzugefügt', 'success')
    else:
        flash(f'Wert "{value}" existiert bereits in Filter {filter_num}', 'error')
    
    return redirect(url_for('manage_filters'))

@app.route('/remove_filter_value/<int:filter_num>/<string:value>', methods=['POST'])
def remove_filter_value(filter_num, value):
    """
    Remove a predefined value from the specified filter.
    
    Args:
        filter_num (int): Filter number (1 or 2)
        value (str): Value to remove
        
    Returns:
        flask.Response: Redirect to filter management page
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'success': False, 'error': 'Not authorized'}), 403
    
    # Remove the value from the filter
    success = it.remove_predefined_filter_value(filter_num, value)
    
    if success:
        flash(f'Wert "{value}" wurde aus Filter {filter_num} entfernt', 'success')
    else:
        flash(f'Fehler beim Entfernen des Wertes "{value}" aus Filter {filter_num}', 'error')
    
    return redirect(url_for('manage_filters'))

@app.route('/get_predefined_filter_values/<int:filter_num>')
def get_predefined_filter_values(filter_num):
    """
    API endpoint to get predefined values for a specific filter.
    
    Args:
        filter_num (int): Filter number (1 or 2)

        
    Returns:
        dict: Dictionary containing predefined filter values
    """
    values = it.get_predefined_filter_values(filter_num)
    return jsonify({'values': values})

@app.route('/search_word/<path:word>')
def search_word(word):
    """Search items by Titel (Name) and Beschreibung, case-insensitive.

    Returns: JSON with list of matching item IDs.
    """
    try:
        term = (word or "").strip()
        if not term:
            return jsonify({"success": True, "response": []})

        term_lower = term.lower()
        id_set = set()
        for i in it.get_items():
            beschreibung = i.get("Beschreibung", "")
            titel = i.get("Name", "")
            # Normalize Beschreibung to string
            try:
                if isinstance(beschreibung, (list, tuple)):
                    text = " ".join([str(x) for x in beschreibung])
                else:
                    text = str(beschreibung)
            except Exception:
                text = ""

            # Normalize title
            try:
                title_text = str(titel)
            except Exception:
                title_text = ""

            if (term_lower in text.lower()) or (term_lower in title_text.lower()):
                _id = i.get("_id")
                if _id is not None:
                    id_set.add(str(_id))

        return jsonify({"success": True, "response": list(id_set)})
    except Exception as e:
        return jsonify({"success": False, "response": str(e)})

@app.route('/fetch_book_info/<isbn>')
def fetch_book_info(isbn):
    """
    API endpoint to fetch book information by ISBN using Google Books API
    
    Args:
        isbn (str): ISBN to look up
        
    Returns:
        dict: Book information or error message
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({"error": "Not authorized"}), 403

    if not cfg.MODULES.is_enabled('library'):
        return jsonify({"error": "Bibliotheks-Modul ist deaktiviert."}), 403

    try:
        clean_isbn = normalize_and_validate_isbn(isbn)
        if not clean_isbn:
            return jsonify({"error": "Ungültige ISBN. Bitte ISBN-10 oder ISBN-13 verwenden."}), 400

        # First source: Google Books
        response = requests.get(
            f"https://www.googleapis.com/books/v1/volumes?q=isbn:{clean_isbn}",
            timeout=10
        )

        if response.status_code == 200:
            data = response.json()
            if data.get('totalItems', 0) > 0 and data.get('items'):
                book_info = data['items'][0].get('volumeInfo', {})
                sale_info = data['items'][0].get('saleInfo', {})

                price = None
                retail_price = sale_info.get('retailPrice', {})
                list_price = sale_info.get('listPrice', {})
                if retail_price and 'amount' in retail_price:
                    price = f"{retail_price['amount']} {retail_price.get('currencyCode', '€')}"
                elif list_price and 'amount' in list_price:
                    price = f"{list_price['amount']} {list_price.get('currencyCode', '€')}"

                thumbnail = book_info.get('imageLinks', {}).get('thumbnail', '')
                if thumbnail:
                    thumbnail = thumbnail.replace('http:', 'https:')

                return jsonify({
                    "title": book_info.get('title', 'Unknown Title'),
                    "authors": ', '.join(book_info.get('authors', ['Unknown Author'])),
                    "publisher": book_info.get('publisher', 'Unknown Publisher'),
                    "publishedDate": book_info.get('publishedDate', 'Unknown Date'),
                    "description": book_info.get('description', 'No description available'),
                    "pageCount": book_info.get('pageCount', 'Unknown'),
                    "price": price,
                    "thumbnail": thumbnail,
                    "isbn": clean_isbn,
                    "source": "google-books"
                })

        # Fallback: OpenLibrary
        ol_response = requests.get(
            f"https://openlibrary.org/isbn/{clean_isbn}.json",
            timeout=10
        )
        if ol_response.status_code == 200:
            ol_data = ol_response.json()
            author_names = []
            for author_ref in ol_data.get('authors', []):
                key = author_ref.get('key')
                if not key:
                    continue
                try:
                    author_resp = requests.get(f"https://openlibrary.org{key}.json", timeout=8)
                    if author_resp.status_code == 200:
                        author_names.append(author_resp.json().get('name'))
                except Exception:
                    continue

            return jsonify({
                "title": ol_data.get('title', 'Unknown Title'),
                "authors": ', '.join([a for a in author_names if a]) if author_names else 'Unknown Author',
                "publisher": ', '.join(ol_data.get('publishers', [])) if ol_data.get('publishers') else 'Unknown Publisher',
                "publishedDate": ol_data.get('publish_date', 'Unknown Date'),
                "description": (ol_data.get('description', {}).get('value') if isinstance(ol_data.get('description'), dict) else ol_data.get('description')) or 'No description available',
                "pageCount": ol_data.get('number_of_pages', 'Unknown'),
                "price": None,
                "thumbnail": f"https://covers.openlibrary.org/b/isbn/{clean_isbn}-L.jpg",
                "isbn": clean_isbn,
                "source": "openlibrary"
            })

        return jsonify({"error": f"Kein Buch zu dieser ISBN gefunden: {clean_isbn}"}), 404
        
    except Exception as e:
        print(f"Error fetching book data: {e}")
        return jsonify({"error": f"Failed to fetch book information: {str(e)}"}), 500

@app.route('/download_book_cover', methods=['POST'])
def download_book_cover():
    """
    API endpoint to download and save a book cover image from URL
    
    Returns:
        dict: Success status and filename or error message
    """
    if 'username' not in session:
        return jsonify({"error": "Not authorized"}), 403
    if not us.check_admin(session['username']):
        return jsonify({"error": "Admin privileges required"}), 403
    if not cfg.MODULES.is_enabled('library'):
        return jsonify({"error": "Bibliotheks-Modul ist deaktiviert."}), 403
    
    try:
        data = request.get_json()
        image_url = data.get('url')
        
        if not image_url:
            return jsonify({"error": "No image URL provided"}), 400

        parsed_url = urlparse(image_url)
        if parsed_url.scheme != 'https' or not parsed_url.netloc:
            return jsonify({"error": "Only public HTTPS URLs are allowed"}), 400

        hostname = parsed_url.hostname or ''
        if not _is_public_host(hostname):
            return jsonify({"error": "Target host is not allowed"}), 400
        
        # Download the image
        response = requests.get(image_url, stream=True, timeout=10, allow_redirects=False)
        
        if response.status_code != 200:
            return jsonify({"error": f"Failed to download image: Status {response.status_code}"}), 400
        
        # Check content type to ensure it's an image of allowed format
        content_type = response.headers.get('content-type', '')
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
        
        if not any(allowed_type in content_type.lower() for allowed_type in allowed_types):
            return jsonify({
                "error": f"Nicht unterstütztes Bildformat: {content_type}. Erlaubte Formate: JPG, JPEG, PNG, GIF"
            }), 400

        content_length = response.headers.get('Content-Length')
        if content_length:
            try:
                if int(content_length) > 5 * 1024 * 1024:
                    return jsonify({"error": "Image is too large"}), 413
            except ValueError:
                pass
        
        # Generate a fully unique filename using UUID
        import uuid
        import time
        
        unique_id = str(uuid.uuid4())
        timestamp = time.strftime("%Y%m%d%H%M%S")
        
        # Use appropriate extension based on content type
        extension = '.jpg'  # default
        if 'image/png' in content_type.lower():
            extension = '.png'
        elif 'image/gif' in content_type.lower():
            extension = '.gif'
            
        filename = f"book_cover_{unique_id}_{timestamp}{extension}"
        
        # Save the image to uploads folder
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        with open(filepath, 'wb') as f:
            written = 0
            for chunk in response.iter_content(chunk_size=8192):
                written += len(chunk)
                if written > 5 * 1024 * 1024:
                    return jsonify({"error": "Image is too large"}), 413
                f.write(chunk)
        
        return jsonify({
            "success": True,
            "filename": filename,
            "message": "Image downloaded successfully"
        })
        
    except Exception as e:
        print(f"Error downloading book cover: {e}")
       
        return jsonify({"error": f"Failed to download image: {str(e)}"}), 500
"""
@app.route('/proxy_image')
def proxy_image():
    Proxy endpoint to fetch images from external sources,
    bypassing CORS restrictions
    
    Returns:
        flask.Response: The image data or an error response
    url = request.args.get('url')
    if not url:
        return jsonify({"error": "No URL provided"}), 400

    parsed_url = urlparse(url)
    if parsed_url.scheme != 'https' or not parsed_url.netloc:
        return jsonify({"error": "Only public HTTPS URLs are allowed"}), 400

    hostname = parsed_url.hostname or ''
    if not _is_public_host(hostname):
        return jsonify({"error": "Target host is not allowed"}), 400
    
    try:
        # Fetch the image from the external source
        response = requests.get(url, stream=True, timeout=5, allow_redirects=False)
        
        # Check if the request was successful
        if response.status_code != 200:
            return jsonify({"error": f"Failed to fetch image: Status {response.status_code}"}), response.status_code
        
        # Get the content type
        content_type = response.headers.get('Content-Type', 'image/jpeg')
        if not content_type.lower().startswith('image/'):
            return jsonify({"error": "Target URL did not return an image"}), 400

        content_length = response.headers.get('Content-Length')
        if content_length:
            try:
                if int(content_length) > 5 * 1024 * 1024:
                    return jsonify({"error": "Image is too large"}), 413
            except ValueError:
                pass

        payload = bytearray()
        for chunk in response.iter_content(chunk_size=8192):
            if not chunk:
                continue
            payload.extend(chunk)
            if len(payload) > 5 * 1024 * 1024:
                return jsonify({"error": "Image is too large"}), 413
        
        # Return the image data with appropriate headers
        return Response(
            response=bytes(payload),
            status=200,
            headers={
                'Content-Type': content_type
            }
        )
    except Exception as e:
        print(f"Error in proxy_image: {e}")
        return jsonify({"error": f"Error fetching image: {str(e)}"}), 500
"""


def get_period_times(booking_date, period_num):
    """
    Get the start and end times for a given period on a specific date
    
    Args:
        booking_date (datetime): The date for the booking
        period_num (int): The period number
    
    Returns:
        dict: {"start": datetime, "end": datetime} or None if invalid
    """
    try:
        # Convert period_num to string for lookup in SCHOOL_PERIODS
        period_str = str(period_num)
        
        if period_str not in SCHOOL_PERIODS:
            return None
            
        period_info = SCHOOL_PERIODS[period_str]
        
        # Extract start and end times from period info
        start_time_str = period_info.get('start')
        end_time_str = period_info.get('end')
        
        if not start_time_str or not end_time_str:
            return None
            
        # Parse hours and minutes
        start_hour, start_min = map(int, start_time_str.split(':'))
        end_hour, end_min = map(int, end_time_str.split(':'))
        
        # Create datetime objects for start and end times
        start_datetime = datetime.datetime.combine(
            booking_date.date(),
            datetime.time(start_hour, start_min)
        )
        
        end_datetime = datetime.datetime.combine(
            booking_date.date(),
            datetime.time(end_hour, end_min)
        )
        
        return {
            'start': start_datetime,
            'end': end_datetime
        }
    except Exception as e:
        print(f"Error getting period times: {e}")
        return None

"""---------------------------------------------------------Borrowing-----------------------------------------------------------------"""

@app.route('/my_borrowed_items')
def my_borrowed_items():
    """
    Zeigt alle vom aktuellen Benutzer ausgeliehenen und geplanten Objekte an.
    
    Returns:
        Response: Gerendertes Template mit den ausgeliehenen und geplanten Objekten des Benutzers
    """
    if 'username' not in session:
        flash('Bitte melden Sie sich an, um Ihre ausgeliehenen Objekte anzuzeigen', 'error')
        return redirect(url_for('login', next=request.path))
    
    username = session['username']
    client = MongoClient(MONGODB_HOST, MONGODB_PORT)
    db = client[MONGODB_DB]
    items_collection = db.items
    ausleihungen_collection = db.ausleihungen
    
    # Get current time for comparison
    current_time = datetime.datetime.now()
    
    # Check if user is admin
    user_is_admin = False
    if 'is_admin' in session:
        user_is_admin = session['is_admin']
    
    # Get items currently borrowed by the user (where Verfuegbar=false and User=username)
    borrowed_items = list(items_collection.find({'Verfuegbar': False, 'User': username}))
    
    # Get active and planned ausleihungen for the user
    active_ausleihungen = list(ausleihungen_collection.find({
        'User': username,
        'Status': 'active'
    }))
    
    planned_ausleihungen = list(ausleihungen_collection.find({
        'User': username,
        'Status': 'planned'
    }))
    
    # DEBUG: Log the number of planned appointments found
    app.logger.info(f"Found {len(planned_ausleihungen)} planned appointments for user {username}")
    for appt in planned_ausleihungen:
        app.logger.info(f"Planned appointment: ID={str(appt['_id'])}, Item={str(appt.get('Item'))}, Start={appt.get('Start')}")
    
    # Process items
    active_items = []
    planned_items = []
    processed_item_ids = set()  # Keep track of processed item IDs to avoid duplicates
    
    # First, process items that are directly marked as borrowed by the user
    for item in borrowed_items:
        # Convert ObjectId to string for template
        item['_id'] = str(item['_id'])
        active_items.append(item)
        processed_item_ids.add(item['_id'])
    
    # Process active appointments
    for appointment in active_ausleihungen:
        # Get the item ID from the appointment
        item_id = appointment.get('Item')
        
        if not item_id or str(item_id) in processed_item_ids:
            continue  # Skip if we already processed this item or no item ID
        
        # Get item details
        item_obj = items_collection.find_one({'_id': ObjectId(item_id)})
        
        if item_obj:
            # Convert ObjectId to string for template
            item_obj['_id'] = str(item_obj['_id'])
            
            # Add appointment data
            item_obj['AppointmentData'] = {
                'id': str(appointment['_id']),
                'start': appointment.get('Start'),
                'end': appointment.get('End'),
                'notes': appointment.get('Notes'),
                'period': appointment.get('Period'),
                'status': appointment.get('VerifiedStatus', appointment.get('Status')),
            }
            
            # Mark that this item is part of an active appointment
            item_obj['ActiveAppointment'] = True
            
            # Add to the list only if not already there
            if str(item_obj['_id']) not in processed_item_ids:
                active_items.append(item_obj)
                processed_item_ids.add(str(item_obj['_id']))
    
    # Process planned appointments
    for appointment in planned_ausleihungen:
        item_id = appointment.get('Item')
        
        if not item_id:
            continue
        
        item_obj = items_collection.find_one({'_id': ObjectId(item_id)})
        
        if item_obj:
            item_obj['_id'] = str(item_obj['_id'])
            
            # Add appointment data
            item_obj['AppointmentData'] = {
                'id': str(appointment['_id']),
                'start': appointment.get('Start'),
                'end': appointment.get('End'),
                'notes': appointment.get('Notes'),
                'period': appointment.get('Period'),
                'status': appointment.get('Status'),
            }
            
            planned_items.append(item_obj)
    
    client.close()
    
    # DEBUG: Log what we're passing to the template
    app.logger.info(f"Passing {len(active_items)} active items and {len(planned_items)} planned items to template")
    if planned_items:
        for i, item in enumerate(planned_items):
            app.logger.info(f"Planned item {i+1}: {item['Name']}, Appointment ID: {item['AppointmentData']['id']}")
    
    return render_template(
        'my_borrowed_items.html',
        items=active_items,
        planned_items=planned_items,
        user_is_admin=user_is_admin
    )


@app.route('/notifications')
def notifications_view():
    """Notification center for users and admins."""
    if 'username' not in session:
        flash('Bitte melden Sie sich an, um Benachrichtigungen zu sehen.', 'error')
        return redirect(url_for('login'))

    username = session['username']
    is_admin_user = False
    try:
        is_admin_user = us.check_admin(username)
    except Exception:
        is_admin_user = False

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        notifications = _get_notifications_for_user(db, username, is_admin=is_admin_user, limit=200)

        user_notifications = []
        admin_notifications = []
        for notif in notifications:
            is_read = username in (notif.get('ReadBy') or [])
            row = {
                'id': str(notif.get('_id')),
                'title': notif.get('Title', 'Benachrichtigung'),
                'message': notif.get('Message', ''),
                'severity': notif.get('Severity', 'info'),
                'type': notif.get('Type', ''),
                'created_at': notif.get('CreatedAt'),
                'is_read': is_read,
                'reference': notif.get('Reference', {}) or {},
            }
            if notif.get('Audience') == 'admin':
                admin_notifications.append(row)
            else:
                user_notifications.append(row)

        return render_template(
            'notifications.html',
            user_notifications=user_notifications,
            admin_notifications=admin_notifications,
            is_admin_user=is_admin_user,
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        )
    except Exception as exc:
        app.logger.error(f"Error loading notifications: {exc}")
        flash('Fehler beim Laden der Benachrichtigungen.', 'error')
        return redirect(url_for('home'))
    finally:
        if client:
            client.close()


@app.route('/notifications/mark_read/<notification_id>', methods=['POST'])
def mark_notification_read(notification_id):
    """Mark a single notification as read for the current user."""
    if 'username' not in session:
        return redirect(url_for('login'))

    username = session['username']
    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        result = db['notifications'].update_one(
            {'_id': ObjectId(notification_id)},
            {
                '$addToSet': {'ReadBy': username},
                '$set': {'UpdatedAt': datetime.datetime.now()}
            }
        )
        if result.modified_count > 0:
            _bump_notification_version(f'user:{username}')
    except Exception as exc:
        app.logger.warning(f"Could not mark notification as read {notification_id}: {exc}")
    finally:
        if client:
            client.close()

    return redirect(url_for('notifications_view'))


@app.route('/notifications/mark_all_read', methods=['POST'])
def mark_all_notifications_read():
    """Mark all visible notifications as read for the current user."""
    if 'username' not in session:
        return redirect(url_for('login'))

    username = session['username']
    is_admin_user = False
    try:
        is_admin_user = us.check_admin(username)
    except Exception:
        is_admin_user = False

    query = {
        '$or': [
            {'Audience': 'user', 'TargetUser': username},
        ]
    }
    if is_admin_user:
        query['$or'].append({'Audience': 'admin'})

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        result = db['notifications'].update_many(
            query,
            {
                '$addToSet': {'ReadBy': username},
                '$set': {'UpdatedAt': datetime.datetime.now()}
            }
        )
        if result.modified_count > 0:
            _bump_notification_version(f'user:{username}')
    except Exception as exc:
        app.logger.warning(f"Could not mark all notifications as read for {username}: {exc}")
    finally:
        if client:
            client.close()

    return redirect(url_for('notifications_view'))


@app.route('/notifications/unread_status', methods=['GET'])
def notifications_unread_status():
    """Return unread notification count and latest unread message metadata."""
    if 'username' not in session:
        return jsonify({'ok': False, 'error': 'not_authenticated'}), 401

    username = session['username']
    is_admin_user = False
    try:
        is_admin_user = us.check_admin(username)
    except Exception:
        is_admin_user = False

    cached_payload, version_tag = _get_cached_unread_status(username, is_admin=is_admin_user)
    if cached_payload is not None:
        cached_etag = _build_unread_status_etag(version_tag, cached_payload)
        return _build_cached_json_response(cached_payload, cached_etag)

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]

        visibility_query = {
            '$or': [
                {'Audience': 'user', 'TargetUser': username},
            ]
        }
        if is_admin_user:
            visibility_query['$or'].append({'Audience': 'admin'})

        unread_query = {
            '$and': [
                visibility_query,
                {'ReadBy': {'$ne': username}},
            ]
        }

        unread_count = db['notifications'].count_documents(unread_query)
        latest_unread = db['notifications'].find_one(
            unread_query,
            {
                'Title': 1,
                'Message': 1,
                'CreatedAt': 1,
                'Type': 1,
                'Severity': 1,
            },
            sort=[('CreatedAt', -1)]
        )

        latest_payload = None
        if latest_unread:
            latest_payload = {
                'title': latest_unread.get('Title', 'Benachrichtigung'),
                'message': latest_unread.get('Message', ''),
                'created_at': latest_unread.get('CreatedAt').isoformat() if isinstance(latest_unread.get('CreatedAt'), datetime.datetime) else '',
                'type': latest_unread.get('Type', ''),
                'severity': latest_unread.get('Severity', 'info'),
            }

        payload = {
            'ok': True,
            'unread_count': unread_count,
            'latest_unread': latest_payload,
        }

        _set_cached_unread_status(username, is_admin_user, version_tag, payload)
        etag_value = _build_unread_status_etag(version_tag, payload)
        return _build_cached_json_response(payload, etag_value)
    except Exception as exc:
        app.logger.warning(f"Could not fetch unread notification status for {username}: {exc}")
        return jsonify({'ok': False, 'error': 'status_fetch_failed'}), 500
    finally:
        if client:
            client.close()


@app.route('/admin/damaged_items')
def admin_damaged_items():
    """Dedicated admin management window for damaged items."""
    if 'username' not in session or not us.check_admin(session['username']):
        flash('Administratorrechte erforderlich.', 'error')
        return redirect(url_for('login'))

    client = None
    try:
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        items_col = db['items']
        ausleihungen_col = db['ausleihungen']

        items = list(items_col.find(
            {
                'Deleted': {'$ne': True},
                '$or': [
                    {'HasDamage': True},
                    {'Condition': 'destroyed'},
                    {'DamageReports.0': {'$exists': True}},
                ]
            },
            {
                'Name': 1,
                'Code_4': 1,
                'ItemType': 1,
                'Author': 1,
                'ISBN': 1,
                'Condition': 1,
                'DamageReports': 1,
                'DamageRepairs': 1,
                'Verfuegbar': 1,
                'User': 1,
                'LastUpdated': 1,
            }
        ).sort('LastUpdated', -1))

        damaged_rows = []
        for item_doc in items:
            item_id = str(item_doc.get('_id'))
            active_borrow = ausleihungen_col.find_one(
                {'Item': item_id, 'Status': {'$in': ['active', 'planned']}},
                {'_id': 1, 'User': 1, 'Status': 1, 'End': 1}
            )
            reports = item_doc.get('DamageReports', []) or []
            latest_report = reports[0] if reports else {}

            damaged_rows.append({
                'id': item_id,
                'name': item_doc.get('Name', ''),
                'code': item_doc.get('Code_4', ''),
                'item_type': item_doc.get('ItemType', ''),
                'author': item_doc.get('Author', ''),
                'isbn': item_doc.get('ISBN', ''),
                'condition': item_doc.get('Condition', ''),
                'available': bool(item_doc.get('Verfuegbar', False)),
                'borrow_user': item_doc.get('User', ''),
                'damage_count': len(reports),
                'damage_reports': reports,
                'latest_damage_description': latest_report.get('description', ''),
                'latest_damage_by': latest_report.get('reported_by', ''),
                'latest_damage_at': latest_report.get('reported_at'),
                'active_borrow': active_borrow,
                'last_updated': item_doc.get('LastUpdated'),
            })

        return render_template(
            'admin_damaged_items.html',
            damaged_items=damaged_rows,
            library_module_enabled=cfg.MODULES.is_enabled('library'),
            student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
        )
    except Exception as exc:
        app.logger.error(f"Error loading damaged-items admin view: {exc}")
        flash('Fehler beim Laden der Defekte-Items-Verwaltung.', 'error')
        return redirect(url_for('home_admin'))
    finally:
        if client:
            client.close()

@app.route('/favicon.ico')
def favicon():
    """
    Serve the favicon directly from the static directory.
    
    Returns:
        flask.Response: The favicon.ico file
    """
    return send_from_directory(app.static_folder, 'favicon.ico')

@app.route('/get_predefined_locations')
def get_predefined_locations_route():
    """
    API endpoint to get predefined locations.
    
    Returns:
        dict: Dictionary containing predefined location values
    """
    values = it.get_predefined_locations()
    return jsonify({'locations': values})

@app.route('/add_location_value', methods=['POST'])
def add_location_value():
    """
    Add a new predefined location value.
    
    Returns:
        flask.Response: Redirect to location management page
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'success': False, 'error': 'Not authorized'}), 403
    
    value = sanitize_form_value(request.form.get('value'))
    
    if not value:
        flash('Bitte geben Sie einen Wert ein', 'error')
        return redirect(url_for('manage_locations'))
    
    # Add the value to locations
    success = it.add_predefined_location(value)
    
    if success:
        flash(f'Ort "{value}" wurde zur Liste hinzugefügt', 'success')
    else:
        flash(f'Ort "{value}" existiert bereits', 'error')
    
    return redirect(url_for('manage_locations'))

@app.route('/remove_location_value/<string:value>', methods=['POST'])
def remove_location_value(value):
    """
    Remove a predefined location value.
    
    Args:
        value (str): Value to remove
        
    Returns:
        flask.Response: Redirect to location management page
    """
    if 'username' not in session or not us.check_admin(session['username']):
        return jsonify({'success': False, 'error': 'Not authorized'}), 403
    
    # Remove the value from locations
    success = it.remove_predefined_location(value)
    
    if success:
        flash(f'Ort "{value}" wurde aus der Liste entfernt', 'success')
    else:
        flash(f'Fehler beim Entfernen des Ortes "{value}"', 'error')
    
    return redirect(url_for('manage_locations'))

@app.route('/manage_locations')
def manage_locations():
    """
    Admin page to manage predefined location values.
    
    Returns:
        flask.Response: Rendered location management template or redirect
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
    
    # Get predefined location values
    location_values = it.get_predefined_locations()
    
    return render_template(
        'manage_locations.html',
        location_values=location_values,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards')
    )


@app.route('/admin/school-settings', methods=['GET', 'POST'])
def admin_school_settings():
    """Admin page for configuring school metadata used in exports."""
    if 'username' not in session:
        flash('Bitte melden Sie sich mit einem administrativen Konto an.', 'error')
        return redirect(url_for('login'))
    if not us.check_admin(session['username']):
        flash('Für diese Seite sind Administratorrechte erforderlich.', 'error')
        return redirect(url_for('login'))

    permissions = _get_current_user_permissions()
    if not _action_access_allowed(permissions, 'can_manage_settings'):
        flash('Sie haben keine Berechtigung, die Schulstammdaten zu ändern.', 'error')
        return redirect(url_for('home_admin'))

    current_school = _get_school_info_for_export()
    tenant_context = None
    tenant_id = None
    tenant_db = cfg.MONGODB_DB
    try:
        tenant_context = get_tenant_context()
    except Exception:
        tenant_context = None
    if tenant_context:
        tenant_id = tenant_context.tenant_id
        tenant_db = tenant_context.db_name or tenant_db

    if request.method == 'POST':
        school_info = {
            'name': sanitize_form_value(request.form.get('name')),
            'address': sanitize_form_value(request.form.get('address')),
            'postal_code': sanitize_form_value(request.form.get('postal_code')),
            'city': sanitize_form_value(request.form.get('city')),
            'school_number': sanitize_form_value(request.form.get('school_number')),
            'it_admin': sanitize_form_value(request.form.get('it_admin')),
            'logo_path': sanitize_form_value(request.form.get('logo_path')),
            'logo_thumb': sanitize_form_value(request.form.get('logo_thumb')),
        }

        uploaded_logo = request.files.get('logo_upload')
        if uploaded_logo and getattr(uploaded_logo, 'filename', ''):
            try:
                previous_logo_path = current_school.get('logo_path') if isinstance(current_school, dict) else ''
                previous_thumb_path = current_school.get('logo_thumb') if isinstance(current_school, dict) else ''
                saved = _save_school_logo_upload(uploaded_logo, tenant_id=tenant_id, tenant_db=tenant_db)
                if isinstance(saved, (list, tuple)):
                    saved_logo_filename, saved_thumb_filename = saved
                else:
                    saved_logo_filename, saved_thumb_filename = saved, None

                if saved_logo_filename:
                    school_info['logo_path'] = saved_logo_filename
                    school_info['logo_thumb'] = saved_thumb_filename or ''

                    # remove previous files if different
                    if previous_logo_path and previous_logo_path != saved_logo_filename:
                        previous_logo_file = os.path.join(app.config['UPLOAD_FOLDER'], previous_logo_path)
                        if os.path.exists(previous_logo_file):
                            try:
                                os.remove(previous_logo_file)
                            except Exception:
                                pass
                    if previous_thumb_path and previous_thumb_path != saved_thumb_filename:
                        previous_thumb_file = os.path.join(app.config['UPLOAD_FOLDER'], previous_thumb_path)
                        if os.path.exists(previous_thumb_file):
                            try:
                                os.remove(previous_thumb_file)
                            except Exception:
                                pass
            except Exception as exc:
                flash(f'Logo konnte nicht hochgeladen werden: {exc}', 'error')
                return redirect(url_for('admin_school_settings'))

        if not school_info.get('logo_path') and current_school.get('logo_path'):
            school_info['logo_path'] = current_school.get('logo_path', '')
        if not school_info.get('logo_thumb') and current_school.get('logo_thumb'):
            school_info['logo_thumb'] = current_school.get('logo_thumb', '')

        missing_fields = [
            label for label, key in [
                ('Schulname', 'name'),
                ('Adresse', 'address'),
                ('PLZ', 'postal_code'),
                ('Ort', 'city'),
                ('Schulnummer', 'school_number'),
            ] if not school_info.get(key)
        ]

        if missing_fields:
            flash(f'Bitte füllen Sie die Pflichtfelder aus: {", ".join(missing_fields)}.', 'error')
        else:
            try:
                updated_school = cfg.update_school_info(school_info)
                current_school = updated_school
                flash('Schulstammdaten wurden erfolgreich gespeichert.', 'success')
            except Exception as exc:
                app.logger.error(f'Could not update school settings: {exc}\n{traceback.format_exc()}')
                flash('Die Schulstammdaten konnten nicht gespeichert werden.', 'error')

    return render_template(
        'admin_school_settings.html',
        school_info=current_school,
        tenant_id=tenant_id,
        tenant_db=tenant_db,
        library_module_enabled=cfg.MODULES.is_enabled('library'),
        student_cards_module_enabled=cfg.MODULES.is_enabled('student_cards'),
    )

@app.route('/check_code_unique/<code>')
def check_code_unique(code):
    """
    API endpoint to check if a code is unique
    
    Args:
        code (str): Code to check
        exclude_id (str, optional): ID of item to exclude from check (for edit operations)
        
    Returns:
        dict: JSON response with is_unique boolean
    """
    exclude_id = request.args.get('exclude_id')
    is_unique = it.is_code_unique(code, exclude_id)
    
    return jsonify({
        'is_unique': is_unique,
        'code': code
    })

@app.route('/schedule_appointment', methods=['POST'])
def schedule_appointment():
    """
    Schedule an appointment for an item
    """
    if 'username' not in session:
        return jsonify({'error': 'Nicht angemeldet'}), 401
        
    try:
        # Extract form data
        item_id = request.form.get('item_id')
        specific_item_id = (request.form.get('specific_item_id') or '').strip()
        schedule_date = request.form.get('schedule_date')
        start_period = request.form.get('start_period')
        end_period = request.form.get('end_period')
        notes = request.form.get('notes', '')
        
        # Check for multi-day
        is_multi_day = request.form.get('is_multi_day') == 'on'
        schedule_end_date = request.form.get('schedule_end_date')
        
        # Validate inputs
        if not all([item_id, schedule_date, start_period, end_period]):
            return jsonify({'success': False, 'message': 'Pflichtfelder fehlen'}), 400
            
        # Parse the start date
        try:
            appointment_date_obj = datetime.datetime.strptime(schedule_date, '%Y-%m-%d')
            appointment_date = appointment_date_obj.date()  # Get date part only
        except ValueError:
            return jsonify({'success': False, 'message': 'Ungültiges Datumsformat'}), 400
            
        # Parse end date if multi-day
        appointment_end_date = appointment_date
        if is_multi_day and schedule_end_date:
            try:
                appointment_end_date_obj = datetime.datetime.strptime(schedule_end_date, '%Y-%m-%d')
                appointment_end_date = appointment_end_date_obj.date()
                
                if appointment_end_date < appointment_date:
                    return jsonify({'success': False, 'message': 'Enddatum kann nicht vor Startdatum liegen'}), 400
            except ValueError:
                return jsonify({'success': False, 'message': 'Ungültiges Enddatumsformat'}), 400
            
        # Validate periods
        try:
            start_period_num = int(start_period)
            end_period_num = int(end_period)
            
            # Only check period order if it's the same day
            if appointment_date == appointment_end_date and start_period_num > end_period_num:
                return jsonify({'success': False, 'message': 'Startperiode kann nicht nach Endperiode liegen'}), 400
                
            if not (1 <= start_period_num <= 10) or not (1 <= end_period_num <= 10):
                return jsonify({'success': False, 'message': 'Ungültige Periodennummern'}), 400
                
        except ValueError:
            return jsonify({'success': False, 'message': 'Ungültige Periodenwerte'}), 400
            
        # Check if item exists
        item = it.get_item(item_id)
        if not item:
            return jsonify({'success': False, 'message': 'Element nicht gefunden'}), 404

        # Grouped inventory mode: allow scheduling for a specific physical unit.
        try:
            client = MongoClient(MONGODB_HOST, MONGODB_PORT)
            db = client[MONGODB_DB]
            items_col = db['items']
            grouped_children = list(items_col.find({'ParentItemId': item_id, 'IsGroupedSubItem': True}))
            client.close()
        except Exception:
            grouped_children = []

        if grouped_children:
            grouped_units = [item] + [{**child, '_id': str(child.get('_id'))} for child in grouped_children]
            available_units = [unit for unit in grouped_units if unit.get('Verfuegbar', True)]

            chosen_unit = None
            if specific_item_id:
                chosen_unit = next((unit for unit in grouped_units if str(unit.get('_id')) == specific_item_id), None)
                if not chosen_unit:
                    return jsonify({'success': False, 'message': 'Der gewählte Unterartikel wurde nicht gefunden.'}), 400
            else:
                chosen_unit = available_units[0] if available_units else grouped_units[0]

            item_id = str(chosen_unit.get('_id'))
            item = chosen_unit
            
        # Check if item is reservable
        if not item.get('Reservierbar', True):
             return jsonify({'success': False, 'message': 'Dieses Item kann nicht reserviert werden.'}), 400
            
        # Calculate start and end times for the appointment
        # Make sure we're passing the correct datetime object
        period_start_datetime = datetime.datetime.combine(appointment_date, datetime.time())
        period_times_start = get_period_times(period_start_datetime, start_period_num)
        
        period_end_datetime = datetime.datetime.combine(appointment_end_date, datetime.time())
        period_times_end = get_period_times(period_end_datetime, end_period_num)
        
        # Check if we got valid period times
        if not period_times_start or not period_times_end:
            print(f"Invalid period times: start={period_times_start}, end={period_times_end}")
            return jsonify({'success': False, 'message': 'Ungültige Periodenzeiten'}), 400
            
        start_datetime = period_times_start['start']
        end_datetime = period_times_end['end']
        
        # Determine if we should use period-based booking or time-based (multi-day)
        # If it's multi-day (dates differ), we treat it as a continuous block (Period=None)
        # If it's single day, we use the period range logic
        
        booking_period = None
        booking_period_end = None
        
        if appointment_date == appointment_end_date:
            booking_period = start_period_num
            booking_period_end = end_period_num
        
        # Check for conflicts (use full period-range aware check)
        try:
            has_conflict = au.check_booking_period_range_conflict(
                item_id,
                start_datetime,
                end_datetime,
                period=booking_period,
                period_end=booking_period_end
            )
            if has_conflict:
                return jsonify({'success': False, 'message': 'Termin kollidiert mit bestehender Buchung'}), 409
        except Exception as e:
            print(f"Error checking for booking conflicts: {e}")
            return jsonify({'success': False, 'message': f'Fehler beim Prüfen der Verfügbarkeit: {str(e)}'}), 500
            
        # Check if the appointment should already be active
        now = datetime.datetime.now()
        initial_status = 'active' if start_datetime <= now else 'planned'
        
        # Create the appointment
        try:
            # Use add_ausleihung directly to set the correct initial status
            appointment_id = au.add_ausleihung(
                item_id=item_id,
                user=session['username'],
                start_date=start_datetime,
                end_date=end_datetime,
                notes=notes,
                status=initial_status,
                period=booking_period # Will be None for multi-day
            )
            
            # If it became active immediately, log it and send a notification
            if initial_status == 'active' and appointment_id:
                app.logger.info(f"Appointment {appointment_id} scheduled retroactively as active.")
                
                # Make the item unavailable since it is now actively borrowed
                try:
                    it.update_item_status(item_id, False, session['username'])
                except Exception as update_err:
                    app.logger.warning(f"Failed to update item status when retroactively activating: {update_err}")

                # We can also notify the user right away
                item_name = item.get('Name', 'Unbekannt')
                
                # Log audit event
                _append_audit_event_standalone(
                    event_type='ausleihung_started',
                    payload={
                        'borrow_id': str(appointment_id),
                        'item_id': item_id,
                        'item_name': item_name,
                        'user': session['username'],
                        'status_before': 'planned',
                        'status_after': 'active'
                    }
                )
                
                # Send notification
                try:
                    client_temp = MongoClient(MONGODB_HOST, MONGODB_PORT)
                    db_temp = client_temp[MONGODB_DB]
                    _create_notification(
                        db_temp,
                        audience='user',
                        notif_type='appointment_activated',
                        title='Reservierung ist jetzt aktiv',
                        message=f"Deine geplante Ausleihe für {item_name} startet jetzt.",
                        target_user=session['username'],
                        reference={
                            'appointment_id': str(appointment_id),
                            'item_id': str(item_id),
                            'event': 'activated',
                        },
                        unique_key=f"appointment:activated:{appointment_id}",
                        severity='info'
                    )
                    client_temp.close()
                except Exception as notif_err:
                    app.logger.error(f"Error sending immediate active notification: {notif_err}")

            if not appointment_id:
                return jsonify({'success': False, 'message': 'Termin konnte nicht erstellt werden'}), 500
        except Exception as e:
            print(f"Error creating booking: {e}")
            return jsonify({'success': False, 'message': f'Fehler beim Erstellen des Termins: {str(e)}'}), 500
        
        # If we got this far, we have a valid appointment_id
        try:
            # Update item with next scheduled appointment info
            # Convert date to datetime for MongoDB storage if needed
            if isinstance(appointment_date, datetime.date) and not isinstance(appointment_date, datetime.datetime):
                appointment_datetime = datetime.datetime.combine(appointment_date, datetime.time())
            else:
                appointment_datetime = appointment_date
                
            # Handle end date conversion
            if isinstance(appointment_end_date, datetime.date) and not isinstance(appointment_end_date, datetime.datetime):
                appointment_end_datetime = datetime.datetime.combine(appointment_end_date, datetime.time())
            else:
                appointment_end_datetime = appointment_end_date
                
            result = it.update_item_next_appointment(item_id, {
                'date': appointment_datetime,
                'end_date': appointment_end_datetime,
                'start_period': start_period_num,
                'end_period': end_period_num,
                'user': session['username'],
                'notes': notes,
                'appointment_id': str(appointment_id)
            })
            
            if result:
                return jsonify({'success': True, 'appointment_id': str(appointment_id)})
            else:
                print("Failed to update item with appointment info")
                return jsonify({'success': False, 'message': 'Element konnte nicht mit Termininformationen aktualisiert werden'}), 500
                
        except Exception as e:
            print(f"Error updating item with appointment info: {e}")
            return jsonify({'success': False, 'message': f'Fehler beim Aktualisieren des Elements: {str(e)}'}), 500
            
    except Exception as e:
        print(f"Error creating appointment: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Serverfehler aufgetreten: {str(e)}'}), 500

@app.route('/cancel_ausleihung/<id>', methods=['POST'])
def cancel_ausleihung_route(id):
    """
    Route for canceling a planned or active ausleihung.
    
    Args:
        id (str): ID of the ausleihung to cancel
        
    Returns:
        flask.Response: Redirect to My Ausleihungen page
    """
    if 'username' not in session:
        flash('Ihnen ist es nicht gestattet auf dieser Internetanwendung, die eben besuchte Adrrese zu nutzen, versuchen sie es erneut nach dem sie sich mit einem berechtigten Nutzer angemeldet haben!', 'error')
        return redirect(url_for('login'))
        
    username = session['username']
    
    try:
        print(f"Attempting to cancel ausleihung with ID: {id}")
        
        # Get the ausleihung record to check if it belongs to the current user
        ausleihung = au.get_ausleihung(id)
        
        if not ausleihung:
            print(f"Ausleihung not found with ID: {id}")
            flash('Ausleihung nicht gefunden', 'error')
            return redirect(url_for('my_borrowed_items'))
            
        # Log ausleihung details for debugging
        ausleihung_status = ausleihung.get('Status', 'unknown')
        ausleihung_user = ausleihung.get('User', 'unknown')
        print(f"Found ausleihung: ID={id}, User={ausleihung_user}, Status={ausleihung_status}")
            
        # Check if the ausleihung belongs to the current user
        if ausleihung_user != username and not us.check_admin(username):
            print(f"Authorization failure: {username} attempted to cancel ausleihung belonging to {ausleihung_user}")
            flash('Sie sind nicht berechtigt, diese Ausleihung zu stornieren', 'error')
            return redirect(url_for('my_borrowed_items'))
            
        # Cancel the ausleihung
        if au.cancel_ausleihung(id):
            print(f"Successfully canceled ausleihung with ID: {id}")
            flash('Ausleihung wurde erfolgreich storniert', 'success')
            
            # If the booking was already active, make the item available again
            item_id = str(ausleihung.get('Item')) if ausleihung.get('Item') is not None else None
            if ausleihung_status == 'active' and item_id:
                try:
                    it.update_item_status(item_id, True)
                    print(f"Restored availability of item {item_id} after active cancellation")
                except Exception as status_err:
                    print(f"Warning: could not restore availability of item {item_id}: {status_err}")

            _append_audit_event_standalone(
                event_type='ausleihung_cancelled',
                payload={
                    'borrow_id': id,
                    'item_id': str(ausleihung.get('Item') or ''),
                    'cancelled_by': username,
                    'owner_user': str(ausleihung_user or ''),
                    'status_before': str(ausleihung_status or ''),
                }
            )
            # Also clear NextAppointment on the related item if it matches this appointment
            try:
                item_id = str(ausleihung.get('Item')) if ausleihung.get('Item') is not None else None
                if item_id:
                    item_doc = it.get_item(item_id)
                    if item_doc:
                        next_appt = item_doc.get('NextAppointment', {})
                        if next_appt and str(next_appt.get('appointment_id')) == str(id):
                            cleared = it.clear_item_next_appointment(item_id)
                            print(f"Cleared NextAppointment for item {item_id}: {cleared}")
            except Exception as clear_err:
                print(f"Warning: could not clear NextAppointment for cancelled ausleihung {id}: {clear_err}")
        else:
            print(f"Failed to cancel ausleihung with ID: {id}")
            flash('Fehler beim Stornieren der Ausleihung', 'error')
            
    except Exception as e:
        print(f"Error canceling ausleihung: {e}")
        flash(f'Fehler: {str(e)}', 'error')
        
    return redirect(url_for('my_borrowed_items'))

@app.route('/reset_item/<id>', methods=['POST'])
def reset_item(id):
    """
    Route for completely resetting an item's borrowing status.
    This handles items that have inconsistent borrowing states.
    
    Args:
        id (str): ID of the item to reset
        
    Returns:
        JSON: Success status and details
    """
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    if not us.check_admin(session['username']):
        return jsonify({'success': False, 'error': 'Not authorized'}), 403
    
    try:
        # Import the ausleihung module
        import ausleihung as au
        
        result = au.reset_item_completely(id)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'details': result.get('details', {})
            })
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        print(f"Error in reset_item route: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Serverfehler: {str(e)}'
        }), 500

# New image and video optimization functions
def is_image_file(filename):
    """
    Check if a file is an image based on its extension.
    
    Args:
        filename (str): Name of the file to check
        
    Returns:
        bool: True if the file is an image, False otherwise
    """
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff', '.tif', '.svg'}
    extension = filename.lower()[filename.rfind('.'):]
    return extension in image_extensions


def is_video_file(filename):
    """
    Check if a file is a video based on its extension.
    
    Args:
        filename (str): Name of the file to check
        
    Returns:
        bool: True if the file is a video, False otherwise
    """
    video_extensions = {'.mp4', '.mov', '.avi', '.mkv', '.webm', '.flv', '.m4v', '.3gp'}
    extension = filename.lower()[filename.rfind('.'):]
    return extension in video_extensions


def normalize_image_orientation(img, log_prefix=""):
    """
    Normalize image orientation using EXIF metadata.

    Many phone images are stored as "rotated + EXIF orientation tag".
    Converting them to other formats (e.g. WebP) without applying the EXIF
    transform can make portrait images appear sideways.
    """
    try:
        return ImageOps.exif_transpose(img)
    except Exception as exif_err:
        if log_prefix:
            app.logger.warning(f"{log_prefix} Could not apply EXIF orientation: {str(exif_err)}")
        return img


def create_image_thumbnail(image_path, thumbnail_path, size, debug_prefix=""):
    """
    Create a thumbnail for an image file, always converting to WebP format.
    
    Args:
        image_path (str): Path to the original image
        thumbnail_path (str): Path where the thumbnail should be saved
        size (tuple): Thumbnail size as (width, height)
        debug_prefix (str, optional): Prefix for debug logs
        
    Returns:
        bool: True if thumbnail was created successfully, False otherwise
    """
    # Check if this is a PNG file
    is_png = image_path.lower().endswith('.png')
    log_prefix = debug_prefix if debug_prefix else (f"PNG DEBUG: [{os.path.basename(image_path)}]" if is_png else "")
    
    try:
        if is_png and log_prefix:
            app.logger.info(f"{log_prefix} Creating thumbnail from PNG: {image_path} -> {thumbnail_path}")
            
        try:
            with Image.open(image_path) as img:
                img = normalize_image_orientation(img, log_prefix)

                if is_png and log_prefix:
                    app.logger.info(f"{log_prefix} PNG opened successfully: Format={img.format}, Mode={img.mode}, Size={img.size}")
                
                # Create thumbnail with proper aspect ratio
                if is_png and log_prefix:
                    app.logger.info(f"{log_prefix} Resizing PNG to {size}")
                try:
                    img.thumbnail(size, Image.Resampling.LANCZOS)
                except Exception as resize_err:
                    if is_png and log_prefix:
                        app.logger.error(f"{log_prefix} Error during PNG resize: {str(resize_err)}")
                        app.logger.info(f"{log_prefix} Trying alternative resize method")
                    # Try alternative resize method
                    img = img.resize((min(img.width, size[0]), min(img.height, size[1])), Image.Resampling.BILINEAR)
                
                # Create a new image with the exact size (add padding if needed)
                # Use RGBA for transparency support in WebP
                thumb = Image.new('RGBA', size, (255, 255, 255, 0))
                
                # Calculate position to center the image
                x = (size[0] - img.size[0]) // 2
                y = (size[1] - img.size[1]) // 2
                
                # Convert image to RGBA if it's not already
                if img.mode != 'RGBA':
                    img = img.convert('RGBA')
                
                thumb.paste(img, (x, y), img)

                # Ensure the thumbnail path ends with .webp
                if not thumbnail_path.lower().endswith('.webp'):
                    thumbnail_path = os.path.splitext(thumbnail_path)[0] + '.webp'

                # Ensure target directory exists
                os.makedirs(os.path.dirname(thumbnail_path), exist_ok=True)

                # Save with optimization
                thumb.save(thumbnail_path, 'WEBP', quality=85, method=6)
                return True
        except Exception as img_err:
            # Special handling for corrupted PNGs
            if is_png and log_prefix:
                app.logger.error(f"{log_prefix} Error opening PNG with PIL: {str(img_err)}")
                
                # Try to fix the PNG if possible
                app.logger.info(f"{log_prefix} Attempting to fix corrupt PNG")
                try:
                    # Create a placeholder thumbnail since we can't process this PNG
                    thumb = Image.new('RGBA', size, (200, 200, 200, 255))
                    # Add text indicating error
                    from PIL import ImageDraw
                    draw = ImageDraw.Draw(thumb)
                    text = "PNG Error"
                    draw.text((size[0]//4, size[1]//2), text, fill=(0, 0, 0, 255))
                    # Continue with saving this placeholder
                    app.logger.info(f"{log_prefix} Created placeholder for corrupt PNG")
                except Exception as fix_err:
                    app.logger.error(f"{log_prefix} Failed to create PNG placeholder: {str(fix_err)}")
                    raise img_err  # Re-raise the original error if we couldn't create a placeholder
            else:
                # For non-PNG files, just propagate the error
                raise
            
            if not thumbnail_path.lower().endswith('.webp'):
                thumbnail_path = os.path.splitext(thumbnail_path)[0] + '.webp'
            os.makedirs(os.path.dirname(thumbnail_path), exist_ok=True)
            # Save with optimization
            thumb.save(thumbnail_path, 'WEBP', quality=85, method=6)
            return True
            
    except Exception as e:
        print(f"Error creating image thumbnail for {image_path}: {str(e)}")
        return False


def create_video_thumbnail(video_path, thumbnail_path, size):
    """
    Create a thumbnail for a video file using ffmpeg.
    
    Args:
        video_path (str): Path to the original video
        thumbnail_path (str): Path where the thumbnail should be saved
        size (tuple): Thumbnail size as (width, height)
        
    Returns:
        bool: True if thumbnail was created successfully, False otherwise
    """
    try:
        # Use ffmpeg to extract a frame from the video (at 1 second)
        cmd = [
            'ffmpeg', 
            '-i', video_path,
            '-ss', '00:00:01.000',  # Extract frame at 1 second
            '-vframes', '1',
            '-y',  # Overwrite output file
            thumbnail_path + '.temp.jpg'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0 and os.path.exists(thumbnail_path + '.temp.jpg'):
            # Now resize the extracted frame using PIL
            success = create_image_thumbnail(thumbnail_path + '.temp.jpg', thumbnail_path, size)
            
            # Clean up temporary file
            try:
                os.remove(thumbnail_path + '.temp.jpg')
            except:
                pass
                
            return success
        else:
            print(f"ffmpeg failed for {video_path}: {result.stderr}")
            return False
            
    except Exception as e:
        print(f"Error creating video thumbnail for {video_path}: {str(e)}")
        return False


def generate_optimized_versions(filename, max_original_width=500, target_size_kb=80, debug_prefix=""):
    """
    Generate optimized version of uploaded files.
    Convert all image files to WebP format.
    Also resizes and compresses the original image to save storage space.
    No separate thumbnail or preview files are generated.
    
    Args:
        filename (str): Name of the uploaded file
        max_original_width (int): Maximum width for the original image (default: 500px)
        target_size_kb (int): Target file size in kilobytes (default: 80KB)
        
    Returns:
        dict: Dictionary with paths to generated files
    """
    # Create a process ID for logging
    process_id = str(uuid.uuid4())[:6]
    log_prefix = f"[Optimize-{process_id}][{filename}]"
    app.logger.info(f"{log_prefix} Starting optimization")
    
    # Make sure all required directories exist
    for directory in [app.config['UPLOAD_FOLDER']]:
        os.makedirs(directory, exist_ok=True)
    
    original_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    # Fallback to production path if dev path missing
    if not os.path.exists(original_path):
        prod_upload = "/var/Inventarsystem/Web/uploads"
        alt_path = os.path.join(prod_upload, filename)
        if os.path.exists(alt_path):
            original_path = alt_path
    
    # Generate file paths
    name_part, ext = os.path.splitext(filename)
    ext = ext.lower()
    is_webp_ext = ext == '.webp'
    
    # If already a WebP, keep filename to avoid same-file writes
    converted_filename = filename if is_webp_ext else f"{name_part}.webp"
    converted_path = os.path.join(app.config['UPLOAD_FOLDER'], converted_filename)
    
    result = {
        'original': converted_filename,  # Use WebP name; if already WebP, this equals input
        'thumbnail': None,
        'preview': None,
        'is_image': False,
        'is_video': False,
        'success': False
    }
    
    # Check if the file actually exists
    if not os.path.exists(original_path):
        app.logger.error(f"{log_prefix} Original file not found: {original_path}")
        
        # Check if we need to use a placeholder
        placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.svg')
        if not os.path.exists(placeholder_path):
            placeholder_path = os.path.join(app.static_folder, 'img', 'no-image.png')
        # Also check production static dir
        if not os.path.exists(placeholder_path):
            prod_static = "/var/Inventarsystem/Web/static/img"
            fallback_svg = os.path.join(prod_static, 'no-image.svg')
            fallback_png = os.path.join(prod_static, 'no-image.png')
            if os.path.exists(fallback_svg):
                placeholder_path = fallback_svg
            elif os.path.exists(fallback_png):
                placeholder_path = fallback_png
            
        if os.path.exists(placeholder_path):
            app.logger.info(f"{log_prefix} Using placeholder image instead")
            try:
                # Copy placeholder to uploads folder with the original filename
                shutil.copy2(placeholder_path, original_path)
                result['original'] = filename
                result['is_placeholder'] = True
                result['success'] = True
                return result
            except Exception as e:
                app.logger.error(f"{log_prefix} Failed to use placeholder: {str(e)}")
                return result
        else:
            app.logger.error(f"{log_prefix} No placeholder found, cannot continue")
            return result
    
    # Check if it's an image or video file
    is_png = filename.lower().endswith('.png')
    
    if is_image_file(filename):
        result['is_image'] = True
        app.logger.info(f"{log_prefix} Processing as image file")
        
        # Special logging for PNG files
        if is_png:
            if debug_prefix:
                app.logger.info(f"{debug_prefix} Processing PNG in optimization function")
            else:
                app.logger.info(f"PNG DEBUG: {log_prefix} Processing PNG in optimization function")
    elif is_video_file(filename):
        result['is_video'] = True
        app.logger.info(f"{log_prefix} Processing as video file")
        # For videos, we might want to extract a frame as the "image" representation
        # But for now, we'll just leave it as is, assuming the video itself is the asset
        # If a thumbnail is needed for video, we might need to generate one "poster" image
        # But the requirement says "only the downsized Image", which implies for images.
        # For videos, we'll just return success.
        result['success'] = True
        return result
    else:
        app.logger.info(f"{log_prefix} Not an image or video file, skipping optimization")
        return result
    
    try:
        # Get file info before processing
        original_size = os.path.getsize(original_path)
        app.logger.info(f"{log_prefix} Original size: {original_size/1024:.1f}KB")
        
    # Try to open and process the image
        try:
            with Image.open(original_path) as img:
                img = normalize_image_orientation(img, log_prefix)

                # Special handling for PNG
                is_png = filename.lower().endswith('.png')
                if is_png:
                    debug_msg = debug_prefix if debug_prefix else f"PNG DEBUG: {log_prefix}"
                    app.logger.info(f"{debug_msg} Processing PNG image in optimization function")
                    app.logger.info(f"{debug_msg} PNG details - Format: {img.format}, Mode: {img.mode}")
                
                # Log original dimensions
                original_width, original_height = img.size
                app.logger.info(f"{log_prefix} Original dimensions: {original_width}x{original_height}")
                
                # Resize if needed
                resized = False
                if original_width > max_original_width:
                    try:
                        scaling_factor = max_original_width / original_width
                        new_width = max_original_width
                        new_height = int(original_height * scaling_factor)
                        # Resize with high quality resampling
                        img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                        app.logger.info(f"{log_prefix} Resized to {new_width}x{new_height}")
                        if is_png:
                            app.logger.info(f"{debug_msg} PNG resized to {new_width}x{new_height}")
                        resized = True
                    except Exception as e:
                        app.logger.error(f"{log_prefix} Resize failed: {str(e)}")
                        if is_png:
                            app.logger.error(f"{debug_msg} PNG resize failed: {str(e)}")
                            app.logger.error(f"{debug_msg} Error type: {type(e).__name__}")
                        # Continue without resizing
                
                # Save as WebP with compression to target file size
                try:
                    # Get optimal quality setting to reach target size
                    # For WebP we can use a fixed quality or adapt it
                    quality = 80 # Default quality for WebP
                    app.logger.info(f"{log_prefix} Using quality setting: {quality}")
                    
                    # Standard save for WebP
                    if not is_webp_ext:
                        # Only create a new WebP if source wasn't already WebP
                        img.save(converted_path, 'WEBP', quality=quality, method=6)
                        app.logger.info(f"{log_prefix} Saved optimized WebP: {converted_path}")
                        
                        # Remove the original non-WebP file after successful conversion
                        if os.path.exists(converted_path):
                            try:
                                os.remove(original_path)
                                app.logger.info(f"{log_prefix} Removed original file after conversion")
                            except Exception as e:
                                app.logger.warning(f"{log_prefix} Error removing original file: {str(e)}")
                    else:
                        # Already a WebP: don't overwrite original; we'll use it for thumbs
                        app.logger.info(f"{log_prefix} Original is already WebP; skip in-place re-save")
                        
                except Exception as save_err:
                    app.logger.error(f"{log_prefix} Failed to save optimized WebP: {str(save_err)}")
                    
                    # Try with default quality as fallback (only when not already WebP)
                    try:
                        if not is_webp_ext:
                            app.logger.info(f"{log_prefix} Attempting save with default quality")
                            img.save(converted_path, 'WEBP', quality=80, method=6)
                            app.logger.info(f"{log_prefix} Saved WebP with default quality")
                        else:
                            app.logger.info(f"{log_prefix} Skipping fallback save; original is WebP and won't be overwritten")
                    except Exception as default_save_err:
                        app.logger.error(f"{log_prefix} WebP fallback save also failed: {str(default_save_err)}")
                        
                        # If WebP conversion fails entirely and different path, copy original
                        if not is_webp_ext and os.path.abspath(original_path) != os.path.abspath(converted_path):
                            shutil.copy2(original_path, converted_path)
                            app.logger.warning(f"{log_prefix} Used original file without optimization")
                    
                    # Compare file sizes
                    if not is_webp_ext and os.path.exists(converted_path):
                        new_size = os.path.getsize(converted_path)
                        reduction = (1 - (new_size / original_size)) * 100 if original_size > 0 else 0
                        app.logger.info(f"{log_prefix} Size reduction: {original_size/1024:.1f}KB -> {new_size/1024:.1f}KB ({reduction:.1f}%)")
                    
                    # Remove the original non-WebP file if it was converted or resized
                    if not is_webp_ext and os.path.exists(converted_path) and (not filename.lower().endswith('.webp') or resized):
                        try:
                            os.remove(original_path)
                            app.logger.info(f"{log_prefix} Removed original file after conversion")
                        except Exception as e:
                            app.logger.warning(f"{log_prefix} Error removing original file: {str(e)}")
                    
                except Exception as e:
                    app.logger.error(f"{log_prefix} Compression error: {str(e)}")
                    # Use original file if optimization fails
                    if not os.path.exists(converted_path):
                        shutil.copy2(original_path, converted_path)
                        app.logger.warning(f"{log_prefix} Used original file as fallback")
                
                # Use the converted file for thumbnails if it exists
                # If we produced a converted WebP (non-WebP source), use it as the basis for thumbs
                if not is_webp_ext and os.path.exists(converted_path):
                    original_path = converted_path
        
        except Exception as e:
            app.logger.error(f"{log_prefix} Failed to process image: {str(e)}")
            traceback.print_exc()
            # Just copy the original file as is
            if not is_webp_ext and os.path.exists(original_path) and not os.path.exists(converted_path):
                try:
                    shutil.copy2(original_path, converted_path)
                    app.logger.warning(f"{log_prefix} Used original file after processing error")
                except Exception as copy_err:
                    app.logger.error(f"{log_prefix} Failed to copy original file: {str(copy_err)}")
        
        # Mark success if we have at least the original or converted file
        if os.path.exists(original_path) or os.path.exists(converted_path):
            result['success'] = True
            app.logger.info(f"{log_prefix} Optimization completed successfully")
            
            # Log verification of all created files
            for file_type, file_path in [
                ('Original', original_path),
                ('Converted', converted_path)
            ]:
                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path) / 1024.0  # KB
                    app.logger.info(f"{log_prefix} {file_type}: {os.path.basename(file_path)} ({file_size:.1f}KB)")
                else:
                    app.logger.warning(f"{log_prefix} {file_type} file missing: {os.path.basename(file_path)}")
            
        return result
        
    except Exception as e:
        app.logger.error(f"{log_prefix} Unhandled exception in optimization: {str(e)}")
        traceback.print_exc()
        
        # If anything went wrong but the original file exists, just use it
        if os.path.exists(original_path):
            try:
                # Copy original to all required outputs as last resort
                for target_path in [converted_path]:
                    if not os.path.exists(os.path.dirname(target_path)):
                        os.makedirs(os.path.dirname(target_path), exist_ok=True)
                    shutil.copy2(original_path, target_path)
                
                result['original'] = filename
                result['success'] = True
                result['recovery'] = True
                app.logger.warning(f"{log_prefix} Recovery completed: using original file for all outputs")
                return result
            except Exception as recovery_err:
                app.logger.error(f"{log_prefix} Recovery failed: {str(recovery_err)}")
        
        return result

def get_thumbnail_info(filename):
    """
    Get thumbnail and preview information for a file.
    Returns the main image URL for all requests as we only use one image version now.
    
    Args:
        filename (str): Original filename
        
    Returns:
        dict: Dictionary with thumbnail and preview information
    """
    if not filename:
        return {'has_thumbnail': False, 'has_preview': False}
    
    name_part, ext_part = os.path.splitext(filename)
    
    # Check if the file exists (either as WebP or original extension)
    # We prefer WebP if available
    webp_filename = f"{name_part}.webp"
    webp_path = os.path.join(app.config['UPLOAD_FOLDER'], webp_filename)
    
    original_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    final_filename = filename
    has_image = False
    
    if os.path.exists(webp_path):
        final_filename = webp_filename
        has_image = True
    elif os.path.exists(original_path):
        final_filename = filename
        has_image = True
    else:
        # Check production paths
        prod_upload = "/var/Inventarsystem/Web/uploads"
        if os.path.exists(os.path.join(prod_upload, webp_filename)):
            final_filename = webp_filename
            has_image = True
        elif os.path.exists(os.path.join(prod_upload, filename)):
            final_filename = filename
            has_image = True
            
    # Build URLs based on the filename
    # We use the same URL for thumbnail and preview since we only have one image
    image_url = f"/uploads/{final_filename}" if has_image else None
    
    # Backward compatibility: Check for legacy thumbnails/previews
    thumbnail_url = image_url
    has_thumbnail = has_image
    
    # Check for legacy thumbnail files if we have a main image or not
    legacy_thumb_names = [
        f"{name_part}_thumb.webp",
        f"{name_part}_thumb.jpg",
        f"{name_part}_thumb{ext_part}"
    ]
    
    for thumb_name in legacy_thumb_names:
        if os.path.exists(os.path.join(app.config['THUMBNAIL_FOLDER'], thumb_name)):
            thumbnail_url = f"/thumbnails/{thumb_name}"
            has_thumbnail = True
            break
        elif os.path.exists(os.path.join("/var/Inventarsystem/Web/thumbnails", thumb_name)):
            thumbnail_url = f"/thumbnails/{thumb_name}"
            has_thumbnail = True
            break

    # Check for legacy preview files
    preview_url = image_url
    has_preview = has_image
    
    legacy_preview_names = [
        f"{name_part}_preview.webp",
        f"{name_part}_preview.jpg",
        f"{name_part}_preview{ext_part}"
    ]
    
    for preview_name in legacy_preview_names:
        if os.path.exists(os.path.join(app.config['PREVIEW_FOLDER'], preview_name)):
            preview_url = f"/previews/{preview_name}"
            has_preview = True
            break
        elif os.path.exists(os.path.join("/var/Inventarsystem/Web/previews", preview_name)):
            preview_url = f"/previews/{preview_name}"
            has_preview = True
            break
    
    return {
        'has_thumbnail': has_thumbnail,
        'has_preview': has_preview,
        'thumbnail_url': thumbnail_url,
        'preview_url': preview_url,
        'original_ext': os.path.splitext(final_filename)[1].lower(),
        'is_image': is_image_file(final_filename),
        'is_video': is_video_file(final_filename)
    }

# Mobile device detection utilities
def is_mobile_device(request):
    """Determine if the request is coming from a mobile device"""
    user_agent = request.headers.get('User-Agent', '').lower()
    mobile_identifiers = ['iphone', 'ipad', 'android', 'mobile', 'tablet']
    return any(identifier in user_agent for identifier in mobile_identifiers)

def is_ios_device(request):
    """Determine if the request is coming from an iOS device"""
    user_agent = request.headers.get('User-Agent', '').lower()
    return 'iphone' in user_agent or 'ipad' in user_agent or 'ipod' in user_agent

def log_mobile_action(action, request, success=True, details=None):
    """Log mobile-specific actions for debugging"""
    device_info = request.headers.get('User-Agent', 'Unknown device')
    status = "SUCCESS" if success else "FAILED"
    message = f"MOBILE {action} {status} - Device: {device_info}"
    if details:
        message += f" - Details: {details}"
    
    if success:
        app.logger.info(message)
    else:
        app.logger.error(message)
        
# Add explicit static file routes to handle CSS serving issues
@app.route('/static/<path:filename>')
def serve_static(filename):
    """
    Explicitly serve static files to resolve 403 Forbidden errors.
    This ensures CSS and JS files are properly accessible.
    
    Args:
        filename (str): The static file path
        
    Returns:
        flask.Response: The requested static file
    """
    return send_from_directory(app.static_folder, filename)

@app.route('/static/css/<filename>')
def serve_css(filename):
    """
    Explicitly serve CSS files from the static/css directory.
    
    Args:
        filename (str): Name of the CSS file to serve
        
    Returns:
        flask.Response: The requested CSS file
    """
    css_folder = os.path.join(app.static_folder, 'css')
    return send_from_directory(css_folder, filename)

@app.route('/static/js/<filename>')
def serve_js(filename):
    """
    Explicitly serve JavaScript files from the static/js directory.
    
    Args:
        filename (str): Name of the JS file to serve
        
    Returns:
        flask.Response: The requested JS file
    """
    js_folder = os.path.join(app.static_folder, 'js')
    return send_from_directory(js_folder, filename)


def cleanup_old_optimized_images(max_age_days=30):
    """
    Clean up old optimized images to save disk space.
    Optimized images are re-created on demand, so old ones can be safely deleted.
    
    Args:
        max_age_days (int): Delete cached images older than this many days. Default 30.
        
    Returns:
        dict: Statistics about cleanup (deleted count, freed space in MB)
    """
    try:
        import time
        import shutil
        
        cache_dir = os.path.join(app.config['THUMBNAIL_FOLDER'], 'optimized_480p')
        if not os.path.exists(cache_dir):
            return {'deleted': 0, 'freed_mb': 0, 'error': None}
        
        current_time = time.time()
        max_age_seconds = max_age_days * 24 * 60 * 60
        deleted_count = 0
        freed_bytes = 0
        
        for filename in os.listdir(cache_dir):
            file_path = os.path.join(cache_dir, filename)
            if not os.path.isfile(file_path):
                continue
            
            file_age_seconds = current_time - os.path.getmtime(file_path)
            if file_age_seconds > max_age_seconds:
                try:
                    file_size = os.path.getsize(file_path)
                    os.remove(file_path)
                    deleted_count += 1
                    freed_bytes += file_size
                except Exception as e:
                    app.logger.warning(f"Failed to delete optimized image {filename}: {str(e)}")
        
        freed_mb = freed_bytes / (1024 * 1024)
        app.logger.info(f"Cleanup complete: Deleted {deleted_count} images, freed {freed_mb:.2f} MB")
        
        return {
            'deleted': deleted_count,
            'freed_mb': round(freed_mb, 2),
            'error': None
        }
    except Exception as e:
        app.logger.error(f"Error during optimized image cleanup: {str(e)}")
        return {'deleted': 0, 'freed_mb': 0, 'error': str(e)}


@app.route('/log_mobile_issue', methods=['POST'])
def log_mobile_issue():
    """
    Route for logging mobile-specific issues.
    Used for tracking and debugging mobile browser problems.
    
    Returns:
        flask.Response: JSON response with success status
    """
    try:
        # Get issue data from request
        issue_data = request.json
        
        # Add timestamp if not present
        if 'timestamp' not in issue_data:
            issue_data['timestamp'] = datetime.now().isoformat()
            
        # Format the log message
        log_message = f"MOBILE ISSUE: {issue_data.get('action', 'unknown')} - "
        log_message += f"Error: {issue_data.get('error', 'none')} - "
        log_message += f"Browser: {issue_data.get('browser', 'unknown')}"
        
        # Create a structured log entry
        log_entry = {
            'type': 'mobile_issue',
            'timestamp': issue_data.get('timestamp'),
            'data': issue_data
        }
        
        # Log to application log file
        app.logger.warning(log_message)
        
        # Store in database for analytics
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        logs_collection = db['system_logs']
        logs_collection.insert_one(log_entry)
        client.close()
        
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f"Error logging mobile issue: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

def delete_item_images(filenames):
    """
    Delete all images associated with an item.
    Only deletes the main image file as we no longer use separate thumbnails/previews.
    
    Args:
        filenames (list): List of image filenames to delete
        
    Returns:
        dict: Statistics of deleted files
    """
    stats = {
        'originals': 0,
        'thumbnails': 0,
        'previews': 0,
        'errors': 0
    }
    
    if not filenames:
        return stats
        
    for filename in filenames:
        if not filename:
            continue
            
        try:
            # Generate paths based on filename pattern
            name_part = os.path.splitext(filename)[0]
            
            # Potential original files (WebP or JPG or original extension)
            files_to_check = [
                (os.path.join(app.config['UPLOAD_FOLDER'], f"{name_part}.webp"), 'originals'),
                (os.path.join(app.config['UPLOAD_FOLDER'], f"{name_part}.jpg"), 'originals'),
                (os.path.join(app.config['UPLOAD_FOLDER'], filename), 'originals'),
                
                # Also try to clean up legacy thumbnails/previews if they exist
                (os.path.join(app.config['THUMBNAIL_FOLDER'], f"{name_part}_thumb.webp"), 'thumbnails'),
                (os.path.join(app.config['THUMBNAIL_FOLDER'], f"{name_part}_thumb.jpg"), 'thumbnails'),
                (os.path.join(app.config['PREVIEW_FOLDER'], f"{name_part}_preview.webp"), 'previews'),
                (os.path.join(app.config['PREVIEW_FOLDER'], f"{name_part}_preview.jpg"), 'previews')
            ]
            
            # Delete all found files
            for file_path, category in files_to_check:
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        stats[category] += 1
                    except Exception as del_err:
                        app.logger.error(f"Failed to delete {file_path}: {del_err}")
                        stats['errors'] += 1
                
        except Exception as e:
            app.logger.error(f"Error deleting image files for {filename}: {str(e)}")
            stats['errors'] += 1
    
    return stats

def get_optimal_image_quality(img, target_size_kb=80):
    """
    Find the optimal JPEG quality setting to achieve a target file size.
    Uses a binary search approach to efficiently find the best quality.
    
    Args:
        img (PIL.Image): The PIL Image object
        target_size_kb (int): Target file size in kilobytes
        
    Returns:
        int: Quality setting (1-95)
    """
    import io
    
    # Initialize search range
    min_quality = 30  # We don't want to go lower than this
    max_quality = 95  # No need to go higher than this
    best_quality = 80  # Default quality
    best_diff = float('inf')
    target_size_bytes = target_size_kb * 1024
    
    # Binary search for optimal quality
    for _ in range(5):  # 5 iterations is usually enough
        quality = (min_quality + max_quality) // 2
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=quality, optimize=True)
        size = buffer.tell()
        
        # Check how close we are to target size
        diff = abs(size - target_size_bytes)
        if diff < best_diff:
            best_diff = diff
            best_quality = quality
        
        # Adjust search range
        if size > target_size_bytes:
            max_quality = quality - 1
        else:
            min_quality = quality + 1
            
        # If we’re within 10% of target, that’s good enough
        if abs(size - target_size_bytes) < (target_size_bytes * 0.1):
            return quality
    
    return best_quality


# ============================================================================
# PUSH NOTIFICATION API ENDPOINTS
# ============================================================================

@app.route('/health')
def health_check():
    return 'OK', 200

@app.route('/api/push/subscribe', methods=['POST'])
def subscribe_to_push():
    """
    Subscribe a user to push notifications
    
    Expects JSON payload:
    {
        'subscription': {
            'endpoint': '...',
            'keys': {'p256dh': '...', 'auth': '...'}
        }
    }
    """
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    try:
        data = request.get_json() or {}
        subscription = data.get('subscription')
        
        if not subscription or not subscription.get('endpoint'):
            return jsonify({'success': False, 'error': 'Invalid subscription'}), 400
        
        username = session['username']
        success = pn.save_push_subscription(username, subscription)
        
        if success:
            app.logger.info(f'Push subscription saved for {username}')
            return jsonify({
                'success': True,
                'message': 'Successfully subscribed to push notifications'
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to save subscription'}), 500
            
    except Exception as e:
        app.logger.error(f'Error subscribing to push: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/push/unsubscribe', methods=['POST'])
def unsubscribe_from_push():
    """
    Unsubscribe a user from push notifications
    
    Expects JSON payload:
    {
        'endpoint': '...'
    }
    """
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    try:
        data = request.get_json() or {}
        endpoint = data.get('endpoint')
        
        if not endpoint:
            return jsonify({'success': False, 'error': 'Missing endpoint'}), 400
        
        username = session['username']
        success = pn.remove_push_subscription(username, endpoint)
        
        if success:
            app.logger.info(f'Push subscription removed for {username}')
            return jsonify({
                'success': True,
                'message': 'Successfully unsubscribed from push notifications'
            })
        else:
            return jsonify({'success': False, 'error': 'Subscription not found'}), 404
            
    except Exception as e:
        app.logger.error(f'Error unsubscribing from push: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/push/subscriptions', methods=['GET'])
def get_push_subscriptions():
    """
    Get all push subscriptions for the current user
    """
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    try:
        username = session['username']
        subscriptions = pn.get_user_subscriptions(username)
        
        # Convert ObjectIds to strings for JSON serialization
        subs_data = []
        for sub in subscriptions:
            subs_data.append({
                'id': str(sub['_id']),
                'endpoint': sub.get('Endpoint', ''),
                'created_at': sub.get('CreatedAt', '').isoformat() if sub.get('CreatedAt') else '',
                'last_used': sub.get('LastUsed', '').isoformat() if sub.get('LastUsed') else '',
                'user_agent': sub.get('UserAgent', ''),
            })
        
        return jsonify({
            'success': True,
            'subscriptions': subs_data,
            'count': len(subs_data)
        })
        
    except Exception as e:
        app.logger.error(f'Error getting push subscriptions: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/push/vapid-key', methods=['GET'])
def get_vapid_key():
    """
    Get the VAPID public key for push notifications
    Used by the service worker to communicate with push service
    """
    try:
        vapid_key = pn.VAPID_PUBLIC_KEY
        if not vapid_key:
            app.logger.warning('VAPID_PUBLIC_KEY not configured')
            return jsonify({
                'success': False,
                'error': 'Push notifications not configured on server'
            }), 501
        
        return jsonify({
            'success': True,
            'vapid_key': vapid_key
        })
        
    except Exception as e:
        app.logger.error(f'Error getting VAPID key: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/push/test', methods=['POST'])
def test_push_notification():
    """
    Send a test push notification (admin only)
    """
    if 'username' not in session:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    if not us.is_admin(session['username']):
        return jsonify({'success': False, 'error': 'Admin access required'}), 403
    
    try:
        data = request.get_json() or {}
        target_user = data.get('target_user', session['username'])
        
        sent = pn.send_push_notification(
            target_user,
            'Test Benachrichtigung',
            'Dies ist eine Test-Benachrichtigung von Inventarsystem',
            url='/',
            tag='test-notification'
        )
        
        if sent > 0:
            return jsonify({
                'success': True,
                'message': f'Test push sent to {sent} subscription(s)'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'No active subscriptions for user'
            }), 404
            
    except Exception as e:
        app.logger.error(f'Error sending test push: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500
